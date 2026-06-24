"""Seed rental data from EstateCFO_Rental_SampleData.xlsx into RDS.
Run from backend/: python scripts/seed_rental_from_excel.py [path_to_xlsx]
Default path: C:/Users/HCSUSER/Downloads/EstateCFO_Rental_SampleData.xlsx
Pass --clear to wipe existing rental data first.
"""
import sys
import os
from datetime import date, datetime
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

import models.tenancy  # noqa
import models.rentals.models  # noqa
import models.rentals.maintenance  # noqa
import models.rentals.vendor  # noqa

from database import Base, SessionLocal, engine
from models.rentals.models import (
    RentalCollection, RentalCompany, RentalExpense, RentalExpenseCategory,
    RentalInvoice, RentalLease, RentalLeaseStatus, RentalOwnership,
    RentalPartnerRole, RentalProp, RentalTenant, RentalUnit, RentalUnitStatus,
)
from models.rentals.maintenance import (
    MaintenanceCategory, MaintenancePriority, MaintenanceRequest, MaintenanceStatus,
)
from models.rentals.vendor import RentalVendor, VendorCategory
from models.tenancy import TenantUser
from services.local_auth import DEMO_EMAIL

import openpyxl

Base.metadata.create_all(bind=engine)

EXCEL_PATH = r"C:\Users\HCSUSER\Downloads\EstateCFO_Rental_SampleData.xlsx"


def parse_date(val):
    if not val:
        return None
    if isinstance(val, date) and not isinstance(val, datetime):
        return val
    if isinstance(val, datetime):
        return val.date()
    try:
        from dateutil.parser import parse as dparse
        return dparse(str(val)).date()
    except Exception:
        return None


def safe_float(val, default=0.0):
    if val is None:
        return default
    try:
        return float(val)
    except (ValueError, TypeError):
        return default


def sheet_rows(ws):
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return [], []
    headers = [str(h).strip().lower().replace(' ', '_') if h else '' for h in rows[0]]
    data = []
    for row in rows[1:]:
        if all(v is None for v in row):
            continue
        data.append(dict(zip(headers, row)))
    return headers, data


def get(row, *keys):
    for k in keys:
        v = row.get(k.lower().replace(' ', '_'))
        if v is not None:
            return v
    return None


def clear_rental_data(db, tid):
    from sqlalchemy import delete
    db.execute(delete(RentalCollection).where(RentalCollection.tenant_id == tid))
    db.execute(delete(RentalInvoice).where(RentalInvoice.tenant_id == tid))
    db.execute(delete(RentalLease).where(RentalLease.tenant_id == tid))
    db.execute(delete(RentalTenant).where(RentalTenant.tenant_id == tid))
    db.execute(delete(MaintenanceRequest).where(MaintenanceRequest.tenant_id == tid))
    db.execute(delete(RentalExpense).where(RentalExpense.tenant_id == tid))
    db.execute(delete(RentalOwnership).where(RentalOwnership.tenant_id == tid))
    db.execute(delete(RentalUnit).where(RentalUnit.tenant_id == tid))
    db.execute(delete(RentalProp).where(RentalProp.tenant_id == tid))
    db.execute(delete(RentalCompany).where(RentalCompany.tenant_id == tid))
    db.execute(delete(RentalVendor).where(RentalVendor.tenant_id == tid))
    db.commit()
    print("Cleared existing rental data.")


def seed(xlsx_path=None, force_clear=False):
    if xlsx_path is None:
        xlsx_path = EXCEL_PATH

    db = SessionLocal()
    try:
        user = db.query(TenantUser).filter(TenantUser.email == DEMO_EMAIL).first()
        if not user:
            print("ERROR: Demo user not found. Start the backend once to auto-create it.")
            sys.exit(1)
        tid = user.tenant_id
        print(f"Tenant: {user.email}  (tenant_id={tid})")

        existing = db.query(RentalCompany).filter(RentalCompany.tenant_id == tid).count()
        if existing and not force_clear:
            print(f"Rental data already exists ({existing} companies). Use --clear to reset.")
            return
        if existing and force_clear:
            clear_rental_data(db, tid)

        wb = openpyxl.load_workbook(xlsx_path, data_only=True)
        print(f"Opened: {xlsx_path}")
        print(f"Sheets: {wb.sheetnames}")

        # ── Companies & Properties ─────────────────────────────────────────
        company_map: dict[str, RentalCompany] = {}  # company_name -> model
        prop_map: dict[tuple, RentalProp] = {}       # (company_name, property_name) -> model

        co_sheet = wb['Companies'] if 'Companies' in wb.sheetnames else None
        if co_sheet:
            _, rows = sheet_rows(co_sheet)
            for row in rows:
                co_name = str(get(row, 'company_name', 'company') or '').strip()
                prop_name = str(get(row, 'property_name', 'property') or '').strip()
                city = str(get(row, 'city') or '')
                state = str(get(row, 'state') or '')
                zip_ = str(get(row, 'zip') or '')
                ptype = str(get(row, 'property_type') or 'residential').lower()
                if not co_name:
                    continue

                co = company_map.get(co_name)
                if not co:
                    co = RentalCompany(tenant_id=tid, company_name=co_name, created_by=DEMO_EMAIL)
                    db.add(co)
                    db.flush()
                    company_map[co_name] = co
                    print(f"  Company: {co_name}")

                key = (co_name, prop_name or co_name)
                if key not in prop_map:
                    address = ', '.join(filter(None, [city, state, zip_])) or None
                    prop = RentalProp(
                        tenant_id=tid, company_id=co.id,
                        property_name=prop_name or co_name,
                        address=address, property_type=ptype,
                    )
                    db.add(prop)
                    db.flush()
                    prop_map[key] = prop

        # ── Units ─────────────────────────────────────────────────────────
        unit_map: dict[str, RentalUnit] = {}  # unit_number -> model
        rt_map: dict[str, 'RentalTenant'] = {}  # unit_number -> RentalTenant (current)

        status_xlat = {
            'occupied': RentalUnitStatus.occupied,
            'vacant': RentalUnitStatus.vacant,
            'notice': RentalUnitStatus.notice,
            'reserved': RentalUnitStatus.reserved,
        }

        unit_sheet = wb['Units'] if 'Units' in wb.sheetnames else None
        if unit_sheet:
            _, rows = sheet_rows(unit_sheet)
            for row in rows:
                unit_num = str(get(row, 'unit_id', 'unit') or '').strip()
                co_name = str(get(row, 'company') or '').strip()
                prop_name = str(get(row, 'property_name', 'property') or '').strip()
                if not unit_num or not co_name:
                    continue

                co = company_map.get(co_name)
                if not co:
                    co = RentalCompany(tenant_id=tid, company_name=co_name, created_by=DEMO_EMAIL)
                    db.add(co)
                    db.flush()
                    company_map[co_name] = co

                key = (co_name, prop_name or co_name)
                prop = prop_map.get(key)
                if not prop:
                    city = str(get(row, 'city') or '')
                    state = str(get(row, 'state') or '')
                    zip_ = str(get(row, 'zip') or '')
                    ptype = str(get(row, 'property_type') or 'residential').lower()
                    address = ', '.join(filter(None, [city, state, zip_])) or None
                    prop = RentalProp(
                        tenant_id=tid, company_id=co.id,
                        property_name=prop_name or co_name,
                        address=address, property_type=ptype,
                    )
                    db.add(prop)
                    db.flush()
                    prop_map[key] = prop

                status_str = str(get(row, 'status') or 'vacant').strip().lower()
                unit_status = status_xlat.get(status_str, RentalUnitStatus.vacant)
                rent = safe_float(get(row, 'actual_rent', 'monthly_rent', 'market_rent'))

                unit = RentalUnit(
                    tenant_id=tid,
                    property_id=prop.id,
                    company_id=co.id,
                    unit_number=unit_num,
                    status=unit_status,
                    monthly_rent=rent,
                )
                db.add(unit)
                db.flush()
                unit_map[unit_num] = unit

                tenant_name = str(get(row, 'tenant_name') or '').strip()
                if tenant_name and unit_status in (RentalUnitStatus.occupied, RentalUnitStatus.notice):
                    rt = RentalTenant(
                        tenant_id=tid,
                        unit_id=unit.id,
                        tenant_name=tenant_name,
                        is_current=True,
                    )
                    db.add(rt)
                    db.flush()
                    rt_map[unit_num] = rt

        db.flush()
        print(f"  Created {len(unit_map)} units, {len(rt_map)} tenants")

        # ── Leases ────────────────────────────────────────────────────────
        lease_sheet = wb['Leases'] if 'Leases' in wb.sheetnames else None
        lease_count = 0
        if lease_sheet:
            _, rows = sheet_rows(lease_sheet)
            for row in rows:
                unit_num = str(get(row, 'unit_id', 'unit') or '').strip()
                unit = unit_map.get(unit_num)
                if not unit:
                    continue

                ls = parse_date(get(row, 'lease_start'))
                le = parse_date(get(row, 'lease_end'))
                if not ls or not le:
                    continue

                tenant_name = str(get(row, 'tenant_name') or '').strip()
                rt = rt_map.get(unit_num)
                if not rt and tenant_name:
                    rt = db.query(RentalTenant).filter(
                        RentalTenant.tenant_id == tid,
                        RentalTenant.unit_id == unit.id,
                        RentalTenant.tenant_name == tenant_name,
                    ).first()
                    if not rt:
                        email = str(get(row, 'email') or '').strip() or None
                        phone = str(get(row, 'phone') or '').strip() or None
                        rt = RentalTenant(
                            tenant_id=tid, unit_id=unit.id,
                            tenant_name=tenant_name,
                            tenant_email=email,
                            tenant_phone=phone,
                            is_current=True,
                        )
                        db.add(rt)
                        db.flush()
                        rt_map[unit_num] = rt

                status_raw = str(get(row, 'status') or 'active').lower().strip()
                if status_raw == 'active':
                    l_status = RentalLeaseStatus.active
                elif 'notice' in status_raw:
                    l_status = RentalLeaseStatus.notice_given
                elif status_raw == 'expired':
                    l_status = RentalLeaseStatus.expired
                else:
                    l_status = RentalLeaseStatus.active

                deposit = safe_float(get(row, 'security_deposit', 'deposit'))
                monthly = safe_float(get(row, 'monthly_rent', 'rent'))

                lease = RentalLease(
                    tenant_id=tid,
                    unit_id=unit.id,
                    r_tenant_id=rt.id if rt else None,
                    lease_start=ls,
                    lease_end=le,
                    deposit_amount=deposit,
                    status=l_status,
                )
                db.add(lease)
                db.flush()

                # Create paid invoice history: monthly invoices + collections
                today = date.today()
                cur = date(ls.year, ls.month, 1)
                end_period = date(today.year, today.month, 1)
                while cur <= end_period and cur <= le:
                    inv = RentalInvoice(
                        tenant_id=tid,
                        unit_id=unit.id,
                        lease_id=lease.id,
                        billing_period=cur,
                        amount_billed=monthly or unit.monthly_rent,
                    )
                    db.add(inv)
                    db.flush()
                    col = RentalCollection(
                        tenant_id=tid,
                        invoice_id=inv.id,
                        amount_collected=monthly or unit.monthly_rent,
                        collected_date=date(cur.year, cur.month, min(5, 28)),
                    )
                    db.add(col)
                    # advance month
                    m = cur.month + 1
                    y = cur.year + (m - 1) // 12
                    m = ((m - 1) % 12) + 1
                    cur = date(y, m, 1)
                lease_count += 1

        db.flush()
        print(f"  Created {lease_count} leases with invoice history")

        # ── AR Dashboard — create unpaid invoices ─────────────────────────
        ar_sheet = wb['AR Dashboard'] if 'AR Dashboard' in wb.sheetnames else None
        ar_count = 0
        if ar_sheet:
            _, rows = sheet_rows(ar_sheet)
            for row in rows:
                unit_num = str(get(row, 'unit_id', 'unit') or '').strip()
                unit = unit_map.get(unit_num)
                if not unit:
                    continue
                ar_amount = safe_float(get(row, 'ar_amount'))
                if ar_amount <= 0:
                    continue

                # Find active lease for unit
                lease = db.query(RentalLease).filter(
                    RentalLease.tenant_id == tid,
                    RentalLease.unit_id == unit.id,
                    RentalLease.status == RentalLeaseStatus.active,
                ).first()
                if not lease:
                    lease = db.query(RentalLease).filter(
                        RentalLease.tenant_id == tid,
                        RentalLease.unit_id == unit.id,
                    ).first()

                as_of = parse_date(get(row, 'as_of_date')) or date.today()
                bp = date(as_of.year, as_of.month, 1)

                # Check if invoice for this period already exists (from lease seeding)
                existing_inv = db.query(RentalInvoice).filter(
                    RentalInvoice.tenant_id == tid,
                    RentalInvoice.unit_id == unit.id,
                    RentalInvoice.billing_period == bp,
                ).first()
                if existing_inv:
                    # Adjust: remove collection to make it appear as AR
                    col = db.query(RentalCollection).filter(
                        RentalCollection.tenant_id == tid,
                        RentalCollection.invoice_id == existing_inv.id,
                    ).first()
                    if col:
                        db.delete(col)
                    ar_count += 1
                elif lease:
                    inv = RentalInvoice(
                        tenant_id=tid,
                        unit_id=unit.id,
                        lease_id=lease.id,
                        billing_period=bp,
                        amount_billed=ar_amount,
                    )
                    db.add(inv)
                    ar_count += 1

        db.flush()
        print(f"  Created {ar_count} AR (unpaid invoice) records")

        # ── Expenses ──────────────────────────────────────────────────────
        exp_cat_map = {
            'repairs & maintenance': RentalExpenseCategory.repairs,
            'repairs': RentalExpenseCategory.repairs,
            'maintenance': RentalExpenseCategory.maintenance,
            'landscaping': RentalExpenseCategory.maintenance,
            'utilities': RentalExpenseCategory.utilities,
            'management': RentalExpenseCategory.management,
            'management fee': RentalExpenseCategory.management,
            'insurance': RentalExpenseCategory.insurance,
            'tax': RentalExpenseCategory.tax,
            'property tax': RentalExpenseCategory.tax,
            'cam': RentalExpenseCategory.cam,
        }

        exp_sheet = wb['Expenses'] if 'Expenses' in wb.sheetnames else None
        exp_count = 0
        if exp_sheet:
            _, rows = sheet_rows(exp_sheet)
            for row in rows:
                co_name = str(get(row, 'company') or '').strip()
                prop_name = str(get(row, 'property_name', 'property') or '').strip()
                co = company_map.get(co_name)
                if not co:
                    continue

                key = (co_name, prop_name or co_name)
                prop = prop_map.get(key) or next(
                    (p for (cn, pn), p in prop_map.items() if cn == co_name), None
                )
                if not prop:
                    continue

                exp_date = parse_date(get(row, 'date')) or date.today()
                cat_raw = str(get(row, 'category') or 'other').lower().strip()
                cat = exp_cat_map.get(cat_raw, RentalExpenseCategory.other)
                vendor = str(get(row, 'vendor') or '').strip()
                amt = safe_float(get(row, 'amount'))

                e = RentalExpense(
                    tenant_id=tid,
                    property_id=prop.id,
                    company_id=co.id,
                    expense_date=exp_date,
                    category=cat,
                    amount=amt,
                    description=f"{vendor} - {cat_raw}" if vendor else cat_raw,
                    created_by=DEMO_EMAIL,
                )
                db.add(e)
                exp_count += 1

        db.flush()
        print(f"  Created {exp_count} expense records")

        # ── Maintenance ───────────────────────────────────────────────────
        maint_cat_map = {
            'plumbing': MaintenanceCategory.plumbing,
            'electrical': MaintenanceCategory.electrical,
            'hvac': MaintenanceCategory.hvac,
            'ac': MaintenanceCategory.hvac,
            'appliance': MaintenanceCategory.appliance,
            'structural': MaintenanceCategory.structural,
            'pest': MaintenanceCategory.pest_control,
            'pest control': MaintenanceCategory.pest_control,
        }
        maint_priority_map = {
            'low': MaintenancePriority.low,
            'medium': MaintenancePriority.medium,
            'high': MaintenancePriority.high,
            'emergency': MaintenancePriority.emergency,
            'urgent': MaintenancePriority.high,
        }
        maint_status_map = {
            'open': MaintenanceStatus.open,
            'in progress': MaintenanceStatus.in_progress,
            'in_progress': MaintenanceStatus.in_progress,
            'completed': MaintenanceStatus.completed,
            'closed': MaintenanceStatus.closed,
            'assigned': MaintenanceStatus.assigned,
        }

        maint_sheet = wb['Maintenance'] if 'Maintenance' in wb.sheetnames else None
        maint_count = 0
        if maint_sheet:
            _, rows = sheet_rows(maint_sheet)
            for row in rows:
                co_name = str(get(row, 'company') or '').strip()
                prop_name = str(get(row, 'property') or '').strip()
                issue = str(get(row, 'issue', 'title', 'description') or '').strip()
                if not co_name or not issue:
                    continue

                co = company_map.get(co_name)
                key = (co_name, prop_name or co_name)
                prop = prop_map.get(key) or next(
                    (p for (cn, pn), p in prop_map.items() if cn == co_name), None
                )
                if not prop:
                    continue

                # Find first unit for this property
                unit = next(
                    (u for u in unit_map.values() if u.property_id == prop.id), None
                )
                if not unit:
                    # use any unit for this company
                    unit = next(
                        (u for u in unit_map.values() if u.company_id == (co.id if co else None)),
                        None,
                    )
                if not unit:
                    continue

                prio_raw = str(get(row, 'priority') or 'low').lower().strip()
                status_raw = str(get(row, 'status') or 'open').lower().strip()
                vendor_name = str(get(row, 'vendor') or '').strip() or None
                reported_date = parse_date(get(row, 'date_raised', 'date')) or date.today()
                cost = safe_float(get(row, 'estimated_cost', 'cost'), None)  # type: ignore

                # Guess category from issue text
                cat = MaintenanceCategory.general
                for kw, mapped_cat in maint_cat_map.items():
                    if kw in issue.lower():
                        cat = mapped_cat
                        break

                mr = MaintenanceRequest(
                    tenant_id=tid,
                    unit_id=unit.id,
                    property_id=prop.id,
                    title=issue[:512],
                    category=cat,
                    priority=maint_priority_map.get(prio_raw, MaintenancePriority.low),
                    status=maint_status_map.get(status_raw, MaintenanceStatus.open),
                    vendor_name=vendor_name,
                    cost=cost if cost else None,
                    reported_date=reported_date,
                    created_by=DEMO_EMAIL,
                )
                db.add(mr)
                maint_count += 1

        db.flush()
        print(f"  Created {maint_count} maintenance requests")

        # ── Vendors ───────────────────────────────────────────────────────
        vendor_cat_map = {
            'maintenance': VendorCategory.maintenance,
            'landscaping': VendorCategory.landscaping,
            'security': VendorCategory.security,
            'cleaning': VendorCategory.cleaning,
            'utilities': VendorCategory.utilities,
            'property management': VendorCategory.property_mgmt,
            'property_mgmt': VendorCategory.property_mgmt,
            'insurance': VendorCategory.insurance,
            'accounting': VendorCategory.accounting,
            'legal': VendorCategory.legal,
            'pool maintenance': VendorCategory.maintenance,
            'pool': VendorCategory.maintenance,
        }

        vendor_sheet = wb['Vendors'] if 'Vendors' in wb.sheetnames else None
        vendor_count = 0
        if vendor_sheet:
            _, rows = sheet_rows(vendor_sheet)
            for row in rows:
                vname = str(get(row, 'vendor_name', 'vendor') or '').strip()
                if not vname:
                    continue
                existing_v = db.query(RentalVendor).filter(
                    RentalVendor.tenant_id == tid,
                    RentalVendor.vendor_name == vname,
                ).first()
                if existing_v:
                    continue
                cat_raw = str(get(row, 'category') or 'other').lower().strip()
                phone = str(get(row, 'phone', 'contact_phone') or '').strip() or None
                email = str(get(row, 'email', 'contact_email') or '').strip() or None
                v = RentalVendor(
                    tenant_id=tid,
                    vendor_name=vname,
                    vendor_category=vendor_cat_map.get(cat_raw, VendorCategory.other),
                    contact_phone=phone,
                    contact_email=email,
                    created_by=DEMO_EMAIL,
                )
                db.add(v)
                vendor_count += 1

        db.commit()
        print(f"  Created {vendor_count} vendors")
        print("\nDone! Rental data seeded from Excel.")

    finally:
        db.close()


if __name__ == '__main__':
    force = '--clear' in sys.argv
    path = next((a for a in sys.argv[1:] if not a.startswith('--')), None)
    seed(path, force_clear=force)
