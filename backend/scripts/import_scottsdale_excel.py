"""
Import PR456 Scottsdale Promenade Center from Excel or JSON.

Canonical source: backend/data/PR456_full_data_package.json

Usage:
  cd backend
  python scripts/import_scottsdale_excel.py --tenant-id <uuid>
  python scripts/import_scottsdale_excel.py --tenant-id <uuid> --excel data/PR456_Scottsdale_Promenade_Center.xlsx
  python scripts/import_scottsdale_excel.py --transform-only
"""
from __future__ import annotations

import argparse
import re
import sys
import uuid
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from database import SessionLocal
from scripts.scottsdale_import import (
    EXCEL_DEFAULT_PATH,
    RAW_PACKAGE_PATH,
    final_state_check,
    persist_scottsdale,
    prepare_seed_from_raw,
    transform_raw_to_seed,
    verify_seed_data,
)

DATA_DIR = Path(__file__).resolve().parents[1] / "data"


def _blank(val) -> bool:
    if val is None:
        return True
    s = str(val).strip()
    return s in ("", "—", "-", "N/A", "n/a")


def _num(val) -> float | None:
    if _blank(val):
        return None
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).replace(",", "").replace("$", "").replace("%", "").strip()
    if not s:
        return None
    return float(s)


def _int_num(val) -> int | None:
    n = _num(val)
    return int(n) if n is not None else None


def _date_str(val) -> str | None:
    if _blank(val):
        return None
    if hasattr(val, "isoformat"):
        return val.date().isoformat() if hasattr(val, "date") else val.isoformat()[:10]
    return str(val).strip()[:10]


def _snake(s: str) -> str:
    return re.sub(r"[^a-z0-9]+", "_", s.lower()).strip("_")


def _parse_address(raw: str) -> dict:
    parts = [p.strip() for p in raw.split(",")]
    if len(parts) >= 3:
        city = parts[1]
        state_zip = parts[2].split()
        state = state_zip[0] if state_zip else "AZ"
        zip_code = state_zip[1] if len(state_zip) > 1 else None
        return {"address": parts[0], "city": city, "state": state, "zip_code": zip_code}
    return {"address": raw, "city": None, "state": None, "zip_code": None}


def _parse_schedule_range(raw: str) -> tuple[str | None, str | None]:
    if _blank(raw):
        return None, None
    if " to " in str(raw):
        a, b = str(raw).split(" to ", 1)
        return a.strip(), b.strip()
    return str(raw).strip(), None


PROJECT_TYPE_MAP = {
    "commercial (retail + office)": "commercial_for_sale",
    "commercial_for_sale": "commercial_for_sale",
    "mixed use": "mixed_use",
    "mixed_use": "mixed_use",
}

STATUS_MAP = {
    "under construction": "under_construction",
    "under_construction": "under_construction",
}

PERMIT_STATUS_MAP = {
    "under review": "under_review",
    "revisions requested": "revisions_requested",
    "not started": "not_started",
    "approval pending": "pending_approval",
    "signed": "approved",
}

CO_STATUS_MAP = {
    "approval pending": "approval_pending",
    "signed": "signed",
}


def parse_excel_to_raw(excel_path: Path) -> dict:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise ImportError("openpyxl is required for Excel import: pip install openpyxl") from exc

    wb = load_workbook(excel_path, data_only=True)
    summary = wb["Project Summary"]
    sov = wb["Cost Divisions (SOV)"]
    permits_ws = wb["Permits"]
    co_ws = wb["Change Orders"]
    sched_ws = wb["Schedule Health"]
    comp_ws = wb["Compliance Center"]
    fin_ws = wb["Financials & ROI"]

    addr_parts = _parse_address(str(summary["B8"].value or ""))
    sched = _parse_schedule_range(summary["B13"].value)
    sqft_raw = str(summary["B10"].value or "0")
    sqft = _int_num(re.sub(r"[^\d.]", "", sqft_raw))
    acres_raw = str(summary["B11"].value or "0")
    acres = _num(re.sub(r"[^\d.]", "", acres_raw))

    project_type_raw = str(summary["B6"].value or "").lower()
    status_raw = str(summary["B12"].value or "").lower()

    project_master = {
        "project_code": "PR456",
        "project_name": str(summary["B5"].value or "Scottsdale Promenade Center"),
        "project_type": PROJECT_TYPE_MAP.get(project_type_raw, "commercial_for_sale"),
        "address": addr_parts["address"],
        "city": addr_parts["city"],
        "state": addr_parts["state"],
        "zip_code": addr_parts["zip_code"],
        "county": str(summary["B9"].value or "Maricopa"),
        "total_saleable_sqft": sqft,
        "total_land_acres": acres,
        "status": STATUS_MAP.get(status_raw, "under_construction"),
        "start_date": _date_str(sched[0]),
        "target_completion_date": _date_str(sched[1]),
        "contract_value": _num(summary["B15"].value),
        "flood_zone": False,
        "wildfire_risk_zone": False,
        "hurricane_zone": False,
    }

    divisions = []
    for row in range(5, 22):
        div_num = sov.cell(row, 1).value
        if _blank(div_num) or str(div_num).upper() == "TOTAL":
            continue
        divisions.append({
            "division_number": str(div_num).zfill(2) if str(div_num).isdigit() else str(div_num),
            "division_name": str(sov.cell(row, 2).value or ""),
            "contract_value": _num(sov.cell(row, 3).value) or 0,
            "actual_cost_to_date": _num(sov.cell(row, 4).value) or 0,
            "committed_cost": _num(sov.cell(row, 5).value) or 0,
            "pct_complete": _num(sov.cell(row, 7).value) or 0,
        })

    permits = []
    for row in range(5, 14):
        ptype = permits_ws.cell(row, 1).value
        if _blank(ptype):
            continue
        status_raw_p = str(permits_ws.cell(row, 4).value or "").lower()
        blocking_raw = str(permits_ws.cell(row, 7).value or "").lower()
        permits.append({
            "permit_type": _snake(str(ptype)),
            "issuing_authority": str(permits_ws.cell(row, 2).value or ""),
            "budgeted_cost": _num(permits_ws.cell(row, 3).value),
            "actual_cost": None,
            "status": PERMIT_STATUS_MAP.get(status_raw_p, _snake(status_raw_p)),
            "is_blocking": blocking_raw in ("yes", "true", "1"),
            "application_date": None,
            "target_approval_date": _date_str(permits_ws.cell(row, 5).value),
            "actual_approval_date": _date_str(permits_ws.cell(row, 6).value),
            "notes": str(permits_ws.cell(row, 8).value or "") or None,
        })

    change_orders = []
    for row in range(5, 9):
        cr = co_ws.cell(row, 1).value
        if _blank(cr):
            continue
        sched_impact = co_ws.cell(row, 5).value
        days = _int_num(re.sub(r"[^\d]", "", str(sched_impact or "0")))
        status_co = str(co_ws.cell(row, 3).value or "").lower()
        change_orders.append({
            "change_request_number": str(cr),
            "description": str(co_ws.cell(row, 2).value or ""),
            "status": CO_STATUS_MAP.get(status_co, _snake(status_co)),
            "cost_impact": _num(co_ws.cell(row, 4).value) or 0,
            "schedule_impact_days": days,
            "due_date": _date_str(co_ws.cell(row, 6).value),
            "requested_by": str(co_ws.cell(row, 7).value or ""),
            "created_date": _date_str(co_ws.cell(row, 8).value),
        })

    schedule_tasks = []
    for row in range(5, 11):
        task = sched_ws.cell(row, 1).value
        if _blank(task):
            continue
        schedule_tasks.append({
            "task_name": str(task),
            "division_number": str(sched_ws.cell(row, 2).value or "").zfill(2),
            "planned_start": _date_str(sched_ws.cell(row, 3).value),
            "planned_end": _date_str(sched_ws.cell(row, 4).value),
            "actual_start": _date_str(sched_ws.cell(row, 5).value),
            "actual_end": _date_str(sched_ws.cell(row, 6).value),
            "days_late": _int_num(sched_ws.cell(row, 7).value) or 0,
        })

    compliance_docs = []
    for row in range(5, 16):
        vendor = comp_ws.cell(row, 1).value
        if _blank(vendor):
            continue
        compliance_docs.append({
            "vendor_name": str(vendor),
            "document_type": _snake(str(comp_ws.cell(row, 2).value or "")),
            "status": _snake(str(comp_ws.cell(row, 3).value or "missing")),
            "expiration_date": _date_str(comp_ws.cell(row, 4).value),
        })

    financials = {
        "period_start": "2026-01-01",
        "period_end": "2026-06-20",
        "received_from_owner": _num(fin_ws["B6"].value) or 0,
        "paid_to_subcontractors": abs(_num(fin_ws["B7"].value) or 0),
        "other_expenses": abs(_num(fin_ws["B8"].value) or 0),
        "retainage_held": _num(fin_ws["B11"].value) or 0,
        "retainage_receivable": _num(fin_ws["B12"].value) or 0,
    }

    roi_summary = {
        "total_project_cost": _num(fin_ws["B16"].value) or 62_732_000,
        "equity_pct": 0.40,
        "debt_pct": 0.60,
        "interest_rate_annual": _num(fin_ws["B19"].value) or 0.0825,
        "construction_months": _int_num(fin_ws["B20"].value) or 20,
        "stabilized_noi": _num(fin_ws["B22"].value) or 4_950_000,
        "exit_cap_rate": _num(fin_ws["B23"].value) or 0.0675,
        "selling_costs_pct": 0.025,
        "exit_strategy": "forward_sale",
    }

    return {
        "project_master": project_master,
        "divisions": divisions,
        "permits": permits,
        "change_orders": change_orders,
        "schedule_tasks": schedule_tasks,
        "financials": financials,
        "compliance_docs": compliance_docs,
        "roi_summary": roi_summary,
    }


def _print_final_summary(summary: dict) -> None:
    print("\n=== PR456 Import Final State ===")
    print(f"  Project: {summary['project_code']} @ {summary['address']}")
    print(f"  Scottsdale projects in tenant: {summary['scottsdale_project_count']} (expected 1)")
    print(f"  contract_value (SOV): ${summary['contract_value']:,.0f} (expected 41,160,000)")
    print(f"  roi_assumptions.total_project_cost: ${summary['total_project_cost']:,.0f} (expected 62,732,000)")
    print(f"  ROI verified: {summary['roi_verified']:.1%} | MOIC verified: {summary['moic_verified']:.2f}x")
    print(f"  SOV verified: ${summary['sov_verified']:,.0f}")
    print(f"  Divisions: {summary['divisions']} | Permits: {summary['permits']} | "
          f"COs: {summary['change_orders']} | Schedule: {summary['schedule_tasks']} | "
          f"Compliance: {summary['compliance_docs']}")
    print("  All verification assertions passed.")


def import_scottsdale(
    tenant_id: uuid.UUID,
    *,
    json_path: Path | None = None,
    excel_path: Path | None = None,
    replace: bool = True,
) -> uuid.UUID:
    if excel_path:
        raw = parse_excel_to_raw(excel_path)
        seed = transform_raw_to_seed(raw)
        from scripts.scottsdale_import import verify_seed_counts
        verify_seed_counts(seed)
        verify_seed_data(seed)
    else:
        seed = prepare_seed_from_raw(json_path or RAW_PACKAGE_PATH)

    db = SessionLocal()
    try:
        project_id = persist_scottsdale(db, tenant_id, seed, replace=replace)
        checks = verify_seed_data(seed)
        summary = final_state_check(db, tenant_id, project_id, seed, checks)
    except Exception:
        db.rollback()
        raise
    finally:
        db.close()

    master = seed["project_master"]
    print(f"Imported {master['project_name']} ({master['project_code']}) -> {project_id}")
    _print_final_summary(summary)
    return project_id


def main():
    parser = argparse.ArgumentParser(description="Import PR456 Scottsdale from Excel or JSON")
    parser.add_argument("--tenant-id", help="Tenant UUID (required unless --transform-only)")
    parser.add_argument("--json", type=Path, help="Path to PR456_full_data_package.json")
    parser.add_argument("--excel", type=Path, help="Path to PR456 Excel workbook")
    parser.add_argument(
        "--replace",
        action=argparse.BooleanOptionalAction,
        default=True,
        help="Replace existing PR456 / purge stale placeholder (default: true)",
    )
    parser.add_argument(
        "--transform-only",
        action="store_true",
        help="Transform + verify only (no DB)",
    )
    args = parser.parse_args()

    if args.transform_only:
        seed = prepare_seed_from_raw(args.json or RAW_PACKAGE_PATH)
        checks = verify_seed_data(seed)
        print("Transform + verification OK")
        print(f"  SOV: ${checks['sov_total']:,.0f} | ROI: {checks['roi']:.1%} | MOIC: {checks['moic']:.2f}x")
        return

    if not args.tenant_id:
        parser.error("--tenant-id is required unless --transform-only")

    import_scottsdale(
        uuid.UUID(args.tenant_id),
        json_path=args.json,
        excel_path=args.excel,
        replace=args.replace,
    )


if __name__ == "__main__":
    main()
