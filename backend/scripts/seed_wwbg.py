"""
Seed WWBG land-dev company from the 4 Excel files into RDS.
Run from backend/: python scripts/seed_wwbg.py

Files expected in C:/Users/HCSUSER/Downloads/WWBG LLC (1)/WWBG LLC/
  BS.xlsx, P&L.xlsx, LOANS.xlsx, cash Flows.xlsx

Adds ALTER TABLE for new columns before inserting.
"""
import sys
import os
from pathlib import Path
import datetime

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

import openpyxl
import sqlalchemy as sa

import models.tenancy            # noqa  register all models
import models.propdev.company    # noqa
import models.propdev.lot        # noqa
import models.propdev.partner    # noqa
import models.propdev.loan       # noqa
import models.propdev.capital_call  # noqa
import models.propdev.expense    # noqa

from database import Base, SessionLocal, engine
from models.propdev.company import PropDevCompany
from models.propdev.partner import PropDevPartner
from models.propdev.loan import PropDevLoan
from models.tenancy import TenantUser
from services.local_auth import DEMO_EMAIL

EXCEL_DIR = r"C:\Users\HCSUSER\Downloads\WWBG LLC (1)\WWBG LLC"

# ── 1. Ensure new columns exist ───────────────────────────────────────────────

def ensure_columns():
    ddls = [
        "ALTER TABLE propdev_companies ADD COLUMN IF NOT EXISTS interest_capitalised NUMERIC(16,2) DEFAULT 0 NOT NULL",
        "ALTER TABLE propdev_companies ADD COLUMN IF NOT EXISTS improvements NUMERIC(16,2) DEFAULT 0 NOT NULL",
        "ALTER TABLE propdev_companies ADD COLUMN IF NOT EXISTS yearly_pl JSONB",
        "ALTER TABLE propdev_companies ADD COLUMN IF NOT EXISTS yearly_bs JSONB",
        "ALTER TABLE propdev_companies ADD COLUMN IF NOT EXISTS yearly_cf JSONB",
    ]
    with engine.connect() as conn:
        for ddl in ddls:
            try:
                conn.execute(sa.text(ddl))
                conn.commit()
                print(f"  ✓ {ddl[:60]}...")
            except Exception as e:
                print(f"  (skipped) {e}")


# ── 2. Parse Excel files ──────────────────────────────────────────────────────

YEARS = ["2021", "2022", "2023", "2024", "2025", "2026"]

def _row_vals(row, max_cols=8):
    return [v for i, v in enumerate(row) if i < max_cols]

def _float(v):
    try:
        return float(v) if v is not None else 0.0
    except (TypeError, ValueError):
        return 0.0

def parse_bs(path: str) -> dict:
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    rows = {i+1: list(r) for i, r in enumerate(ws.iter_rows(values_only=True))}

    # Row 5: years in cols 2-7
    year_col = {}  # year_str -> col index (0-based)
    for j, v in enumerate(rows.get(5, [])):
        sv = str(v or '').strip()
        if sv in YEARS:
            year_col[sv] = j

    def find_row(label_kw):
        for rnum, row in rows.items():
            if row and str(row[0] or '').strip().lower().startswith(label_kw.lower()):
                return row
        return None

    def extract(label_kw):
        row = find_row(label_kw)
        if not row:
            return {y: 0.0 for y in YEARS}
        return {y: _float(row[c]) for y, c in year_col.items()}

    bank      = extract("Great Plains Bank")
    impr      = extract("Improvements")
    int_cap   = extract("Interest Capitalised")
    land      = extract("WWBL")
    total_ass = extract("Total for Assets")
    lt_loans  = extract("Total for Long-term")
    total_lia = extract("Total for Liabilities")

    return {
        "year_col": year_col,
        "bank": bank,
        "improvements": impr,
        "interest_capitalised": int_cap,
        "land": land,
        "total_assets": total_ass,
        "lt_loans": lt_loans,
        "total_liabilities": total_lia,
    }


def parse_pl(path: str) -> dict:
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    rows = {i+1: list(r) for i, r in enumerate(ws.iter_rows(values_only=True))}

    year_col = {}
    total_col = None
    for j, v in enumerate(rows.get(5, [])):
        sv = str(v or '').strip()
        if sv in YEARS:
            year_col[sv] = j
        elif sv.lower() == 'total':
            total_col = j

    def extract(label_kw):
        for rnum, row in rows.items():
            if row and str(row[0] or '').strip().lower().startswith(label_kw.lower()):
                return {y: _float(row[c]) for y, c in year_col.items()}
        return {y: 0.0 for y in YEARS}

    def extract_total(label_kw):
        for rnum, row in rows.items():
            if row and str(row[0] or '').strip().lower().startswith(label_kw.lower()):
                if total_col and total_col < len(row):
                    return _float(row[total_col])
        return 0.0

    net_income   = extract("Net Income")
    total_exp    = extract("Total for Expenses")
    interest_row = None
    mgmt_fee     = 0.0
    int_paid     = 0.0
    prop_tax     = 0.0
    prof_fee     = 0.0
    bk_charges   = 0.0
    engineering  = 0.0
    loan_proc    = 0.0
    appraisal    = 0.0
    land_survey  = 0.0

    for rnum, row in rows.items():
        if not row:
            continue
        label = str(row[0] or '').strip().lower()
        tc = _float(row[total_col]) if total_col and total_col < len(row) else 0.0
        if label == 'management fee':
            mgmt_fee = abs(tc)
        elif 'business loan interest' in label:
            int_paid = abs(tc)
        elif label.startswith('property tax'):
            prop_tax += abs(tc)
        elif 'professional service fee' in label:
            prof_fee += abs(tc)
        elif 'book keeping' in label or 'accounting fee' in label:
            bk_charges += abs(tc)
        elif 'engineering' in label:
            engineering += abs(tc)
        elif 'loan processing' in label:
            loan_proc = abs(tc)
        elif 'appraisal fee' in label:
            appraisal += abs(tc)
        elif 'land survey' in label:
            land_survey += abs(tc)

    return {
        "net_income": net_income,
        "total_expenses": total_exp,
        "totals": {
            "management_fee": mgmt_fee,
            "interest_paid": int_paid,
            "property_tax": prop_tax,
            "professional_charges": prof_fee + bk_charges,
            "engineering": engineering,
            "loan_processing": loan_proc,
            "appraisal": appraisal,
            "land_survey": land_survey,
        }
    }


def parse_loans(path: str) -> list:
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        return []

    headers = [str(h or '').strip() for h in rows[0]]
    loans = []
    for row in rows[1:]:
        if not row or not any(row):
            continue
        d = {headers[i]: row[i] for i in range(min(len(headers), len(row)))}
        company_name  = str(d.get('Company Name', '') or '').strip()
        property_name = str(d.get('Property Name', '') or '').strip()
        bank          = str(d.get('Loan Bank Name', '') or '').strip()
        acc_no        = str(d.get('Loan Acc No', '') or '').strip()
        amount        = _float(d.get('Loan Amount', 0))
        rate_raw      = d.get('Loan interest Rate', 0)
        rate          = _float(rate_raw) if _float(rate_raw) < 1 else _float(rate_raw) / 100
        emi           = _float(d.get('Loan EMI ', 0) or d.get('Loan EMI', 0))
        loan_date_raw = d.get('Loan Date')
        maturity_raw  = d.get('Loan Maurity Date')

        def to_date(v):
            if v is None:
                return None
            if isinstance(v, (datetime.date, datetime.datetime)):
                return v.date() if isinstance(v, datetime.datetime) else v
            try:
                return datetime.date.fromisoformat(str(v)[:10])
            except Exception:
                return None

        loans.append({
            'company_name': company_name,
            'property_name': property_name,
            'bank': bank,
            'account_no': acc_no,
            'loan_amount': amount,
            'interest_rate': rate,
            'emi': emi,
            'loan_date': to_date(loan_date_raw),
            'maturity_date': to_date(maturity_raw),
            'emi_day': 10,
        })
    return loans


def parse_cf(path: str) -> dict:
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    rows = {i+1: list(r) for i, r in enumerate(ws.iter_rows(values_only=True))}

    year_col = {}
    for j, v in enumerate(rows.get(5, [])):
        sv = str(v or '').strip()
        if sv in YEARS:
            year_col[sv] = j

    def extract(label_kw):
        for rnum, row in rows.items():
            if row and str(row[0] or '').strip().lower().startswith(label_kw.lower()):
                return {y: _float(row[c]) for y, c in year_col.items()}
        return {y: 0.0 for y in YEARS}

    operating  = extract("Net cash provided by operating")
    investing  = extract("Net cash provided by investing")
    financing  = extract("Net cash provided by financing")
    net_change = extract("NET CASH INCREASE")
    net_income = extract("Net Income")

    # Partner investments total per year (sum all partner rows)
    partner_inv = {y: 0.0 for y in YEARS}
    for rnum, row in rows.items():
        if not row:
            continue
        label = str(row[0] or '').strip().lower()
        if label.startswith('partner investments:'):
            for y, c in year_col.items():
                if c < len(row):
                    partner_inv[y] += _float(row[c])

    return {
        "operating": operating,
        "investing": investing,
        "financing": financing,
        "net_change": net_change,
        "net_income": net_income,
        "partner_investments": partner_inv,
    }


def parse_partners_from_bs(bs_path: str) -> list:
    """Extract partner names and total capital from BS equity section."""
    wb = openpyxl.load_workbook(bs_path, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))

    year_col = {}
    for j, v in enumerate(rows[4] if len(rows) > 4 else []):
        sv = str(v or '').strip()
        if sv in YEARS:
            year_col[sv] = j

    # Find "Equity" section; collect capital rows until end
    in_equity = False
    partners = []
    SKIP = {'equity', 'opening balance equity', 'partner investments', 'retained earnings',
            'total for equity', 'total for liabilities and equity', ''}
    for row in rows:
        if not row:
            continue
        label = str(row[0] or '').strip()
        ll = label.lower()
        if ll == 'equity':
            in_equity = True
            continue
        if in_equity:
            if ll.startswith('total') or ll in SKIP:
                if ll.startswith('total for liabilities'):
                    in_equity = False
                continue
            if ll in SKIP:
                continue
            # Get 2026 value (last year = most recent capital)
            c2026 = year_col.get('2026')
            c2025 = year_col.get('2025')
            capital = 0.0
            if c2026 and c2026 < len(row) and row[c2026] is not None:
                capital = _float(row[c2026])
            elif c2025 and c2025 < len(row) and row[c2025] is not None:
                capital = _float(row[c2025])
            if capital > 0 or label:
                # Strip " - Capital" suffix
                name = label.replace(' - Capital', '').replace(':RM - Capital', '').strip()
                if name and name.lower() not in SKIP:
                    partners.append({'name': name, 'capital': capital})
    return partners


# ── 3. Seed ───────────────────────────────────────────────────────────────────

def seed():
    print("=== WWBG Land Dev Seeder ===")

    print("\n1. Ensuring new columns exist on propdev_companies...")
    ensure_columns()

    print("\n2. Parsing Excel files...")
    base = EXCEL_DIR
    bs_data   = parse_bs(os.path.join(base, "BS.xlsx"))
    pl_data   = parse_pl(os.path.join(base, "P&L.xlsx"))
    loans     = parse_loans(os.path.join(base, "LOANS.xlsx"))
    cf_data   = parse_cf(os.path.join(base, "cash Flows.xlsx"))
    partners  = parse_partners_from_bs(os.path.join(base, "BS.xlsx"))

    print(f"   BS: {len(bs_data)} sections")
    print(f"   P&L: net income 2026 = {pl_data['net_income'].get('2026')}")
    print(f"   Loans: {len(loans)} records")
    print(f"   Partners from BS: {len(partners)}")

    db = SessionLocal()
    try:
        # Get tenant
        user = db.query(TenantUser).filter(TenantUser.email == DEMO_EMAIL).first()
        if not user:
            print(f"ERROR: Demo user '{DEMO_EMAIL}' not found. Run seed_rentals.py first.")
            return
        tid = user.tenant_id
        print(f"\n3. Tenant: {tid}")

        # ── Upsert WWBG company ──────────────────────────────────────────────
        company = db.query(PropDevCompany).filter(
            PropDevCompany.tenant_id == tid,
            PropDevCompany.name.ilike('%WWBG%'),
        ).first()

        totals = pl_data['totals']
        land_val = bs_data['land'].get('2026', 3338438.40)
        impr_val = bs_data['improvements'].get('2026', 389072.24)
        int_cap  = bs_data['interest_capitalised'].get('2026', 165225.36)
        cash_val = bs_data['bank'].get('2026', 18558.90)
        loan_bal = bs_data['lt_loans'].get('2026', 1787411.59)

        # Build yearly dicts for JSON storage
        yearly_bs = {}
        yearly_pl = {}
        yearly_cf = {}
        for y in YEARS:
            yearly_bs[y] = {
                "cash":                bs_data['bank'].get(y, 0),
                "land":                bs_data['land'].get(y, 0),
                "improvements":        bs_data['improvements'].get(y, 0),
                "interest_capitalised":bs_data['interest_capitalised'].get(y, 0),
                "total_assets":        bs_data['total_assets'].get(y, 0),
                "loan_balance":        bs_data['lt_loans'].get(y, 0),
                "total_liabilities":   bs_data['total_liabilities'].get(y, 0),
            }
            yearly_pl[y] = {
                "net_income":    pl_data['net_income'].get(y, 0),
                "total_expenses":pl_data['total_expenses'].get(y, 0),
                "revenue":       0.0,
            }
            yearly_cf[y] = {
                "operating":          cf_data['operating'].get(y, 0),
                "investing":          cf_data['investing'].get(y, 0),
                "financing":          cf_data['financing'].get(y, 0),
                "net_change":         cf_data['net_change'].get(y, 0),
                "partner_investments":cf_data['partner_investments'].get(y, 0),
            }

        fields = dict(
            property_name       = 'WWBL',
            address             = '',
            total_lots          = 1,
            sale_consideration  = 0.0,
            land_cost           = float(land_val),
            hard_cost           = float(impr_val),
            soft_cost           = 0.0,
            title_charges       = 0.0,
            other_charges       = float(totals['management_fee']),
            property_tax        = float(totals['property_tax']),
            loan_processing     = float(totals['loan_processing']),
            professional_charges= float(totals['professional_charges']),
            legal_fees          = 0.0,
            interest_on_loan    = float(totals['interest_paid']),
            cash_available      = float(cash_val),
            interest_capitalised= float(int_cap),
            improvements        = float(impr_val),
            yearly_pl           = yearly_pl,
            yearly_bs           = yearly_bs,
            yearly_cf           = yearly_cf,
        )

        if company:
            for k, v in fields.items():
                setattr(company, k, v)
            print(f"4. Updated existing company: {company.name} (id={company.id})")
        else:
            company = PropDevCompany(
                tenant_id=tid,
                name='WWBG',
                **fields,
            )
            db.add(company)
            db.flush()
            print(f"4. Created new company: WWBG (id={company.id})")

        db.flush()
        cid = company.id

        # ── Clear old loans/partners for this company ──────────────────────
        for ln in company.loans:
            db.delete(ln)
        for p in company.partners:
            db.delete(p)
        db.flush()

        # ── Add loan ───────────────────────────────────────────────────────
        print(f"\n5. Adding {len(loans)} loan(s)...")
        for ln in loans:
            db.add(PropDevLoan(
                tenant_id     = tid,
                company_id    = cid,
                bank          = 'Greater Plains Bank',
                loan_date     = ln['loan_date'],
                account_no    = ln['account_no'],
                loan_amount   = ln['loan_amount'],
                balance       = float(loan_bal),
                interest_rate = ln['interest_rate'],
                emi           = ln['emi'],
                maturity_date = ln['maturity_date'],
                emi_day       = 10,
                lender_name   = 'Greater Plains Bank',
                emi_status    = 'Current',
            ))
            print(f"   Loan: Greater Plains Bank ${ln['loan_amount']:,.0f} @ {ln['interest_rate']*100:.2f}%  EMI ${ln['emi']:,.2f}")

        # ── Add partners ───────────────────────────────────────────────────
        total_capital = sum(p['capital'] for p in partners if p['capital'] > 0) or 1.0
        print(f"\n6. Adding {len(partners)} partner(s) (total capital ${total_capital:,.2f})...")
        for p in partners:
            cap = p['capital']
            share = round(cap / total_capital * 100, 4) if cap > 0 else 0.0
            db.add(PropDevPartner(
                tenant_id            = tid,
                company_id           = cid,
                name                 = p['name'],
                partner_type         = 'Class A',
                share_percent        = share,
                capital_contributed  = float(cap),
                distributions_received = 0.0,
                preferred_return     = 8.0,
                status               = 'Active',
            ))
            print(f"   {p['name']}: ${cap:,.2f}  ({share:.2f}%)")

        db.commit()
        print(f"\n✅ Done — WWBG seeded with:")
        print(f"   Land cost:           ${land_val:>14,.2f}")
        print(f"   Improvements:        ${float(impr_val):>14,.2f}")
        print(f"   Interest capitalised:${float(int_cap):>14,.2f}")
        print(f"   Total invested:      ${land_val + impr_val + int_cap:>14,.2f}")
        print(f"   Outstanding loan:    ${float(loan_bal):>14,.2f}")
        print(f"   LTV:                 {float(loan_bal)/land_val*100:.1f}%")
        print(f"   Cash on hand:        ${float(cash_val):>14,.2f}")

    except Exception as e:
        db.rollback()
        import traceback
        traceback.print_exc()
        print(f"\n❌ Error: {e}")
    finally:
        db.close()


if __name__ == '__main__':
    seed()
