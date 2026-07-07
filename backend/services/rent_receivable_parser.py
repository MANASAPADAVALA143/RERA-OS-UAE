"""
Rent Receivable Parser — flexible header detection.

Supports two sheet layouts:
  Layout A (old):  col 0 = "Unit Name",          cols 1-12 = Jan-2026…Dec-2026
  Layout B (real): col 0 = "SUITE NAMES",  col 1 = "Name Of the Unit",
                   then alternating Mon-YYYY / Sec Dep pairs

Header detection scans the first 10 rows × first 20 columns looking for a
cell whose text matches UNIT_NAME_LABELS.  Whatever column that cell is in
becomes `unit_name_col`.  Month columns are found by matching MON-YYYY
patterns anywhere in the same header row — Sec Dep columns are ignored
automatically because they never match the month pattern.
"""

import re
import openpyxl
from datetime import datetime as _dt
from typing import Dict, List, Optional, Tuple

# All recognised header labels for the "unit name" column (case-insensitive)
UNIT_NAME_LABELS = {'unit name', 'name of the unit', 'unit'}

# Recognise any Mon-YYYY value in a header cell (supports any year)
_MONTH_RE = re.compile(
    r'^(Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)-(\d{4})$',
    re.IGNORECASE,
)

# Canonical 2026 order (used for vacancy-loss lookback sorting)
MONTHS = [
    'Jan-2026', 'Feb-2026', 'Mar-2026', 'Apr-2026',
    'May-2026', 'Jun-2026', 'Jul-2026', 'Aug-2026',
    'Sep-2026', 'Oct-2026', 'Nov-2026', 'Dec-2026',
]

_MONTH_ORDER_MAP = {m: i for i, m in enumerate(MONTHS)}


def _month_sort_key(m: str) -> Tuple[int, int]:
    """Sort key for any Mon-YYYY string, regardless of year."""
    mo = re.match(r'^(Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)-(\d{4})$', m, re.I)
    if not mo:
        return (9999, 99)
    month_names = ['jan','feb','mar','apr','may','jun','jul','aug','sep','oct','nov','dec']
    return (int(mo.group(2)), month_names.index(mo.group(1).lower()))


def count_physical_units(unit_name: str) -> int:
    """'Unit E,F,G' → 3;  'Unit R & S' → 2;  'Unit A' → 1"""
    name = str(unit_name)
    if ',' in name or '&' in name:
        parts = [p.strip() for p in name.replace('&', ',').split(',') if p.strip()]
        return max(1, len(parts))
    return 1


def expand_unit_match_names(raw_name: str) -> List[str]:
    """
    Names to try when matching an Excel row to registry units.
    'Unit K, L' → ['Unit K, L', 'Unit K', 'Unit L'] so split registry rows still sync.
    """
    name = ' '.join(str(raw_name).split()).strip()
    keys = [name]
    m = re.match(r'^(unit\s+)(.+)$', name, re.I)
    if not m:
        return keys
    prefix, rest = m.group(1), m.group(2)
    if ',' not in rest and '&' not in rest:
        return keys
    parts = [p.strip() for p in re.split(r'[,&]', rest) if p.strip()]
    if len(parts) <= 1:
        return keys
    for p in parts:
        keys.append(p if p.lower().startswith('unit') else f"{prefix}{p}")
    seen: set[str] = set()
    out: List[str] = []
    for k in keys:
        nk = k.lower()
        if nk not in seen:
            seen.add(nk)
            out.append(k)
    return out


def scale_amount_map(amounts: Dict[str, float], divisor: float) -> Dict[str, float]:
    """Split a combined-row rent total across multiple registry units."""
    if divisor <= 1:
        return amounts
    return {k: round(v / divisor, 2) for k, v in amounts.items()}


def safe_float(val) -> float:
    """Parse numeric rent cells — handles raw numbers, $1,234.56 strings, and blanks."""
    if val is None:
        return 0.0
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).strip()
    if not s or s in ('—', '–', '-'):
        return 0.0
    # Accounting negatives: ($1,234.56)
    neg = s.startswith('(') and s.endswith(')')
    if neg:
        s = s[1:-1]
    s = s.replace('$', '').replace(',', '').strip()
    if not s:
        return 0.0
    try:
        n = float(s)
        return -n if neg else n
    except (ValueError, TypeError):
        return 0.0


def _cell_to_month(val) -> str:
    """Convert a header cell to 'Mon-YYYY' if possible, else empty string.
    Handles both plain-text 'Jan-2026' and Excel date objects (datetime).
    """
    if val is None:
        return ''
    if isinstance(val, _dt):
        return val.strftime('%b-%Y')   # datetime(2026,1,1) → 'Jan-2026'
    s = ' '.join(str(val).split())     # normalise whitespace / NBSP
    mo = _MONTH_RE.match(s)
    if mo:
        return f"{mo.group(1).capitalize()}-{mo.group(2)}"
    return ''


def _norm(val) -> str:
    """Normalise a cell value to a plain lowercase string.
    Collapses all whitespace variants (including non-breaking spaces \\u00a0)
    so Excel-formatted cells compare correctly against UNIT_NAME_LABELS.
    """
    return ' '.join(str(val).split()).lower() if val is not None else ''


def _find_header(rows: list) -> Optional[Tuple[int, int]]:
    """
    Return (row_index, unit_name_col) of the first cell matching
    UNIT_NAME_LABELS within the first 20 rows × 30 columns.
    Returns None if not found.
    """
    for i, row in enumerate(rows[:20]):
        for j, cell in enumerate(row[:30]):
            if cell and _norm(cell) in UNIT_NAME_LABELS:
                return (i, j)
    return None


# Rows that are financial subtotals on some templates — not leasable units
_SKIP_UNIT_LABELS = frozenset({
    'rent',
    'sec dep',
    'sec-dep',
    'secdep',
    'security deposit',
    'security dep',
    'collected',
    'gross potential',
    'vacancy loss',
    'vacancy',
})

def _is_valid_unit_label(cell) -> bool:
    """Return True if a cell value looks like a unit/property name (not a header/summary)."""
    if not cell:
        return False
    name = _norm(cell)
    if not name or name == 'total' or name in UNIT_NAME_LABELS:
        return False
    if name in _SKIP_UNIT_LABELS:
        return False
    raw = str(cell)
    if '—' in raw or '–' in raw or 'Occupied' in raw or 'Collected' in raw:
        return False
    return True


def _row_has_month_data(row, month_col_map: Dict[str, int]) -> bool:
    """True when any month column has a numeric or non-empty rent value."""
    for col in month_col_map.values():
        if col >= len(row):
            continue
        val = row[col]
        if val is None:
            continue
        if isinstance(val, (int, float)):
            return True
        if str(val).strip():
            return True
    return False


def _extract_unit_name(row, unit_name_col: int) -> Optional[str]:
    """
    Read unit name from the header column; fall back to column A when B is empty.
    Some company sheets list the full property name only under SUITE NAMES.
    """
    if unit_name_col < len(row) and _is_valid_unit_label(row[unit_name_col]):
        return str(row[unit_name_col]).strip()
    if unit_name_col != 0 and 0 < len(row) and _is_valid_unit_label(row[0]):
        return str(row[0]).strip()
    return None


def _is_unit_row(row, unit_name_col: int) -> bool:
    """Return True if this row looks like a real unit data row."""
    return _extract_unit_name(row, unit_name_col) is not None


def _empty_result(co_name: str, error: str) -> Dict:
    return {
        'company': co_name, 'error': error,
        'units': [], 'monthly_totals': {}, 'collected': 0,
        'gross_potential': 0, 'total_physical_units': 0,
        'occupied_count': 0, 'vacant_count': 0,
        'occupancy_rate': 0, 'vacancy_loss': 0, 'vacant_units': [],
    }


def parse_sheet(ws, sheet_name: str, target_month: str) -> Dict:
    """Parse one company sheet.  target_month e.g. 'Jun-2026'."""
    rows = list(ws.iter_rows(values_only=True))
    co_name = sheet_name.strip()

    # ── 1. Locate header row and unit-name column ─────────────────────────────
    found = _find_header(rows)
    if found is None:
        return _empty_result(co_name, 'header not found (no "Unit Name" or "Name Of the Unit" cell)')

    hdr_row_idx, unit_name_col = found
    hdr = rows[hdr_row_idx]

    # ── 2. Map month label → column index (skip Sec Dep and other cols) ───────
    # Cells may be plain text "Jan-2026" OR Excel date objects → use _cell_to_month
    month_col_map: Dict[str, int] = {}
    for j, val in enumerate(hdr):
        if val:
            canonical = _cell_to_month(val)
            if canonical:
                month_col_map[canonical] = j

    if not month_col_map:
        return _empty_result(co_name, 'no month columns found')

    # ── 3. Resolve target column (fallback to latest available month) ─────────
    tgt = target_month
    if tgt not in month_col_map:
        available = sorted(month_col_map.keys(), key=_month_sort_key)
        tgt = available[-1] if available else None

    target_col = month_col_map.get(tgt) if tgt else None
    sorted_months = sorted(month_col_map.keys(), key=_month_sort_key)

    # ── 4. Parse unit rows — track suite name from col 0 ─────────────────────
    # Col 0 = "SUITE NAMES": may contain a suite label on its own row (suite
    # header) or alongside the unit name.  We inherit the last seen suite name
    # so even rows where col 0 is blank get the right suite assigned.
    SUITE_NAME_COL = 0
    units: List[Dict] = []
    current_suite: str = ''

    for row in rows[hdr_row_idx + 1:]:
        # Read col 0 as suite name candidate
        col0_raw = row[SUITE_NAME_COL] if SUITE_NAME_COL < len(row) else None
        col0_norm = _norm(col0_raw) if col0_raw else ''
        col0_display = str(col0_raw).strip() if col0_raw else ''
        # Update tracked suite when col 0 has a real value (not a header/total)
        if col0_norm and col0_norm not in ('suite names', 'total') \
                and col0_norm not in UNIT_NAME_LABELS:
            current_suite = col0_display

        unit_name = _extract_unit_name(row, unit_name_col)
        if not unit_name:
            continue

        # Skip suite-only header rows (col A == col B, no rent in month columns).
        # When both columns carry the same property name AND rent data exists,
        # treat as a valid single-property unit row.
        if (
            unit_name_col != SUITE_NAME_COL
            and col0_norm
            and _norm(unit_name) == col0_norm
            and not _row_has_month_data(row, month_col_map)
        ):
            continue

        # Suite for this unit: col 0 if populated (and not a header/total), else inherited
        suite_name = col0_display if (
            col0_norm and col0_norm not in ('suite names', 'total')
            and col0_norm not in UNIT_NAME_LABELS
        ) else current_suite

        current_amt = safe_float(row[target_col]) if (target_col is not None and target_col < len(row)) else 0.0

        history = {
            m: safe_float(row[col]) if col < len(row) else 0.0
            for m, col in month_col_map.items()
        }

        # Vacant if selected month's rent is $0; occupied if > $0
        is_vacant = current_amt == 0

        # Vacancy-loss: average last 2-3 non-zero months before target
        vacancy_loss = 0.0
        if is_vacant and tgt and tgt in sorted_months:
            target_idx = sorted_months.index(tgt)
            lookback: list[float] = []
            for prev_m in reversed(sorted_months[:target_idx]):
                prev_amt = history.get(prev_m, 0)
                if prev_amt > 0:
                    lookback.append(prev_amt)
                if len(lookback) >= 3:
                    break
            if lookback:
                vacancy_loss = round(sum(lookback) / len(lookback), 2)

        units.append({
            'name': unit_name,
            'suite': suite_name,          # ← suite from col A
            'physical_units': count_physical_units(unit_name),
            'current_amount': current_amt,
            'is_vacant': is_vacant,
            'vacancy_loss': vacancy_loss,
            'history': history,
        })

    # Fill vacancy_loss = 0 gaps with suite average
    occupied_units = [u for u in units if not u['is_vacant']]
    if occupied_units:
        suite_avg = sum(u['current_amount'] for u in occupied_units) / len(occupied_units)
        for u in units:
            if u['is_vacant'] and u['vacancy_loss'] == 0:
                u['vacancy_loss'] = round(suite_avg)

    total_physical    = sum(u['physical_units'] for u in units)
    occupied_physical = sum(u['physical_units'] for u in units if not u['is_vacant'])
    vacant_physical   = sum(u['physical_units'] for u in units if u['is_vacant'])
    collected         = sum(u['current_amount'] for u in units)
    vacancy_loss_total = sum(u['vacancy_loss'] for u in units)

    monthly_totals = {m: sum(u['history'].get(m, 0) for u in units) for m in sorted_months}
    gross_potential = max(monthly_totals.values()) if monthly_totals else collected

    return {
        'company': co_name,
        'units': units,
        'monthly_totals': monthly_totals,
        'target_month': tgt,
        'collected': collected,
        'gross_potential': gross_potential,
        'total_physical_units': total_physical,
        'occupied_count': occupied_physical,
        'vacant_count': vacant_physical,
        'occupancy_rate': round(occupied_physical / total_physical * 100, 1) if total_physical > 0 else 0,
        'vacancy_loss': vacancy_loss_total,
        'vacant_units': [u['name'] for u in units if u['is_vacant']],
    }


def parse_rent_receivable_file(file_path: str, target_month: str = 'Jun-2026') -> Dict:
    """
    Main entry point.
    target_month: 'Mon-YYYY' (e.g. 'Jun-2026').  If not found in the file the
    parser auto-falls-back to the latest available month.
    """
    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)

    results: Dict[str, Dict] = {}
    skipped: List[str] = []

    for sheet_name in wb.sheetnames:
        co_name = sheet_name.strip()
        if co_name.upper() in ('SUMMARY', 'INSTRUCTIONS', 'TEMPLATE'):
            continue

        ws = wb[sheet_name]
        data = parse_sheet(ws, sheet_name, target_month)

        if 'error' in data:
            skipped.append(f"{co_name}: {data['error']}")
        elif data['total_physical_units'] > 0:
            results[co_name] = data
        else:
            skipped.append(f"{co_name}: no units found")

    all_months = sorted(
        {m for d in results.values() for m in d['monthly_totals']},
        key=_month_sort_key,
    )

    total_units    = sum(d['total_physical_units'] for d in results.values())
    total_occupied = sum(d['occupied_count'] for d in results.values())
    total_vacant   = sum(d['vacant_count'] for d in results.values())
    total_collected = sum(d['collected'] for d in results.values())
    total_gross    = sum(d['gross_potential'] for d in results.values())
    total_vac_loss = sum(d['vacancy_loss'] for d in results.values())

    portfolio_monthly = {
        m: sum(d['monthly_totals'].get(m, 0) for d in results.values())
        for m in all_months
    }

    portfolio = {
        'target_month': target_month,
        'total_units': total_units,
        'occupied': total_occupied,
        'vacant': total_vacant,
        'total_collected': total_collected,
        'gross_potential': total_gross,
        'total_vacancy_loss': total_vac_loss,
        'collection_rate': round(total_collected / total_gross * 100, 1) if total_gross > 0 else 0,
        'occupancy_rate': round(total_occupied / total_units * 100, 1) if total_units > 0 else 0,
        'monthly_totals': portfolio_monthly,
        'companies_parsed': len(results),
        'skipped': skipped,
    }

    return {'companies': results, 'portfolio': portfolio}
