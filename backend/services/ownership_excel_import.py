"""Parse ownership/partner Excel uploads — portfolio-wide property-level positions."""
from __future__ import annotations

import re
from dataclasses import dataclass
from io import BytesIO
from typing import Any

from sqlalchemy import func, or_
from sqlalchemy.orm import Session

from models.rentals.models import RentalCompany, RentalOwnership, RentalPartnerRole, RentalProp

_COMPANY_HEADERS = frozenset({"entity name", "company name", "company"})
_ENTITY_LINE_HEADERS = frozenset({
    "entity", "business line", "line of business", "entity line", "business unit",
})

_COL_ALIASES: dict[str, tuple[str, ...]] = {
    "partner": ("owned by", "partner name", "partner", "owner"),
    "address": ("property address", "address"),
    "property": ("property name", "property", "building", "suite"),
    "ownership_pct": ("ownership %", "ownership", "own %", "own%", "ownership percent"),
    "structure": ("entity structure", "structure", "role", "partner type"),
    "cost_basis": ("cost basis", "cost"),
    "book_value": ("book value", "book"),
    "debt": ("existing debt", "debt", "loan balance", "mortgage"),
}

_SKIP_ROW = frozenset({
    "entity name", "company name", "owned by", "partner name", "total", "grand total", "entity",
})

_NON_RENTAL_ENTITY_TOKENS = (
    "construction", "development", "holding", "reit", "prop dev", "property dev",
)

_STRUCTURE_TO_ROLE: dict[str, str] = {
    "gp": "general_partner",
    "general partner": "general_partner",
    "general_partner": "general_partner",
    "lp": "limited_partner",
    "limited partner": "limited_partner",
    "limited_partner": "limited_partner",
    "llc": "limited_partner",
    "jv": "limited_partner",
    "sole": "sole_owner",
    "sole owner": "sole_owner",
    "sole_owner": "sole_owner",
    "100%": "sole_owner",
}


@dataclass
class ParsedOwnershipRow:
    entity_name: str
    partner_name: str
    property_address: str | None
    property_name: str
    ownership_pct: float
    entity_structure: str | None
    cost_basis: float | None
    book_value: float | None
    existing_debt: float | None
    entity_line: str | None
    sheet: str
    row_num: int


@dataclass
class ParseOwnershipResult:
    rows: list[ParsedOwnershipRow]
    skipped_non_rental: int
    has_entity_line_column: bool


def _norm_header(cell: Any) -> str:
    return re.sub(r"\s+", " ", str(cell or "").strip().lower())


def normalize_entity_line(val: Any) -> str | None:
    """Canonical Entity label: Rental, Land, Consulting, Partner, etc."""
    if val is None or not str(val).strip():
        return None
    key = _norm_header(val)
    if not key:
        return None
    if key in ("rental", "rentals", "property management", "real estate rental", "rental portfolio"):
        return "Rental"
    if key in ("land", "lands", "land holding", "land holdings"):
        return "Land"
    if "consult" in key:
        return "Consulting"
    if key in ("partner", "partners", "jv partner"):
        return "Partner"
    # Title-case leftover values for UI dropdown
    return str(val).strip().title()


def is_rental_entity_line(val: Any) -> bool:
    """True when Entity / business-line column indicates Rentals (or is blank)."""
    if val is None or not str(val).strip():
        return True
    key = _norm_header(val)
    if not key:
        return True
    if any(token in key for token in _NON_RENTAL_ENTITY_TOKENS):
        return False
    if "rental" in key:
        return True
    if key in ("rentals", "property management", "real estate rental", "rental portfolio"):
        return True
    return False


def is_ownership_entity_line(val: Any) -> bool:
    """True for Ownership import scope: Rental, Land, or blank Entity."""
    if val is None or not str(val).strip():
        return True
    normalized = normalize_entity_line(val)
    return normalized in ("Rental", "Land")


def _parse_num(v: Any) -> float | None:
    if v is None:
        return None
    s = str(v).strip().replace("$", "").replace(",", "")
    if not s or s in ("-", "—", "n/a", "na"):
        return None
    if s.endswith("%"):
        try:
            return float(s[:-1].strip())
        except ValueError:
            return None
    try:
        return float(s)
    except ValueError:
        return None


def _parse_pct(v: Any) -> float:
    raw = _parse_num(v)
    if raw is None:
        return 0.0
    return raw / 100 if raw > 1 else raw


def _map_headers(header_row: tuple) -> dict[str, int]:
    labels = [_norm_header(c) for c in header_row]
    has_entity_name_col = any(h in _COMPANY_HEADERS for h in labels if h)

    mapping: dict[str, int] = {}
    for idx, cell in enumerate(header_row):
        h = _norm_header(cell)
        if not h:
            continue
        if h in _COMPANY_HEADERS:
            mapping["company"] = idx
            continue
        if h in _ENTITY_LINE_HEADERS:
            # Legacy sheets: lone "Entity" column = company name (no Entity Name col)
            if h == "entity" and not has_entity_name_col:
                mapping["company"] = idx
            else:
                mapping["entity_line"] = idx
            continue
        for field, aliases in _COL_ALIASES.items():
            if field in mapping:
                continue
            if h in aliases or any(a in h for a in aliases if len(a) > 4):
                mapping[field] = idx
                break
    return mapping


def _find_header_row(rows: list[tuple]) -> tuple[int, dict[str, int]] | None:
    for i, row in enumerate(rows[:20]):
        if not row:
            continue
        labels = [_norm_header(c) for c in row if c is not None and str(c).strip()]
        if not labels:
            continue
        has_entity_name = any(x in _COMPANY_HEADERS for x in labels)
        has_legacy_entity = "entity" in labels and not has_entity_name
        has_company = has_entity_name or has_legacy_entity or any(x == "company" for x in labels)
        has_partner = any("owned by" in x or "partner" in x or "owner" in x for x in labels)
        has_property = any("property" in x for x in labels)
        if has_company and (has_partner or has_property):
            col_map = _map_headers(row)
            if "company" in col_map and "partner" in col_map:
                return i, col_map
    return None


def _cell(row: tuple, col_map: dict[str, int], field: str) -> Any:
    if field in col_map and col_map[field] < len(row):
        return row[col_map[field]]
    return None


def _parse_row_mapped(
    row: tuple,
    col_map: dict[str, int],
    sheet: str,
    row_num: int,
    *,
    filter_entity_line: bool,
    carried_book_value: float | None = None,
) -> ParsedOwnershipRow | None:
    company = str(_cell(row, col_map, "company") or "").strip()
    partner = str(_cell(row, col_map, "partner") or "").strip()
    entity_line_raw = _cell(row, col_map, "entity_line") if "entity_line" in col_map else None
    entity_line = normalize_entity_line(entity_line_raw)

    if filter_entity_line and not is_ownership_entity_line(entity_line_raw):
        return None
    if not company or not partner:
        return None
    if _norm_header(company) in _SKIP_ROW or _norm_header(partner) in _SKIP_ROW:
        return None
    prop = str(_cell(row, col_map, "property") or "").strip() or company
    book_value = _parse_num(_cell(row, col_map, "book_value"))
    # Excel often merges Book Value across Cost Basis rows — carry previous non-blank value.
    if book_value is None and carried_book_value is not None:
        book_value = carried_book_value
    return ParsedOwnershipRow(
        entity_name=company,
        partner_name=partner,
        property_address=str(_cell(row, col_map, "address") or "").strip() or None,
        property_name=prop,
        ownership_pct=_parse_pct(_cell(row, col_map, "ownership_pct")),
        entity_structure=str(_cell(row, col_map, "structure") or "").strip() or None,
        cost_basis=_parse_num(_cell(row, col_map, "cost_basis")),
        book_value=book_value,
        existing_debt=_parse_num(_cell(row, col_map, "debt")),
        entity_line=entity_line or ("Rental" if "entity_line" not in col_map else None),
        sheet=sheet,
        row_num=row_num,
    )


def parse_ownership_workbook(content: bytes, *, rental_only: bool = True) -> ParseOwnershipResult:
    """Parse ownership workbook.

    When rental_only=True (default), keep Entity = Rental or Land (plus blank).
    Other entity lines (Consulting, Partner, Construction, …) are skipped.
    """
    import openpyxl

    wb = openpyxl.load_workbook(BytesIO(content), data_only=True)
    out: list[ParsedOwnershipRow] = []
    skipped_non_rental = 0
    has_entity_line_column = False

    for ws in wb.worksheets:
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        header = _find_header_row(rows)
        if not header:
            continue
        hdr_idx, col_map = header
        if "entity_line" in col_map:
            has_entity_line_column = True
        carried_book_value: float | None = None
        for row_num, row in enumerate(rows[hdr_idx + 1:], start=hdr_idx + 2):
            if rental_only and "entity_line" in col_map:
                raw_line = _cell(row, col_map, "entity_line")
                if raw_line is not None and str(raw_line).strip() and not is_ownership_entity_line(raw_line):
                    skipped_non_rental += 1
                    continue
            raw_bv = _parse_num(_cell(row, col_map, "book_value")) if "book_value" in col_map else None
            if raw_bv is not None:
                carried_book_value = raw_bv
            parsed = _parse_row_mapped(
                row, col_map, ws.title, row_num,
                filter_entity_line=rental_only,
                carried_book_value=carried_book_value,
            )
            if parsed and parsed.ownership_pct > 0:
                out.append(parsed)

    return ParseOwnershipResult(
        rows=out,
        skipped_non_rental=skipped_non_rental,
        has_entity_line_column=has_entity_line_column,
    )


def _role_from_structure(structure: str | None) -> RentalPartnerRole:
    if not structure:
        return RentalPartnerRole.limited_partner
    key = structure.strip().lower().replace("_", " ")
    mapped = _STRUCTURE_TO_ROLE.get(key)
    if not mapped:
        for token, role in _STRUCTURE_TO_ROLE.items():
            if token in key:
                mapped = role
                break
    if mapped == "general_partner":
        return RentalPartnerRole.general_partner
    if mapped == "sole_owner":
        return RentalPartnerRole.sole_owner
    return RentalPartnerRole.limited_partner


def _match_company(db: Session, tid, name: str) -> RentalCompany | None:
    return db.query(RentalCompany).filter(
        RentalCompany.tenant_id == tid,
        or_(
            func.lower(func.trim(RentalCompany.company_name)) == name.lower().strip(),
            RentalCompany.company_name.ilike(f"%{name.strip()}%"),
        ),
    ).first()


def _match_property(db: Session, company_id, name: str) -> RentalProp | None:
    prop = name.strip()
    if not prop:
        return None
    return db.query(RentalProp).filter(
        RentalProp.company_id == company_id,
        or_(
            func.lower(func.trim(RentalProp.property_name)) == prop.lower(),
            RentalProp.property_name.ilike(f"%{prop}%"),
        ),
    ).first()


def import_ownership_from_excel(db: Session, tid, content: bytes) -> dict:
    parsed_result = parse_ownership_workbook(content, rental_only=True)
    parsed = parsed_result.rows
    if not parsed:
        hint = (
            "No ownership rows found for Entity = Rental or Land. "
            "Consulting/Partner/Construction rows are skipped."
            if parsed_result.has_entity_line_column
            else "No ownership rows found."
        )
        return {
            "imported_count": 0,
            "skipped_non_rental": parsed_result.skipped_non_rental,
            "errors": [
                f"{hint} Expected columns: Entity (Rental/Land), Entity Name, Owned By, "
                "Property Address, Property Name, Ownership %, Entity Structure, "
                "Cost Basis, Book Value, Existing Debt.",
            ],
            "error": "no_rows",
        }

    errors: list[str] = []
    to_insert: list[RentalOwnership] = []

    for row in parsed:
        co = _match_company(db, tid, row.entity_name)
        if not co:
            errors.append(f"Row {row.row_num}: company '{row.entity_name}' not found in Company Registry")
            continue
        suite = _match_property(db, co.id, row.property_name)
        to_insert.append(RentalOwnership(
            tenant_id=tid,
            company_id=co.id,
            property_id=suite.id if suite else None,
            partner_name=row.partner_name,
            property_name=suite.property_name if suite else row.property_name,
            property_address=row.property_address,
            entity_structure=row.entity_structure,
            entity_line=row.entity_line or "Rental",
            ownership_pct=row.ownership_pct,
            role=_role_from_structure(row.entity_structure),
            cost_basis=row.cost_basis,
            book_value=row.book_value,
            existing_debt=row.existing_debt,
            capital_contributed=row.cost_basis,
        ))

    if not to_insert:
        return {
            "imported_count": 0,
            "skipped_non_rental": parsed_result.skipped_non_rental,
            "errors": errors or ["No rows could be matched to companies in Company Registry."],
            "error": "no_rows",
        }

    existing = db.query(RentalOwnership).filter(RentalOwnership.tenant_id == tid).all()
    for old in existing:
        db.delete(old)
    db.flush()

    for rec in to_insert:
        db.add(rec)
    db.commit()

    companies = sorted({r.company.company_name for r in to_insert if r.company})
    msg = f"Imported {len(to_insert)} rental ownership position(s)."
    if parsed_result.skipped_non_rental:
        msg += f" Skipped {parsed_result.skipped_non_rental} non-rental row(s)."
    if errors:
        msg += f" Skipped {len(errors)} row(s) not in Company Registry."
    return {
        "imported_count": len(to_insert),
        "skipped_non_rental": parsed_result.skipped_non_rental,
        "errors": errors,
        "companies_updated": companies,
        "message": msg,
    }


def build_import_template_bytes() -> bytes:
    import openpyxl
    from openpyxl.styles import Font, PatternFill

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Ownership"
    headers = [
        "Entity", "Entity Name", "Owned By", "Property Address", "Property Name",
        "Ownership %", "Entity Structure", "Cost Basis", "Book Value", "Existing Debt",
    ]
    fill = PatternFill("solid", fgColor="FFFF00")
    bold = Font(bold=True)
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = fill
        cell.font = bold
    ws.append([
        "Rental", "Example LLC", "Partner A", "123 Main St", "Building One",
        0.25, "LLC", 500000, 480000, 300000,
    ])
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
