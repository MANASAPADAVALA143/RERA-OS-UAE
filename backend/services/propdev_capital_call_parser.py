"""Heuristic parser for messy, multi-table Property Dev capital-call workbooks."""
from __future__ import annotations

import os
import re
from dataclasses import dataclass, field
from datetime import date, datetime, timedelta
from io import BytesIO
from typing import Any, Literal

import openpyxl

BlockType = Literal["capital_call", "expense_builder", "property_pl", "unknown"]

OVERDUE_DAYS = int(os.getenv("PROPDEV_CAPITAL_CALL_OVERDUE_DAYS", "30"))
MONEY_TOLERANCE = 1.0


HEADER_ALIASES: dict[str, tuple[str, ...]] = {
    "serial": ("sl no", "sl no.", "s no", "s.no", "s.no.", "serial no", "s no", "#"),
    "partner_name": (
        "partner", "partners", "partner name", "partner details", "name of partner",
        "member", "member name", "shareholder", "investor", "entity",
    ),
    "ownership_pct": (
        "%", "% share", "% of share", "share %", "shareholding",
        "shareholding pattern", "shareholding pattern appfolio",
        "% of shareholding as per appfolio", "ownership %", "ownership percentage",
    ),
    # Period call amount — "Capital Contribution [dates]" / "Capital Call [dates]"
    "amount_called": (
        "amount called", "call amount", "new call",
        "capital contribution", "contribution amount", "amount due",
        "partner share", "call from partners",
        "capital to be called", "this call", "current call",
        "capital call",  # "Capital Call April 25 - September 25"
    ),
    # Opening + period total (wording varies across sheets)
    "total_due": (
        "total contribution", "total receivable", "total receiveable",
        "total due", "total payable", "total amount due",
        "total capital contribution", "total capital receivable",
    ),
    "old_dues": (
        "old dues", "previous balance", "opening balance", "balance brought forward",
        "prior due", "arrears", "previous dues", "opening dues",
    ),
    "received_date": (
        "received date", "date received", "payment date", "date of receipt",
        "contribution date",
    ),
    "amount_received": (
        "amount received", "received amount", "received", "paid",
        "amount paid", "capital received", "contribution received",
    ),
    "balance": (
        "balance", "bal", "bal.", "balance due", "outstanding",
        "amount outstanding", "closing balance", "balance receivable",
        "bal amount", "balance amount",
    ),
    "expense_category": ("expenses", "expense", "category", "particulars", "description"),
    "expense_amount": ("estimated amount", "cost", "total amount"),
}

PROPERTY_PL_TERMS = (
    "rental income", "loan emi", "property taxes", "property tax", "hoa",
    "net profit", "net loss", "net profit/(loss)", "maintenance",
)
CAPITAL_TITLE_TERMS = ("capital call", "capital contribution", "call from", "property tax")
COMPANY_SUFFIX_RE = re.compile(
    r"\b([A-Za-z0-9&.,' -]{2,}?\b(?:LLC|LP|LLP|Inc\.?|Ltd\.?|Corp\.?|"
    r"Holdings|Ventures|Development|Properties|Realty|Homes))\b",
    re.IGNORECASE,
)
DATE_PATTERNS = (
    "%Y-%m-%d", "%m/%d/%Y", "%d/%m/%Y", "%m-%d-%Y", "%d-%m-%Y",
    "%d-%b-%Y", "%b %d %Y", "%B %d %Y",
)

_MONTH_LOOKUP: dict[str, int] = {
    "jan": 1, "january": 1,
    "feb": 2, "february": 2,
    "mar": 3, "march": 3,
    "apr": 4, "april": 4,
    "may": 5,
    "jun": 6, "june": 6,
    "jul": 7, "july": 7,
    "aug": 8, "august": 8,
    "sep": 9, "sept": 9, "september": 9,
    "oct": 10, "october": 10,
    "nov": 11, "november": 11,
    "dec": 12, "december": 12,
}


def _expand_year(year: int) -> int:
    if year < 100:
        return 2000 + year if year < 70 else 1900 + year
    return year


def _month_end(year: int, month: int) -> date:
    if month == 12:
        return date(year, 12, 31)
    return date(year, month + 1, 1) - timedelta(days=1)


def _parse_month_token(token: str) -> int | None:
    return _MONTH_LOOKUP.get(token.strip().lower().rstrip("."))


def parse_contribution_period(text: str) -> tuple[date | None, date | None, str | None]:
    """
    Extract a capital-contribution date range from a header / title.

    Handles forms like:
      Capital Contribution Nov 25 - Apr 26
      Capital Contribution May25 - Jul25          (no space before year)
      Capital Contribution Receiveable May 25 - Oct 25
      Capital Contribution    Feb 25 - July 25
      Capital Call from May'25 to Oct'25        (apostrophe year)
      Jan to Jun 25
      Capital Call from Mar 2025 to May 2025

    Returns (period_start, period_end, label). End date is the last day of the
    end month — used for sorting "latest" periods.
    """
    raw = str(text or "").strip()
    if not raw:
        return None, None, None
    cleaned = re.sub(r"[–—]", "-", raw)
    # May'25 / May'2025 → May 25
    cleaned = re.sub(
        r"(?P<m>jan(?:uary)?|feb(?:ruary)?|mar(?:ch)?|apr(?:il)?|may|jun(?:e)?|"
        r"jul(?:y)?|aug(?:ust)?|sep(?:t(?:ember)?)?|oct(?:ober)?|nov(?:ember)?|"
        r"dec(?:ember)?)'(?P<y>\d{2,4})",
        lambda m: f"{m.group('m')} {m.group('y')}",
        cleaned,
        flags=re.IGNORECASE,
    )
    cleaned = re.sub(r"\s+", " ", cleaned)

    # Month YY/YYYY - Month YY/YYYY  (with optional "to"/"through"; year may be glued: May25)
    range_re = re.compile(
        r"(?P<m1>jan(?:uary)?|feb(?:ruary)?|mar(?:ch)?|apr(?:il)?|may|jun(?:e)?|"
        r"jul(?:y)?|aug(?:ust)?|sep(?:t(?:ember)?)?|oct(?:ober)?|nov(?:ember)?|"
        r"dec(?:ember)?)"
        r"\s*(?P<y1>\d{2,4})?"
        r"\s*(?:-|to|through|thru|until)\s*"
        r"(?P<m2>jan(?:uary)?|feb(?:ruary)?|mar(?:ch)?|apr(?:il)?|may|jun(?:e)?|"
        r"jul(?:y)?|aug(?:ust)?|sep(?:t(?:ember)?)?|oct(?:ober)?|nov(?:ember)?|"
        r"dec(?:ember)?)"
        r"\s*(?P<y2>\d{2,4})",
        re.IGNORECASE,
    )
    match = range_re.search(cleaned)
    if match:
        m1 = _parse_month_token(match.group("m1"))
        m2 = _parse_month_token(match.group("m2"))
        y2 = _expand_year(int(match.group("y2")))
        y1_raw = match.group("y1")
        if y1_raw:
            y1 = _expand_year(int(y1_raw))
        else:
            y1 = y2
            if m1 and m2 and m1 > m2:
                y1 = y2 - 1
        if m1 and m2:
            start = date(y1, m1, 1)
            end = _month_end(y2, m2)
            label = f"{match.group('m1').title()[:3]} {y1 % 100:02d} - {match.group('m2').title()[:3]} {y2 % 100:02d}"
            return start, end, label

    # Single month-year fallback near "capital contribution/call"
    single = re.search(
        r"(?:capital\s+(?:contribution|call)|contribution|call)[^\dA-Za-z]{0,40}"
        r"(?P<m>jan(?:uary)?|feb(?:ruary)?|mar(?:ch)?|apr(?:il)?|may|jun(?:e)?|"
        r"jul(?:y)?|aug(?:ust)?|sep(?:t(?:ember)?)?|oct(?:ober)?|nov(?:ember)?|"
        r"dec(?:ember)?)\s*(?P<y>\d{2,4})",
        cleaned,
        re.IGNORECASE,
    )
    if single:
        m = _parse_month_token(single.group("m"))
        y = _expand_year(int(single.group("y")))
        if m:
            start = date(y, m, 1)
            end = _month_end(y, m)
            label = f"{single.group('m').title()[:3]} {y % 100:02d}"
            return start, end, label

    return None, None, None


def _header_period_text(row: tuple[Any, ...] | list[Any], title: str, context: str) -> str:
    """Prefer the Capital Contribution column header text, then title/context."""
    prefer: list[str] = []
    other: list[str] = []
    for cell in row:
        text = str(cell or "").strip()
        if not text:
            continue
        low = text.lower()
        if (
            "capital contribution" in low
            or "capital call" in low
            or (
                "contribution" in low
                and re.search(r"\b(jan|feb|mar|apr|may|jun|jul|aug|sep|oct|nov|dec|\d{2})", text, re.I)
            )
        ):
            prefer.append(text)
        else:
            other.append(text)
    parts = prefer + other + [title, context]
    return " | ".join(p for p in parts if p)


def _attach_period_metadata(
    block: ParsedBlock,
    header_row: tuple[Any, ...] | list[Any],
    context: str,
) -> None:
    haystack = _header_period_text(header_row, block.title, context)
    start, end, label = parse_contribution_period(haystack)
    if not label:
        # Try each header cell independently (date range may sit only in one column)
        for cell in header_row:
            start, end, label = parse_contribution_period(str(cell or ""))
            if label:
                break
    if not label:
        start, end, label = parse_contribution_period(block.title)
    if not label:
        start, end, label = parse_contribution_period(context)

    block.period_start = start
    block.period_end = end
    block.period_label = label
    if not end:
        block.date_range_unclear = True
        block.warnings.append(
            "Date range unclear — needs manual date range confirmation before trusting this period."
        )
    if end and not block.call_date:
        block.call_date = end


def select_latest_capital_blocks(blocks: list[ParsedBlock]) -> list[ParsedBlock]:
    """
    Per sheet, keep only the capital-call table with the latest period END date.
    When no end dates are parseable, keep the lowest table on the sheet (highest
    header_row) and flag date_range_unclear.
    """
    capital = [b for b in blocks if b.block_type == "capital_call"]
    others = [b for b in blocks if b.block_type != "capital_call"]
    by_sheet: dict[str, list[ParsedBlock]] = {}
    for block in capital:
        by_sheet.setdefault(block.sheet_name, []).append(block)

    selected: list[ParsedBlock] = []
    for sheet_name, sheet_blocks in by_sheet.items():
        for block in sheet_blocks:
            block.selected_as_latest = False
            block.skipped_period_labels = []

        dated = [b for b in sheet_blocks if b.period_end is not None]
        if dated:
            latest = max(dated, key=lambda b: (b.period_end or date.min, b.header_row))
        elif len(sheet_blocks) == 1:
            latest = sheet_blocks[0]
            latest.date_range_unclear = True
        else:
            # Multiple undated tables — prefer the last table on the sheet + flag.
            latest = max(sheet_blocks, key=lambda b: b.header_row)
            latest.date_range_unclear = True
            latest.warnings.append(
                "Multiple contribution tables without parseable date ranges; "
                "selected the last table on the sheet — review before import."
            )

        skipped_labels: list[str] = []
        for block in sheet_blocks:
            if block is latest:
                continue
            label = block.period_label or block.title or f"header row {block.header_row}"
            skipped_labels.append(label)

        latest.selected_as_latest = True
        latest.skipped_period_labels = skipped_labels
        selected.append(latest)

        # Mark non-selected capital blocks as not latest (kept in all blocks for audit)
        for block in sheet_blocks:
            if block is not latest:
                block.selected_as_latest = False

    return others + selected


def build_company_preview(blocks: list[ParsedBlock], all_capital: list[ParsedBlock]) -> list[dict[str, Any]]:
    """
    Preview rows: every sheet/company with capital tables — including unclear date ranges.
    `blocks` should be post-selection (latest-only capital_call_blocks + other types).
    `all_capital` is every capital block before selection (for skipped ranges).
    """
    selected_capital = [b for b in blocks if b.block_type == "capital_call"]
    by_sheet_all: dict[str, list[ParsedBlock]] = {}
    for block in all_capital:
        by_sheet_all.setdefault(block.sheet_name, []).append(block)

    preview: list[dict[str, Any]] = []
    seen_sheets: set[str] = set()

    for block in selected_capital:
        seen_sheets.add(block.sheet_name)
        siblings = by_sheet_all.get(block.sheet_name, [])
        all_ranges = []
        for sib in siblings:
            all_ranges.append({
                "label": sib.period_label or sib.title or f"Row {sib.header_row}",
                "period_start": sib.period_start.isoformat() if sib.period_start else None,
                "period_end": sib.period_end.isoformat() if sib.period_end else None,
                "partner_rows": len(sib.rows),
                "selected": sib is block or sib.selected_as_latest,
                "date_range_unclear": sib.date_range_unclear,
            })
        preview.append({
            "sheet_name": block.sheet_name,
            "company_id": block.company_id,
            "company_name": block.company_name,
            "matched": bool(block.company_id),
            "attribution_confidence": block.attribution_confidence,
            "attribution_reason": block.attribution_reason,
            "latest_date_range": block.period_label,
            "period_start": block.period_start.isoformat() if block.period_start else None,
            "period_end": block.period_end.isoformat() if block.period_end else None,
            "partner_rows": len(block.rows),
            "skipped_date_ranges": block.skipped_period_labels,
            "all_detected_ranges": all_ranges,
            "date_range_unclear": block.date_range_unclear,
            "warning_badge": (
                "Date range unclear — review before import"
                if block.date_range_unclear
                else None
            ),
            "warnings": list(block.warnings),
            "totals": dict(block.computed_totals),
        })

    # Sheets that had capital tables but none selected (shouldn't happen) / unattributed
    for sheet_name, siblings in by_sheet_all.items():
        if sheet_name in seen_sheets:
            continue
        first = siblings[0]
        preview.append({
            "sheet_name": sheet_name,
            "company_id": first.company_id,
            "company_name": first.company_name,
            "matched": bool(first.company_id),
            "attribution_confidence": first.attribution_confidence,
            "attribution_reason": first.attribution_reason,
            "latest_date_range": None,
            "period_start": None,
            "period_end": None,
            "partner_rows": sum(len(s.rows) for s in siblings),
            "skipped_date_ranges": [],
            "all_detected_ranges": [
                {
                    "label": s.period_label or s.title or f"Row {s.header_row}",
                    "period_start": s.period_start.isoformat() if s.period_start else None,
                    "period_end": s.period_end.isoformat() if s.period_end else None,
                    "partner_rows": len(s.rows),
                    "selected": False,
                    "date_range_unclear": s.date_range_unclear,
                }
                for s in siblings
            ],
            "date_range_unclear": True,
            "warning_badge": "Date range unclear — review before import",
            "warnings": ["No latest period could be selected for this sheet."],
            "totals": {},
        })

    preview.sort(key=lambda row: (row.get("company_name") or row["sheet_name"] or "").lower())
    return preview



@dataclass
class CompanyCandidate:
    id: str
    name: str
    property_name: str = ""
    address: str = ""


@dataclass
class PartnerCallRow:
    partner_name: str
    share_percent: float
    amount_called: float
    old_dues: float
    amount_received: float
    balance: float
    received_date: date | None
    row_number: int
    total_due: float = 0.0


@dataclass
class ParsedBlock:
    block_type: BlockType
    sheet_name: str
    title: str
    header_row: int
    start_row: int
    end_row: int
    call_date: date | None = None
    company_id: str | None = None
    company_name: str | None = None
    property_name: str | None = None
    attribution_confidence: str = "unattributed"
    attribution_reason: str = ""
    rows: list[PartnerCallRow] = field(default_factory=list)
    expense_rows: list[dict[str, Any]] = field(default_factory=list)
    pl_rows: list[dict[str, Any]] = field(default_factory=list)
    source_totals: dict[str, float] = field(default_factory=dict)
    computed_totals: dict[str, float] = field(default_factory=dict)
    warnings: list[str] = field(default_factory=list)
    linked_call_title: str | None = None
    # Contribution period (from "Capital Contribution Nov 25 - Apr 26" etc.)
    period_label: str | None = None
    period_start: date | None = None
    period_end: date | None = None
    date_range_unclear: bool = False
    selected_as_latest: bool = False
    skipped_period_labels: list[str] = field(default_factory=list)


@dataclass
class WorkbookParseResult:
    blocks: list[ParsedBlock]
    capital_call_blocks: list[ParsedBlock]
    expense_builder_blocks: list[ParsedBlock]
    property_pl_blocks: list[ParsedBlock]
    unknown_blocks: list[ParsedBlock]
    manual_review: list[dict[str, Any]]
    # Per-company / per-sheet matching summary for Preview UI (before commit)
    company_preview: list[dict[str, Any]] = field(default_factory=list)


def _norm(value: Any) -> str:
    text = str(value or "").strip().lower()
    text = text.replace("_", " ").replace("\n", " ")
    text = re.sub(r"[%$()/:–—-]+", lambda m: "%" if "%" in m.group(0) else " ", text)
    return re.sub(r"\s+", " ", text).strip(" .")


_CORP_SUFFIX_RE = re.compile(
    r"\b(llc|l\.l\.c\.|lp|l\.p\.|llp|inc\.?|ltd\.?|corp\.?|incorporated|limited)\b",
    re.IGNORECASE,
)


def _company_match_key(name: str) -> str:
    """Normalize company names for fuzzy match (strip corp suffixes / punctuation)."""
    text = _norm(name)
    text = _CORP_SUFFIX_RE.sub(" ", text)
    text = re.sub(r"[^a-z0-9\s&]", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def _company_match_compact(name: str) -> str:
    """Space-insensitive key so 'GeorgeTown' matches 'George Town'."""
    return _company_match_key(name).replace(" ", "")


def _fuzzy_find_companies(
    needle: str,
    companies: list[CompanyCandidate],
) -> list[CompanyCandidate]:
    """Exact key match, then unique containment / starts-with against registry names."""
    key = _company_match_key(needle)
    compact = _company_match_compact(needle)
    if not key or len(key) < 3:
        return []
    exact = [c for c in companies if _company_match_key(c.name) == key]
    if exact:
        return exact
    compact_exact = [c for c in companies if _company_match_compact(c.name) == compact]
    if compact_exact:
        return compact_exact
    contained: list[CompanyCandidate] = []
    for company in companies:
        ck = _company_match_key(company.name)
        if not ck:
            continue
        if key in ck or ck in key:
            contained.append(company)
            continue
        cc = _company_match_compact(company.name)
        if compact and (compact in cc or cc in compact):
            contained.append(company)
            continue
        # "sparks shepard" vs "sparks shepard ventures group"
        key_tokens = key.split()
        ck_tokens = ck.split()
        if len(key_tokens) >= 2 and all(t in ck_tokens for t in key_tokens):
            contained.append(company)
    return contained


def _row_text(row: tuple[Any, ...] | list[Any]) -> str:
    return " | ".join(str(v).strip() for v in row if v not in (None, ""))


def _to_float(value: Any) -> float:
    if value is None or value == "":
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace("$", "").replace(",", "")
    if text in {"", "-", "—", "n/a", "N/A"}:
        return 0.0
    negative = text.startswith("(") and text.endswith(")")
    if negative:
        text = text[1:-1]
    text = text.replace("%", "")
    try:
        amount = float(text)
        return -amount if negative else amount
    except ValueError:
        return 0.0


def _to_date(value: Any) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value or "").strip()
    if not text:
        return None
    for fmt in DATE_PATTERNS:
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    match = re.search(
        r"\b(\d{1,2})[-/ ]([A-Za-z]{3,9}|\d{1,2})[-/ ](\d{2,4})\b", text
    )
    if match:
        candidate = match.group(0)
        for fmt in ("%d-%m-%Y", "%d/%m/%Y", "%d %b %Y", "%d-%b-%Y"):
            try:
                return datetime.strptime(candidate, fmt).date()
            except ValueError:
                continue
    return None


def _canonical_header(value: Any) -> str | None:
    """Map a header cell to a canonical field using longest-alias wins (not first match)."""
    normalized = _norm(value)
    if not normalized:
        return None

    # Bare "amount" is ambiguous — handled in _header_map after other fields resolve.
    if normalized in {"amount", "amt"}:
        return None

    best: tuple[int, str] | None = None
    for canonical, aliases in HEADER_ALIASES.items():
        for alias in aliases:
            alias_norm = _norm(alias)
            if not alias_norm:
                continue
            score = 0
            if normalized == alias_norm:
                score = 1000 + len(alias_norm)
            elif alias_norm in normalized:
                # Prefer specific aliases; reject tiny accidental substrings
                if len(alias_norm) < 3:
                    continue
                if canonical == "partner_name" and alias_norm == "partner":
                    if not (
                        normalized in {
                            "partner", "partners", "partner name", "partner details",
                            "name of partner",
                        }
                        or normalized.startswith("partner")
                    ):
                        continue
                # "received" must not steal "balance receivable" / "total receivable"
                if canonical == "amount_received" and alias_norm == "received":
                    if "receivable" in normalized or "receiveable" in normalized:
                        continue
                    if normalized not in {
                        "received", "amount received", "received amount",
                        "capital received", "contribution received",
                    } and not normalized.startswith("received"):
                        # Allow exact-ish forms only
                        if "received" != alias_norm:
                            continue
                # "capital call" / "capital contribution" OK inside longer dated headers
                score = 100 + len(alias_norm)
            if score and (best is None or score > best[0]):
                best = (score, canonical)
    return best[1] if best else None


def _header_map(row: tuple[Any, ...] | list[Any]) -> dict[str, int]:
    mapped: dict[str, int] = {}
    for idx, value in enumerate(row):
        canonical = _canonical_header(value)
        if canonical and canonical not in mapped:
            mapped[canonical] = idx

    # Pecan-style: "Amount" after Date of Receipt = money received (not expense)
    if "amount_received" not in mapped and "received_date" in mapped:
        for idx, value in enumerate(row):
            if _norm(value) in {"amount", "amt"} and idx not in mapped.values():
                mapped["amount_received"] = idx
                break

    return mapped


def _classify_header(
    row: tuple[Any, ...] | list[Any],
    nearby_text: str,
) -> tuple[BlockType, dict[str, int], int]:
    mapped = _header_map(row)
    normalized_cells = [_norm(v) for v in row if _norm(v)]
    row_joined = " ".join(normalized_cells)
    nearby = _norm(nearby_text)

    pl_hits = sum(1 for term in PROPERTY_PL_TERMS if _norm(term) in row_joined)
    if pl_hits >= 2:
        return "property_pl", mapped, pl_hits + 4

    has_partner = "partner_name" in mapped
    has_share = "ownership_pct" in mapped
    money_fields = {
        key for key in ("amount_called", "amount_received", "balance", "old_dues", "total_due")
        if key in mapped
    }
    title_has_call = any(term in nearby for term in CAPITAL_TITLE_TERMS)
    # Require money columns OR an explicit capital-call title — Partner + % Share +
    # year columns (Annexure II operating-expense rollups) must NOT become calls.
    if has_partner and len(money_fields) >= 1 and (has_share or title_has_call or len(money_fields) >= 2):
        return "capital_call", mapped, 8 + len(money_fields)
    if has_partner and has_share and title_has_call and len(money_fields) >= 1:
        return "capital_call", mapped, 7 + len(money_fields)
    if has_share and len(money_fields) >= 2:
        return "capital_call", mapped, 6 + len(money_fields)

    has_expense = "expense_category" in mapped
    has_amount = "expense_amount" in mapped or any(_norm(v) in {"amount", "amt"} for v in row)
    if has_expense and has_amount and not has_partner and not has_share:
        # Re-map bare Amount for expense blocks
        if "expense_amount" not in mapped:
            for idx, value in enumerate(row):
                if _norm(value) in {"amount", "amt", "total amount"}:
                    mapped["expense_amount"] = idx
                    break
        return "expense_builder", mapped, 5

    return "unknown", mapped, len(mapped)


def _nearest_title(rows: list[tuple[Any, ...]], header_idx: int) -> str:
    candidates: list[str] = []
    for idx in range(max(0, header_idx - 5), header_idx):
        text = _row_text(rows[idx])
        if text:
            candidates.append(text)
    if not candidates:
        return ""
    capital = [c for c in candidates if any(t in _norm(c) for t in CAPITAL_TITLE_TERMS)]
    return (capital or candidates)[-1]


def _context_text(rows: list[tuple[Any, ...]], header_idx: int) -> str:
    start = max(0, header_idx - 8)
    return " | ".join(_row_text(r) for r in rows[start:header_idx + 1] if _row_text(r))


def _next_boundary(
    rows: list[tuple[Any, ...]],
    header_idx: int,
    current_type: BlockType,
) -> int:
    blank_run = 0
    saw_data = False
    for idx in range(header_idx + 1, len(rows)):
        row = rows[idx]
        text = _row_text(row)
        if not text:
            blank_run += 1
            if saw_data and blank_run >= 2:
                return idx - blank_run
            continue
        blank_run = 0
        saw_data = True
        nearby = _context_text(rows, idx)
        detected, _, score = _classify_header(row, nearby)
        if detected != "unknown" and score >= 5:
            return idx - 1
        if _norm(text).startswith(("total", "grand total", "subtotal")):
            return idx
    return len(rows) - 1


def _attribute_block(
    block: ParsedBlock,
    context: str,
    companies: list[CompanyCandidate],
) -> None:
    # Primary path: Excel tab name matches Company Registry (exact, then fuzzy).
    sheet_norm = _norm(block.sheet_name)
    if sheet_norm:
        exact = [c for c in companies if _norm(c.name) == sheet_norm]
        if len(exact) == 1:
            company = exact[0]
            block.company_id = company.id
            block.company_name = company.name
            block.property_name = company.property_name
            block.attribution_confidence = "high"
            block.attribution_reason = "sheet tab matches company registry"
            return
        if len(exact) > 1:
            block.warnings.append(
                f"Multiple registry companies match sheet tab '{block.sheet_name}'; block was not imported."
            )
            return
        fuzzy_sheet = _fuzzy_find_companies(block.sheet_name, companies)
        if len(fuzzy_sheet) == 1:
            company = fuzzy_sheet[0]
            block.company_id = company.id
            block.company_name = company.name
            block.property_name = company.property_name
            block.attribution_confidence = "high"
            block.attribution_reason = "sheet tab fuzzy-matches company registry"
            return
        if len(fuzzy_sheet) > 1:
            block.warnings.append(
                f"Multiple registry companies fuzzy-match sheet tab '{block.sheet_name}'; block was not imported."
            )
            return

    haystack = _norm(f"{block.sheet_name} {block.title} {context}")
    scores: list[tuple[int, CompanyCandidate, list[str]]] = []
    for company in companies:
        score = 0
        reasons: list[str] = []
        name = _norm(company.name)
        name_key = _company_match_key(company.name)
        prop = _norm(company.property_name)
        address = _norm(company.address)
        hay_key = _company_match_key(haystack)
        if name and name in haystack:
            score += 10
            reasons.append("company name")
        elif name_key and len(name_key) >= 4 and (name_key in hay_key or hay_key in name_key):
            score += 9
            reasons.append("company name (fuzzy)")
        if prop and len(prop) >= 5 and prop in haystack:
            score += 6
            reasons.append("property name")
        if address:
            address_tokens = [t for t in address.split() if len(t) >= 4]
            token_hits = sum(1 for token in address_tokens if token in haystack)
            if token_hits >= 2:
                score += min(5, token_hits)
                reasons.append("address")
        if score:
            scores.append((score, company, reasons))
    scores.sort(key=lambda item: item[0], reverse=True)

    if scores and (len(scores) == 1 or scores[0][0] > scores[1][0]):
        score, company, reasons = scores[0]
        block.company_id = company.id
        block.company_name = company.name
        block.property_name = company.property_name
        block.attribution_confidence = "high" if score >= 10 else "medium"
        block.attribution_reason = ", ".join(reasons)
        return

    company_matches = COMPANY_SUFFIX_RE.findall(
        f"{block.title} {context}"
    )
    unique_matches = list(dict.fromkeys(m.strip(" -|") for m in company_matches if m.strip()))
    if len(unique_matches) == 1:
        fuzzy = _fuzzy_find_companies(unique_matches[0], companies)
        if len(fuzzy) == 1:
            company = fuzzy[0]
            block.company_id = company.id
            block.company_name = company.name
            block.property_name = company.property_name
            block.attribution_confidence = "high"
            block.attribution_reason = "title company name fuzzy-matches registry"
            return
        block.company_name = unique_matches[0]
        block.attribution_confidence = "medium"
        block.attribution_reason = "explicit company-like title; company record not matched"
        return

    if len(companies) == 1:
        company = companies[0]
        block.company_id = company.id
        block.company_name = company.name
        block.property_name = company.property_name
        block.attribution_confidence = "low"
        block.attribution_reason = "only company in registry"
        block.warnings.append("Attribution uses the only registered company; confirm manually.")
        return

    block.warnings.append("Company/property attribution is ambiguous; block was not imported.")


def _is_total_row(row: tuple[Any, ...] | list[Any]) -> bool:
    """True when any early cell is exactly Total / Grand Total (not a partner named Total X)."""
    for v in row[:8]:
        n = _norm(v)
        if n in {"total", "grand total", "subtotal"}:
            return True
    return False


def _value(row: tuple[Any, ...], mapped: dict[str, int], field: str) -> Any:
    idx = mapped.get(field)
    return row[idx] if idx is not None and idx < len(row) else None


def _parse_capital_block(
    rows: list[tuple[Any, ...]],
    block: ParsedBlock,
    mapped: dict[str, int],
) -> None:
    source_called = source_received = source_balance = 0.0
    has_source_total = False
    for idx in range(block.start_row - 1, block.end_row):
        row = rows[idx]
        if _is_total_row(row):
            has_source_total = True
            # Prefer period call column; fall back to Total Contribution/Receivable
            source_called = _to_float(_value(row, mapped, "amount_called"))
            if source_called <= 0:
                source_called = _to_float(_value(row, mapped, "total_due"))
            source_received = _to_float(_value(row, mapped, "amount_received"))
            source_balance = _to_float(_value(row, mapped, "balance"))
            continue

        partner = str(_value(row, mapped, "partner_name") or "").strip()
        if not partner or _norm(partner) in {
            "partner", "partners", "partner name", "partner details", "entity", "total",
        }:
            continue
        if any(term in _norm(partner) for term in PROPERTY_PL_TERMS):
            continue

        share_raw = _value(row, mapped, "ownership_pct")
        share = _to_float(share_raw)
        if isinstance(share_raw, (int, float)) and 0 < share <= 1:
            share *= 100

        old_dues = max(0.0, _to_float(_value(row, mapped, "old_dues")))
        called = max(0.0, _to_float(_value(row, mapped, "amount_called")))
        total_due_cell = max(0.0, _to_float(_value(row, mapped, "total_due")))
        received = max(0.0, _to_float(_value(row, mapped, "amount_received")))
        balance_cell = _value(row, mapped, "balance")
        balance = max(0.0, _to_float(balance_cell))

        # If period call column empty but Total Receivable present, derive call amount
        if called <= 0 and total_due_cell > 0:
            called = max(0.0, total_due_cell - old_dues) if old_dues > 0 else total_due_cell

        if called <= 0 and (received > 0 or balance > 0):
            called = received + balance
        if balance_cell in (None, "") and called > 0:
            balance = max(0.0, (total_due_cell or (old_dues + called)) - received)

        total_due = total_due_cell if total_due_cell > 0 else round(old_dues + called, 2)
        received_date = _to_date(_value(row, mapped, "received_date"))

        if called <= 0 and received <= 0 and balance <= 0 and total_due <= 0:
            continue
        block.rows.append(PartnerCallRow(
            partner_name=partner,
            share_percent=round(share, 4),
            amount_called=round(called, 2),
            old_dues=round(old_dues, 2),
            amount_received=round(received, 2),
            balance=round(balance, 2),
            received_date=received_date,
            row_number=idx + 1,
            total_due=round(total_due, 2),
        ))

    computed_called = round(sum(r.amount_called for r in block.rows), 2)
    computed_received = round(sum(r.amount_received for r in block.rows), 2)
    computed_balance = round(sum(r.balance for r in block.rows), 2)
    block.computed_totals = {
        "called": computed_called,
        "received": computed_received,
        "balance": computed_balance,
        "total_due": round(sum(r.total_due for r in block.rows), 2),
    }
    if has_source_total:
        block.source_totals = {
            "called": source_called,
            "received": source_received,
            "balance": source_balance,
        }
        for field, source, computed in (
            ("called", source_called, computed_called),
            ("received", source_received, computed_received),
            ("balance", source_balance, computed_balance),
        ):
            if source and abs(source - computed) > MONEY_TOLERANCE:
                block.warnings.append(
                    f"Source Total {field} ${source:,.2f} does not match partner rows "
                    f"${computed:,.2f} (difference ${source - computed:,.2f})."
                )
    method_balance = round(computed_called - computed_received, 2)
    if abs(method_balance - computed_balance) > MONEY_TOLERANCE:
        block.warnings.append(
            f"Balance cross-check mismatch: Called − Received = ${method_balance:,.2f}, "
            f"but Balance column sums to ${computed_balance:,.2f}."
        )
    if not block.rows:
        block.warnings.append("Capital-call header detected, but no partner amount rows were parsed.")


def _parse_expense_block(
    rows: list[tuple[Any, ...]],
    block: ParsedBlock,
    mapped: dict[str, int],
) -> None:
    for idx in range(block.start_row - 1, block.end_row):
        row = rows[idx]
        if _is_total_row(row):
            continue
        category = str(_value(row, mapped, "expense_category") or "").strip()
        amount = _to_float(_value(row, mapped, "expense_amount"))
        if category and amount:
            block.expense_rows.append({
                "category": category,
                "amount": round(amount, 2),
                "row_number": idx + 1,
            })


def _parse_pl_block(
    rows: list[tuple[Any, ...]],
    block: ParsedBlock,
) -> None:
    header = rows[block.header_row - 1]
    columns = [str(v or "").strip() for v in header]
    for idx in range(block.start_row - 1, block.end_row):
        row = rows[idx]
        label = str(row[0] or "").strip()
        if not label:
            continue
        values = {
            columns[col] or f"Column {col + 1}": _to_float(row[col])
            for col in range(1, min(len(columns), len(row)))
            if row[col] not in (None, "")
        }
        if values:
            block.pl_rows.append({"label": label, "values": values, "row_number": idx + 1})


def _link_expense_blocks(blocks: list[ParsedBlock]) -> None:
    calls = [b for b in blocks if b.block_type == "capital_call"]
    for expense in [b for b in blocks if b.block_type == "expense_builder"]:
        candidates = [
            call for call in calls
            if call.sheet_name == expense.sheet_name
            and (
                not expense.company_name or not call.company_name
                or expense.company_name == call.company_name
            )
        ]
        if not candidates:
            expense.warnings.append("Expense justification could not be linked to a capital call.")
            continue
        nearest = min(candidates, key=lambda call: abs(call.header_row - expense.header_row))
        if abs(nearest.header_row - expense.header_row) <= 25:
            expense.linked_call_title = nearest.title
        else:
            expense.warnings.append("Expense block is standalone; no nearby capital call found.")


def parse_workbook(
    content: bytes,
    companies: list[CompanyCandidate] | None = None,
) -> WorkbookParseResult:
    workbook = openpyxl.load_workbook(BytesIO(content), data_only=True)
    candidates = companies or []
    blocks: list[ParsedBlock] = []

    for sheet_name in workbook.sheetnames:
        ws = workbook[sheet_name]
        rows = [tuple(row) for row in ws.iter_rows(values_only=True)]
        used_until = -1
        for header_idx, row in enumerate(rows):
            if header_idx <= used_until:
                continue
            context = _context_text(rows, header_idx)
            block_type, mapped, score = _classify_header(row, context)
            if block_type == "unknown" or score < 5:
                continue
            end_idx = _next_boundary(rows, header_idx, block_type)
            title = _nearest_title(rows, header_idx) or sheet_name
            block = ParsedBlock(
                block_type=block_type,
                sheet_name=sheet_name,
                title=title,
                header_row=header_idx + 1,
                start_row=header_idx + 2,
                end_row=end_idx + 1,
                call_date=_to_date(title) or _to_date(context),
            )
            _attribute_block(block, context, candidates)
            if block_type == "capital_call":
                _attach_period_metadata(block, row, context)
                _parse_capital_block(rows, block, mapped)
            elif block_type == "expense_builder":
                _parse_expense_block(rows, block, mapped)
            elif block_type == "property_pl":
                _parse_pl_block(rows, block)
            blocks.append(block)
            used_until = end_idx

    _link_expense_blocks(blocks)
    all_capital = [b for b in blocks if b.block_type == "capital_call"]
    # Keep only the latest contribution table per sheet (by period END date)
    selected_blocks = select_latest_capital_blocks(blocks)
    capital = [b for b in selected_blocks if b.block_type == "capital_call"]
    expenses = [b for b in selected_blocks if b.block_type == "expense_builder"]
    property_pl = [b for b in selected_blocks if b.block_type == "property_pl"]
    unknown = [b for b in selected_blocks if b.block_type == "unknown"]
    company_preview = build_company_preview(selected_blocks, all_capital)

    manual_review = [
        {
            "sheet": b.sheet_name,
            "title": b.title,
            "type": b.block_type,
            "reason": "; ".join(b.warnings) or "Attribution requires confirmation",
            "date_range_unclear": b.date_range_unclear,
            "latest_date_range": b.period_label,
        }
        for b in capital
        if (
            b.attribution_confidence == "unattributed"
            or b.company_id is None
        )
    ]
    # Also surface capital sheets that never matched a company
    for row in company_preview:
        if not row.get("matched"):
            manual_review.append({
                "sheet": row["sheet_name"],
                "title": row.get("latest_date_range") or row["sheet_name"],
                "type": "capital_call",
                "reason": row.get("attribution_reason") or "Company/property attribution is ambiguous; block was not imported.",
                "date_range_unclear": row.get("date_range_unclear", False),
                "latest_date_range": row.get("latest_date_range"),
            })

    # Dedupe manual_review by sheet+reason
    seen_mr: set[str] = set()
    deduped_mr: list[dict[str, Any]] = []
    for item in manual_review:
        key = f"{item.get('sheet')}|{item.get('reason')}"
        if key in seen_mr:
            continue
        seen_mr.add(key)
        deduped_mr.append(item)

    return WorkbookParseResult(
        blocks=selected_blocks,
        capital_call_blocks=capital,
        expense_builder_blocks=expenses,
        property_pl_blocks=property_pl,
        unknown_blocks=unknown,
        manual_review=deduped_mr,
        company_preview=company_preview,
    )


def call_status(
    amount_called: float,
    amount_received: float,
    balance: float,
    call_date: date | None,
    received_date: date | None,
    overdue_days: int = OVERDUE_DAYS,
) -> tuple[str, date | None]:
    if balance <= MONEY_TOLERANCE or amount_received >= amount_called - MONEY_TOLERANCE:
        return "Paid", call_date + timedelta(days=overdue_days) if call_date else None
    due_date = call_date + timedelta(days=overdue_days) if call_date else None
    if due_date and date.today() > due_date and not received_date:
        return "Overdue", due_date
    if amount_received > 0:
        return "Partial", due_date
    return "Outstanding", due_date


def block_to_dict(block: ParsedBlock) -> dict[str, Any]:
    return {
        "type": block.block_type,
        "sheet": block.sheet_name,
        "title": block.title,
        "header_row": block.header_row,
        "start_row": block.start_row,
        "end_row": block.end_row,
        "call_date": block.call_date.isoformat() if block.call_date else None,
        "company_id": block.company_id,
        "company_name": block.company_name,
        "property_name": block.property_name,
        "attribution_confidence": block.attribution_confidence,
        "attribution_reason": block.attribution_reason,
        "period_label": block.period_label,
        "period_start": block.period_start.isoformat() if block.period_start else None,
        "period_end": block.period_end.isoformat() if block.period_end else None,
        "date_range_unclear": block.date_range_unclear,
        "selected_as_latest": block.selected_as_latest,
        "skipped_period_labels": block.skipped_period_labels,
        "partner_rows": [
            {
                "partner_name": row.partner_name,
                "share_percent": row.share_percent,
                "amount_called": row.amount_called,
                "old_dues": row.old_dues,
                "amount_received": row.amount_received,
                "balance": row.balance,
                "received_date": row.received_date.isoformat() if row.received_date else None,
                "row_number": row.row_number,
            }
            for row in block.rows
        ],
        "expense_rows": block.expense_rows,
        "pl_rows": block.pl_rows,
        "source_totals": block.source_totals,
        "computed_totals": block.computed_totals,
        "warnings": block.warnings,
        "linked_call_title": block.linked_call_title,
    }
