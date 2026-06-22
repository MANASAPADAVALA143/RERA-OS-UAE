"""
CFO Dashboard Excel export — values-only (no formulas).
All numbers sourced from live app calculations, not recomputed here.
"""
from __future__ import annotations

import io
from datetime import date

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ── Palette ────────────────────────────────────────────────────────────────────
_DARK_BLUE  = "1F3864"
_MID_BLUE   = "2E5FA3"
_LT_BLUE_BG = "D9E1F2"
_AMBER_BG   = "FFF2CC"
_GREEN_BG   = "E2EFDA"
_RED_BG     = "FCE4D6"
_WHITE      = "FFFFFF"
_GRAY_BG    = "F2F2F2"
_BROWN      = "7B3F00"


def _border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)


def _hdr(ws, row, col, text, bg=_DARK_BLUE, fg=_WHITE, bold=True, size=10, wrap=False):
    c = ws.cell(row=row, column=col, value=text)
    c.font = Font(bold=bold, color=fg, size=size, name="Calibri")
    c.fill = PatternFill("solid", fgColor=bg)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=wrap)
    c.border = _border()
    return c


def _cell(ws, row, col, value, bg=None, bold=False, align="left", fmt=None):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(bold=bold, color="000000", size=10, name="Calibri")
    if bg:
        c.fill = PatternFill("solid", fgColor=bg)
    c.alignment = Alignment(horizontal=align, vertical="center")
    c.border = _border()
    if fmt:
        c.number_format = fmt
    return c


# ── COMPLIANCE (static reference — NOT live data) ─────────────────────────────

_COMPLIANCE = [
    ("Texas LLC/LP",   "Annual Franchise Tax Report",        "May 15",              "Texas Comptroller — computed on prior-year revenue"),
    ("Texas LLC/LP",   "Franchise Tax Prepayment",           "May 15",              "Applies if prior-year tax > $1,000"),
    ("Texas LLC/LP",   "Public Information Report (PIR)",    "May 15",              "Filed together with Franchise Tax return"),
    ("Federal",        "1065 Partnership Return",            "March 15",            "Extension to Sep 15 available"),
    ("Federal",        "1120-S S-Corp Return",               "March 15",            "Extension to Sep 15 available"),
    ("Federal",        "Schedule K-1 (Partners/Members)",    "March 15",            "Distribute to all partners/members"),
    ("Federal",        "1099-NEC (Vendors >$600)",           "January 31",          "File with IRS and send to payees"),
    ("Federal",        "1099-MISC (Rent >$600)",             "January 31",          "For rent paid to non-corporate landlords"),
    ("Federal",        "W-9 Collection",                     "Before first payment","Collect from all new vendors"),
    ("Federal",        "Estimated Tax Q1",                   "April 15",            "If partners owe >$1,000 federal tax"),
    ("Federal",        "Estimated Tax Q2",                   "June 15",             ""),
    ("Federal",        "Estimated Tax Q3",                   "September 15",        ""),
    ("Federal",        "Estimated Tax Q4",                   "Jan 15 (next yr)",    ""),
    ("Property Tax",   "Texas Protest Deadline",             "May 15",              "30 days after appraisal notice — check county"),
    ("Property Tax",   "Payment — Harris Co.",               "January 31",          "Without penalty; half-pay option Nov 30/Jun 30"),
    ("Property Tax",   "Payment — Travis Co.",               "January 31",          "Without penalty"),
    ("1031 Exchange",  "45-Day Identification Deadline",     "45 days from close",  "Identify replacement properties in writing"),
    ("1031 Exchange",  "180-Day Exchange Deadline",          "180 days from close", "Close on replacement property"),
    ("Texas",          "Sales & Use Tax (if applicable)",    "20th of next month",  "Monthly or quarterly depending on volume"),
    ("Insurance",      "CGL Policy Renewal",                 "Per policy",          "Review annually; confirm adequate limits"),
    ("Insurance",      "Property Policy Renewal",            "Per policy",          "Ensure replacement-cost coverage"),
    ("Insurance",      "Umbrella Policy Renewal",            "Per policy",          "Minimum $2M recommended for multi-entity portfolio"),
]


def _compliance_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("COMPLIANCE")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 36
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 54

    ws.merge_cells("A1:D1")
    c = ws["A1"]
    c.value = "COMPLIANCE & FILING CALENDAR — STATIC REFERENCE (not live data)"
    c.font = Font(bold=True, color=_WHITE, size=13, name="Calibri")
    c.fill = PatternFill("solid", fgColor=_DARK_BLUE)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    ws.merge_cells("A2:D2")
    c = ws["A2"]
    c.value = "This sheet is a hardcoded US/Texas reference calendar — it does NOT update from the app and is not tenant-specific."
    c.font = Font(italic=True, color="888888", size=9, name="Calibri")
    c.alignment = Alignment(horizontal="center")
    ws.row_dimensions[2].height = 14

    for col, label in enumerate(["Category", "Filing / Deadline", "Date / Window", "Notes"], 1):
        _hdr(ws, 3, col, label, bg=_MID_BLUE, wrap=True)
    ws.row_dimensions[3].height = 20

    for i, (cat, filing, deadline, notes) in enumerate(_COMPLIANCE, 4):
        bg = _GRAY_BG if i % 2 == 0 else _WHITE
        _cell(ws, i, 1, cat, bg=bg, bold=True)
        _cell(ws, i, 2, filing, bg=bg)
        _cell(ws, i, 3, deadline, bg=bg, align="center")
        c = _cell(ws, i, 4, notes, bg=bg)
        c.alignment = Alignment(horizontal="left", wrap_text=True)
        ws.row_dimensions[i].height = 15

    ws.freeze_panes = "A4"


# ── AR & AP SHEET ─────────────────────────────────────────────────────────────

def _arap_sheet(wb: Workbook, entities: list[dict]) -> None:
    ws = wb.create_sheet("AR & AP")
    ws.sheet_view.showGridLines = False

    col_widths = [28, 13, 13, 13, 13, 13, 14, 13, 13, 13, 13, 14, 16]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.merge_cells("A1:M1")
    c = ws["A1"]
    c.value = f"AR & AP AGING REPORT — As of {date.today().strftime('%B %d, %Y')}"
    c.font = Font(bold=True, color=_WHITE, size=13, name="Calibri")
    c.fill = PatternFill("solid", fgColor=_DARK_BLUE)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # Group header row
    c = ws["A2"]
    c.value = "Entity"
    c.font = Font(bold=True, color="000000", size=10, name="Calibri")
    c.fill = PatternFill("solid", fgColor=_GRAY_BG)
    c.alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells("B2:G2")
    c = ws["B2"]
    c.value = "ACCOUNTS RECEIVABLE (AR)"
    c.font = Font(bold=True, color=_WHITE, size=10, name="Calibri")
    c.fill = PatternFill("solid", fgColor=_MID_BLUE)
    c.alignment = Alignment(horizontal="center")

    ws.merge_cells("H2:L2")
    c = ws["H2"]
    c.value = "ACCOUNTS PAYABLE (AP)"
    c.font = Font(bold=True, color=_WHITE, size=10, name="Calibri")
    c.fill = PatternFill("solid", fgColor=_BROWN)
    c.alignment = Alignment(horizontal="center")

    ws.merge_cells("M2:M2")
    c = ws["M2"]
    c.value = "NWC"
    c.font = Font(bold=True, color="000000", size=9, name="Calibri")
    c.alignment = Alignment(horizontal="center")

    ws.row_dimensions[2].height = 18

    col_labels = ["Entity", "Current", "1-30", "31-60", "61-90", "90+", "AR Total",
                  "Current", "1-30", "31-60", "60+", "AP Total", "Net WC"]
    col_bgs    = [_GRAY_BG] + [_LT_BLUE_BG]*6 + [_AMBER_BG]*5 + [_GREEN_BG]
    for col_idx, (label, bg) in enumerate(zip(col_labels, col_bgs), 1):
        c = ws.cell(row=3, column=col_idx, value=label)
        c.font = Font(bold=True, size=9, name="Calibri")
        c.fill = PatternFill("solid", fgColor=bg)
        c.alignment = Alignment(horizontal="center", wrap_text=True)
        c.border = _border()
    ws.row_dimensions[3].height = 18

    totals = [0.0] * 12

    for row_idx, ent in enumerate(entities, 4):
        bg = _GRAY_BG if row_idx % 2 == 0 else _WHITE
        ar = ent.get("ar", {})
        ap_agg = ent.get("ap", {})
        ar_t = ent.get("ar_total", 0.0)
        ap_t = ent.get("ap_total", 0.0)
        nwc  = ar_t - ap_t

        row_vals = [
            ar.get("current_amount", 0.0), ar.get("days_1_30", 0.0),
            ar.get("days_31_60", 0.0), ar.get("days_61_90", 0.0),
            ar.get("days_90_plus", 0.0), ar_t,
            ap_agg.get("current_amount", 0.0), ap_agg.get("days_1_30", 0.0),
            ap_agg.get("days_31_60", 0.0), ap_agg.get("days_60_plus", 0.0),
            ap_t, nwc,
        ]

        c = ws.cell(row=row_idx, column=1, value=ent.get("company_name", ""))
        c.font = Font(size=9, name="Calibri", bold=True)
        c.fill = PatternFill("solid", fgColor=bg)
        c.border = _border()

        for col_idx, val in enumerate(row_vals, 2):
            c = ws.cell(row=row_idx, column=col_idx, value=float(val or 0))
            c.font = Font(size=9, name="Calibri")
            c.fill = PatternFill("solid", fgColor=_RED_BG if col_idx == 13 and nwc < 0 else bg)
            c.number_format = '#,##0.00'
            c.alignment = Alignment(horizontal="right")
            c.border = _border()
            totals[col_idx - 2] += float(val or 0)

        ws.row_dimensions[row_idx].height = 15

    # Portfolio total row
    tr = len(entities) + 4
    c = ws.cell(row=tr, column=1, value="PORTFOLIO TOTAL")
    c.font = Font(bold=True, size=9, name="Calibri", color=_WHITE)
    c.fill = PatternFill("solid", fgColor=_DARK_BLUE)
    c.alignment = Alignment(horizontal="left")
    c.border = _border()
    for col_idx, val in enumerate(totals, 2):
        c = ws.cell(row=tr, column=col_idx, value=val)
        c.font = Font(bold=True, size=9, name="Calibri", color=_WHITE)
        c.fill = PatternFill("solid", fgColor=_DARK_BLUE)
        c.number_format = '#,##0.00'
        c.alignment = Alignment(horizontal="right")
        c.border = _border()
    ws.row_dimensions[tr].height = 18
    ws.freeze_panes = "B4"


# ── OPEX SHEET ────────────────────────────────────────────────────────────────

def _opex_sheet(wb: Workbook, expense_breakdown: list[dict]) -> None:
    ws = wb.create_sheet("OpEx Composition")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 18

    ws.merge_cells("A1:C1")
    c = ws["A1"]
    c.value = f"OPERATING EXPENSE COMPOSITION — {date.today().strftime('%B %Y')}"
    c.font = Font(bold=True, color=_WHITE, size=13, name="Calibri")
    c.fill = PatternFill("solid", fgColor=_DARK_BLUE)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    for col, label in enumerate(["Category", "Amount ($)", "% of Total OpEx"], 1):
        _hdr(ws, 2, col, label, bg=_MID_BLUE)
    ws.row_dimensions[2].height = 20

    total_exp = sum(float(e.get("amount", 0)) for e in expense_breakdown)

    for row_idx, item in enumerate(expense_breakdown, 3):
        amount = float(item.get("amount", 0))
        pct    = amount / total_exp if total_exp else 0.0
        bg     = _GRAY_BG if row_idx % 2 == 0 else _WHITE
        _cell(ws, row_idx, 1, item.get("category", "").replace("_", " ").title(), bg=bg)
        c = ws.cell(row=row_idx, column=2, value=amount)
        c.number_format = '#,##0.00'; c.border = _border(); c.alignment = Alignment(horizontal="right")
        c = ws.cell(row=row_idx, column=3, value=pct)
        c.number_format = '0.00%'; c.border = _border(); c.alignment = Alignment(horizontal="right")

    tr = len(expense_breakdown) + 3
    for col_idx, (val, fmt) in enumerate([("TOTAL", None), (total_exp, '#,##0.00'), (1.0 if total_exp else 0.0, '0.00%')], 1):
        c = ws.cell(row=tr, column=col_idx, value=val)
        c.font = Font(bold=True, color=_WHITE, size=10, name="Calibri")
        c.fill = PatternFill("solid", fgColor=_DARK_BLUE)
        c.border = _border()
        if fmt:
            c.number_format = fmt
            c.alignment = Alignment(horizontal="right")


# ── DASHBOARD SHEET ───────────────────────────────────────────────────────────

def _dashboard_sheet(wb: Workbook, portfolio: dict, loans: list[dict]) -> None:
    ws = wb.create_sheet("DASHBOARD", 0)
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 22

    ws.merge_cells("A1:C1")
    c = ws["A1"]
    c.value = f"EstateCFO — CFO Dashboard   |   {date.today().strftime('%B %d, %Y')}"
    c.font = Font(bold=True, color=_WHITE, size=14, name="Calibri")
    c.fill = PatternFill("solid", fgColor=_DARK_BLUE)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 32

    ws.merge_cells("A2:C2")
    ws["A2"].value = "Values sourced from live EstateCFO app — not independently recomputed in this export."
    ws["A2"].font = Font(italic=True, color="888888", size=9, name="Calibri")
    ws["A2"].alignment = Alignment(horizontal="center")
    ws.row_dimensions[2].height = 14

    # ── KPI headline block ────────────────────────────────────────────────────
    kpi_row = 4
    ws.merge_cells(f"A{kpi_row}:C{kpi_row}")
    c = ws[f"A{kpi_row}"]
    c.value = "KEY PERFORMANCE INDICATORS"
    c.font = Font(bold=True, color=_WHITE, size=10, name="Calibri")
    c.fill = PatternFill("solid", fgColor=_MID_BLUE)
    c.alignment = Alignment(horizontal="left")
    c.border = _border()
    ws.row_dimensions[kpi_row].height = 18

    noi           = portfolio.get("noi_this_month", 0.0)
    occupancy     = portfolio.get("occupancy_pct", 0.0)
    gpr           = portfolio.get("gross_potential_rent", 0.0)
    collected     = portfolio.get("collected_this_month", 0.0)
    vacancy_loss  = portfolio.get("vacancy_loss", 0.0)
    total_expense = portfolio.get("total_expense_this_month", 0.0)
    total_units   = portfolio.get("total_units", 0)
    occupied_u    = portfolio.get("occupied_units", 0)
    vacant_u      = portfolio.get("vacant_units", 0)
    billed        = portfolio.get("billed_this_month", 0.0)
    arrears       = portfolio.get("arrears_total", 0.0)

    # Weighted DSCR/LTV
    total_balance = sum(float(l.get("loan_balance_as_of") or 0) for l in loans)
    dscr_loans    = [l for l in loans if l.get("dscr") is not None and l.get("loan_balance_as_of")]
    ltv_loans     = [l for l in loans if l.get("ltv_current") is not None and l.get("loan_balance_as_of")]
    w_dscr, w_ltv = None, None
    if dscr_loans:
        denom = sum(float(l["loan_balance_as_of"]) for l in dscr_loans)
        if denom:
            w_dscr = round(sum(l["dscr"] * float(l["loan_balance_as_of"]) for l in dscr_loans) / denom, 4)
    if ltv_loans:
        denom = sum(float(l["loan_balance_as_of"]) for l in ltv_loans)
        if denom:
            w_ltv = round(sum(l["ltv_current"] * float(l["loan_balance_as_of"]) for l in ltv_loans) / denom, 4)

    opex_ratio = total_expense / collected if collected else None

    kpis = [
        ("Monthly NOI",               noi,          "#,##0.00"),
        ("Monthly Collections",       collected,    "#,##0.00"),
        ("Gross Potential Rent",       gpr,          "#,##0.00"),
        ("Vacancy Loss",               vacancy_loss, "#,##0.00"),
        ("Occupancy Rate",             occupancy,    "0.00%"),
        ("OpEx Ratio (Expense/Collected)", opex_ratio, "0.00%"),
        ("Total Units",                total_units,  "0"),
        ("Occupied / Vacant",          f"{occupied_u} / {vacant_u}", "@"),
        ("Monthly Billed",             billed,       "#,##0.00"),
        ("Arrears (Total Owed)",       arrears,      "#,##0.00"),
        ("Total Outstanding Debt",     total_balance,"#,##0.00"),
        ("Portfolio DSCR (weighted)",  w_dscr,       "0.00"),
        ("Portfolio LTV (weighted)",   w_ltv,        "0.00%"),
    ]

    for i, (label, value, fmt) in enumerate(kpis, kpi_row + 1):
        c = ws.cell(row=i, column=1, value=label)
        c.font = Font(size=10, name="Calibri"); c.border = _border()
        c.alignment = Alignment(horizontal="left")
        c = ws.cell(row=i, column=2, value=value)
        c.font = Font(bold=True, size=10, name="Calibri"); c.border = _border()
        c.alignment = Alignment(horizontal="right")
        c.number_format = fmt
        ws.row_dimensions[i].height = 17

    # ── Lender risk table ─────────────────────────────────────────────────────
    lr_start = kpi_row + len(kpis) + 2
    ws.merge_cells(f"A{lr_start}:C{lr_start}")
    c = ws[f"A{lr_start}"]
    c.value = "LENDER RISK — ALL LOANS (sorted worst DSCR first)"
    c.font = Font(bold=True, color=_WHITE, size=10, name="Calibri")
    c.fill = PatternFill("solid", fgColor=_MID_BLUE)
    c.alignment = Alignment(horizontal="left")
    c.border = _border()
    ws.row_dimensions[lr_start].height = 18

    for col_idx, h in enumerate(["Company / Property", "Outstanding Balance", "DSCR"], 1):
        _hdr(ws, lr_start + 1, col_idx, h, bg="3C3C3C")

    sorted_loans = sorted(loans, key=lambda l: (l.get("dscr") is None, float(l.get("dscr") or 9999)))
    for i, loan in enumerate(sorted_loans, lr_start + 2):
        dscr_v = loan.get("dscr")
        label  = f"{loan.get('company_name', '')} — {loan.get('property_name', '')}"
        c = ws.cell(row=i, column=1, value=label)
        c.border = _border(); c.font = Font(size=9, name="Calibri")
        c = ws.cell(row=i, column=2, value=float(loan.get("loan_balance_as_of") or 0))
        c.number_format = '#,##0.00'; c.border = _border()
        c.alignment = Alignment(horizontal="right"); c.font = Font(size=9, name="Calibri")
        c = ws.cell(row=i, column=3, value=dscr_v)
        c.number_format = '0.00'; c.border = _border()
        c.alignment = Alignment(horizontal="right"); c.font = Font(size=9, name="Calibri")
        if dscr_v is not None:
            c.fill = PatternFill("solid", fgColor=_RED_BG if dscr_v < 1.0 else (_AMBER_BG if dscr_v < 1.25 else _GREEN_BG))
        ws.row_dimensions[i].height = 15


# ── Public builder ─────────────────────────────────────────────────────────────

def build_cfo_dashboard_workbook(
    portfolio: dict,
    entities: list[dict],
    expense_breakdown: list[dict],
    loans: list[dict],
) -> bytes:
    """Build and return the CFO Dashboard workbook as bytes."""
    wb = Workbook()
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    _dashboard_sheet(wb, portfolio, loans)
    _arap_sheet(wb, entities)
    _opex_sheet(wb, expense_breakdown)
    _compliance_sheet(wb)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
