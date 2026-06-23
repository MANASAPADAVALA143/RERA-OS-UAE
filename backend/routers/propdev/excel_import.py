import uuid
from datetime import datetime
from io import BytesIO

from fastapi import APIRouter, Depends, File, HTTPException, UploadFile
from sqlalchemy.orm import Session
import openpyxl

from database import get_db
from middleware.auth import CurrentUser, get_current_user, require_write_access
from models.propdev.company import PropDevCompany
from models.propdev.lot import PropDevLot
from models.propdev.partner import PropDevPartner
from models.propdev.loan import PropDevLoan
from models.propdev.capital_call import PropDevCapitalCall
from models.propdev.expense import PropDevExpense

router = APIRouter(prefix="/api/propdev", tags=["propdev"])


HEADER_KEYWORDS = {'Company', 'Date', 'Week Starting', 'Sl', 'Partner Name', 'Particulars'}

def parse_excel_file(content: bytes) -> dict:
    # data_only=True returns calculated cell values instead of formula strings
    wb = openpyxl.load_workbook(BytesIO(content), data_only=True)
    sheets = {}
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = []
        headers = None
        for row in ws.iter_rows(values_only=True):
            if headers is None:
                # Skip title/description rows; detect header row by known keywords
                row_vals = [str(v or '').strip() for v in row]
                if any(v in HEADER_KEYWORDS for v in row_vals[:4]):
                    headers = [str(h).strip() if h else f"col_{j}" for j, h in enumerate(row)]
                continue
            if any(cell is not None for cell in row):
                rows.append({headers[j]: row[j] for j in range(min(len(headers), len(row)))})
        sheets[sheet_name] = rows
    return sheets


def extract_summary_data(rows: list) -> dict:
    """Extract company-level data from SUMMARY sheet."""
    data = {}
    for row in rows:
        company_name = row.get('Company', '')
        if company_name and company_name.upper() != 'COMPANY' and company_name.upper() != 'PORTFOLIO TOTAL':
            data[company_name] = {
                'property_name': row.get('Property', ''),
                'total_lots': int(row.get('Lots', 0)) if row.get('Lots') else 0,
                'sold_lots': int(row.get('Sold', 0)) if row.get('Sold') else 0,
                'available_lots': int(row.get('Available', 0)) if row.get('Available') else 0,
                'sale_consideration': float(row.get('Sale ($)', 0)) if row.get('Sale ($)') else 0,
                'net_profit': float(row.get('Net Profit ($)', 0)) if row.get('Net Profit ($)') else 0,
            }
    return data


def extract_expense_data(rows: list) -> dict:
    """Extract expenses from Expense Dashboard sheet."""
    expenses = {}
    expense_totals = {}

    for row in rows:
        if not row or not row.get('Company'):
            continue
        company = row.get('Company', '')

        # Store individual expenses
        if company not in expenses:
            expenses[company] = []
        expenses[company].append({
            'expense_date': row.get('Date'),
            'expense_type': row.get('Expense Type', ''),
            'category': row.get('Category', ''),
            'vendor': row.get('Vendor', ''),
            'invoice_no': row.get('Invoice No'),
            'amount': float(row.get('Amount ($)', 0)) if row.get('Amount ($)') else 0,
            'status': row.get('Status', 'Paid'),
        })

        # Calculate totals by category for company P&L
        if company not in expense_totals:
            expense_totals[company] = {
                'hard_cost': 0, 'soft_cost': 0, 'title_charges': 0,
                'other_charges': 0, 'property_tax': 0, 'loan_processing': 0,
                'professional_charges': 0, 'legal_fees': 0, 'interest_on_loan': 0,
            }

        category = row.get('Category', '').lower()
        amount = float(row.get('Amount ($)', 0)) if row.get('Amount ($)') else 0

        if 'hard' in category:
            expense_totals[company]['hard_cost'] += amount
        elif 'soft' in category:
            expense_totals[company]['soft_cost'] += amount
        elif 'title' in category or 'legal' in category.lower():
            expense_totals[company]['title_charges'] += amount
        elif 'tax' in category:
            expense_totals[company]['property_tax'] += amount
        elif 'loan' in category and 'process' in category:
            expense_totals[company]['loan_processing'] += amount
        elif 'professional' in category or 'prof' in category:
            expense_totals[company]['professional_charges'] += amount
        elif 'legal' in category:
            expense_totals[company]['legal_fees'] += amount
        elif 'interest' in category or 'finance' in category:
            expense_totals[company]['interest_on_loan'] += amount
        else:
            expense_totals[company]['other_charges'] += amount

    return expenses, expense_totals


def extract_deal_pl_data(rows: list) -> dict:
    """Extract expense details from Annexure I (Deal P&L) sheet."""
    deal_pl = {}
    for row in rows:
        if not row or not row.get('Company'):
            continue
        company = row.get('Company', '')
        deal_pl[company] = {
            'land_cost': float(row.get('Land Cost ($)', 0)) if row.get('Land Cost ($)') else 0,
            'hard_cost': float(row.get('Hard Cost ($)', 0)) if row.get('Hard Cost ($)') else 0,
            'soft_cost': float(row.get('Soft Cost ($)', 0)) if row.get('Soft Cost ($)') else 0,
            'title_charges': float(row.get('Title ($)', 0)) if row.get('Title ($)') else 0,
            'other_charges': float(row.get('Other ($)', 0)) if row.get('Other ($)') else 0,
            'property_tax': float(row.get('Prop Tax ($)', 0)) if row.get('Prop Tax ($)') else 0,
            'loan_processing': float(row.get('Loan Proc ($)', 0)) if row.get('Loan Proc ($)') else 0,
            'professional_charges': float(row.get('Prof ($)', 0)) if row.get('Prof ($)') else 0,
            'legal_fees': float(row.get('Legal ($)', 0)) if row.get('Legal ($)') else 0,
            'interest_on_loan': float(row.get('Interest ($)', 0)) if row.get('Interest ($)') else 0,
        }
    return deal_pl


def extract_loan_data(rows: list) -> dict:
    """Extract loans from Loan Sheet."""
    loans = {}
    for row in rows:
        if not row or not row.get('Company'):
            continue
        company = row.get('Company', '')
        if company not in loans:
            loans[company] = []

        # Parse EMI Day (might be "15th" format)
        emi_day_str = row.get('EMI Day', '15')
        try:
            emi_day = int(str(emi_day_str).replace('th', '').replace('st', '').replace('nd', '').replace('rd', ''))
        except:
            emi_day = 15

        loans[company].append({
            'bank': row.get('Bank', ''),
            'loan_date': row.get('Loan Date'),
            'account_no': row.get('Acc No'),
            'loan_amount': float(row.get('Loan Amount ($)', 0)) if row.get('Loan Amount ($)') else 0,
            'balance': float(row.get('Balance ($)', 0)) if row.get('Balance ($)') else 0,
            'interest_rate': float(row.get('Rate %', 0)) if row.get('Rate %') else 0,
            'emi': float(row.get('EMI ($)', 0)) if row.get('EMI ($)') else 0,
            'maturity_date': row.get('Maturity'),
            'emi_day': emi_day,
            'lender_name': row.get('Lender'),
            'lender_email': row.get('Email'),
            'lender_phone': row.get('Phone'),
            'bank_account': row.get('Bank Account'),
            'emi_status': row.get('EMI Status', 'Current'),
        })
    return loans


def extract_partner_data(rows: list) -> dict:
    """Extract partners from Annexure II sheet."""
    partners = {}
    for row in rows:
        if not row or not row.get('Company'):
            continue
        company = row.get('Company', '')
        if company not in partners:
            partners[company] = []
        partners[company].append({
            'name': row.get('Partner Name', ''),
            'partner_type': row.get('Type', 'Class A'),
            'share_percent': float(row.get('% Share', 0)) if row.get('% Share') else 0,
            'capital_contributed': float(row.get('Capital (A) ($)', 0)) if row.get('Capital (A) ($)') else 0,
            'distributions_received': float(row.get('Distributed ($)', 0)) if row.get('Distributed ($)') else 0,
            'preferred_return': float(row.get('Pref Return %', 0)) if row.get('Pref Return %') else 0,
            'status': row.get('Status', 'Active'),
        })
    return partners


def extract_capital_calls(rows: list, partners_map: dict) -> dict:
    """Extract capital calls from Capital Calls sheet."""
    calls = {}
    for row in rows:
        if not row or not row.get('Company'):
            continue
        company = row.get('Company', '')
        if company not in calls:
            calls[company] = []

        # Calculate total_due and call amounts
        bal_due = float(row.get('Bal Due ($)', 0)) if row.get('Bal Due ($)') else 0
        call_6mo = float(row.get('Call 6mo ($)', 0)) if row.get('Call 6mo ($)') else 0
        call_incl_dues = float(row.get('Call incl Dues ($)', 0)) if row.get('Call incl Dues ($)') else bal_due + call_6mo
        amount_received = float(row.get('Amount Received ($)', 0)) if row.get('Amount Received ($)') else 0

        calls[company].append({
            'partner_name': row.get('Partner', ''),
            'period': 'Jan–Jun 2025',
            'share_percent': float(row.get('% Share', 0)) if row.get('% Share') else 0,
            'total_call_amount': call_incl_dues,
            'partner_share': call_incl_dues,
            'old_dues': bal_due,
            'total_due': call_incl_dues,
            'amount_received': amount_received,
            'status': row.get('Status', 'Outstanding'),
        })
    return calls


def extract_lot_data(rows: list) -> dict:
    """Extract lot inventory from Lot Inventory sheet."""
    lots = {}
    for row in rows:
        if not row or not row.get('Company'):
            continue
        company = row.get('Company', '')
        if company not in lots:
            lots[company] = []

        # Get size and convert to int if needed
        size_sqft = row.get('Size (sqft)', 0)
        try:
            size_sqft = float(size_sqft) if size_sqft else 0
        except:
            size_sqft = 0

        sale_price = row.get('Sale Price ($)', None)
        try:
            sale_price = float(sale_price) if sale_price else None
        except:
            sale_price = None

        lots[company].append({
            'lot_no': row.get('Lot #', ''),
            'block': row.get('Block', ''),
            'size_sqft': size_sqft,
            'list_price': float(row.get('List Price ($)', 0)) if row.get('List Price ($)') else 0,
            'sale_price': sale_price,
            'status': row.get('Status', 'available').lower(),
            'buyer_name': row.get('Buyer Name'),
            'contract_date': row.get('Contract Date'),
            'close_date': row.get('Close Date'),
        })
    return lots


@router.post("/import-excel")
async def import_excel(
    file: UploadFile = File(...),
    current_user: CurrentUser = Depends(require_write_access()),
    db: Session = Depends(get_db),
):
    try:
        content = await file.read()
        sheets = parse_excel_file(content)

        # Extract data from different sheets
        summary_data = extract_summary_data(sheets.get('SUMMARY', []))
        expense_data, expense_totals = extract_expense_data(sheets.get('Expense Dashboard', []))
        deal_pl_data = extract_deal_pl_data(sheets.get('Annexure I', []))
        loan_data = extract_loan_data(sheets.get('Loan Sheet', []))
        partner_data = extract_partner_data(sheets.get('Annexure II', []))
        capital_call_data = extract_capital_calls(sheets.get('Capital Calls', []), partner_data)
        lot_data = extract_lot_data(sheets.get('Lot Inventory', []))
        cash_data = {}

        # Create companies and related data
        created_companies = []
        total_companies = len(summary_data)
        print(f"[IMPORT] Processing {total_companies} companies from Excel...")

        for idx, (company_name, summary) in enumerate(summary_data.items(), 1):
            print(f"[IMPORT] [{idx}/{total_companies}] Processing {company_name}...")

            try:
                # Get expense data from Deal P&L sheet (preferred) or Expense Dashboard
                pl_data = deal_pl_data.get(company_name, {})
                exp_totals = expense_totals.get(company_name, {})

                land_cost = pl_data.get('land_cost', 0) or 0
                hard_cost = pl_data.get('hard_cost', 0) or exp_totals.get('hard_cost', 0)
                soft_cost = pl_data.get('soft_cost', 0) or exp_totals.get('soft_cost', 0)
                title_charges = pl_data.get('title_charges', 0) or exp_totals.get('title_charges', 0)
                other_charges = pl_data.get('other_charges', 0) or exp_totals.get('other_charges', 0)
                property_tax = pl_data.get('property_tax', 0) or exp_totals.get('property_tax', 0)
                loan_processing = pl_data.get('loan_processing', 0) or exp_totals.get('loan_processing', 0)
                professional_charges = pl_data.get('professional_charges', 0) or exp_totals.get('professional_charges', 0)
                legal_fees = pl_data.get('legal_fees', 0) or exp_totals.get('legal_fees', 0)
                interest_on_loan = pl_data.get('interest_on_loan', 0) or exp_totals.get('interest_on_loan', 0)

                # Find existing company (matched from registry) or create new
                company = db.query(PropDevCompany).filter(
                    PropDevCompany.tenant_id == current_user.tenant_id,
                    PropDevCompany.name == company_name,
                ).first()

                if company:
                    # Update existing registry entry with Excel data
                    company.property_name = summary.get('property_name', '') or company.property_name
                    company.total_lots = summary.get('total_lots', 0) or company.total_lots
                    company.sale_consideration = summary.get('sale_consideration', 0)
                    company.land_cost = land_cost
                    company.hard_cost = hard_cost
                    company.soft_cost = soft_cost
                    company.title_charges = title_charges
                    company.other_charges = other_charges
                    company.property_tax = property_tax
                    company.loan_processing = loan_processing
                    company.professional_charges = professional_charges
                    company.legal_fees = legal_fees
                    company.interest_on_loan = interest_on_loan
                    # Clear old related records before re-importing
                    for lot in company.lots: db.delete(lot)
                    for p in company.partners: db.delete(p)
                    for ln in company.loans: db.delete(ln)
                    for cc in company.capital_calls: db.delete(cc)
                    for exp in company.expenses: db.delete(exp)
                    db.flush()
                else:
                    # Create new company (not in registry yet)
                    company = PropDevCompany(
                        tenant_id=current_user.tenant_id,
                        name=company_name,
                        property_name=summary.get('property_name', ''),
                        address='',
                        total_lots=summary.get('total_lots', 0),
                        sale_consideration=summary.get('sale_consideration', 0),
                        land_cost=land_cost,
                        hard_cost=hard_cost,
                        soft_cost=soft_cost,
                        title_charges=title_charges,
                        other_charges=other_charges,
                        property_tax=property_tax,
                        loan_processing=loan_processing,
                        professional_charges=professional_charges,
                        legal_fees=legal_fees,
                        interest_on_loan=interest_on_loan,
                        cash_available=0,
                    )
                    db.add(company)
                db.flush()

                # Add lots
                for lot in lot_data.get(company_name, []):
                    db.add(PropDevLot(
                        tenant_id=current_user.tenant_id,
                        company_id=company.id,
                        lot_no=lot.get('lot_no', ''),
                        block=lot.get('block', ''),
                        size_sqft=lot.get('size_sqft', 0),
                        list_price=lot.get('list_price', 0),
                        sale_price=lot.get('sale_price'),
                        status=lot.get('status', 'available'),
                        buyer_name=lot.get('buyer_name'),
                        contract_date=lot.get('contract_date'),
                        close_date=lot.get('close_date'),
                    ))

                # Add partners
                partners_by_company = {}
                for partner in partner_data.get(company_name, []):
                    p = PropDevPartner(
                        tenant_id=current_user.tenant_id,
                        company_id=company.id,
                        name=partner.get('name', ''),
                        partner_type=partner.get('partner_type', 'Class A'),
                        share_percent=partner.get('share_percent', 0),
                        capital_contributed=partner.get('capital_contributed', 0),
                        distributions_received=partner.get('distributions_received', 0),
                        preferred_return=partner.get('preferred_return', 0),
                        status=partner.get('status', 'Active'),
                    )
                    db.add(p)
                    db.flush()
                    partners_by_company[partner.get('name', '')] = p.id

                # Add loans
                for loan in loan_data.get(company_name, []):
                    db.add(PropDevLoan(
                        tenant_id=current_user.tenant_id,
                        company_id=company.id,
                        bank=loan.get('bank', ''),
                        loan_date=loan.get('loan_date'),
                        account_no=loan.get('account_no'),
                        loan_amount=loan.get('loan_amount', 0),
                        balance=loan.get('balance', 0),
                        interest_rate=loan.get('interest_rate', 0),
                        emi=loan.get('emi', 0),
                        maturity_date=loan.get('maturity_date'),
                        emi_day=loan.get('emi_day', 15),
                        lender_name=loan.get('lender_name'),
                        lender_email=loan.get('lender_email'),
                        lender_phone=loan.get('lender_phone'),
                        bank_account=loan.get('bank_account'),
                        emi_status=loan.get('emi_status', 'Current'),
                    ))

                # Add capital calls
                for call in capital_call_data.get(company_name, []):
                    partner_id = partners_by_company.get(call.get('partner_name', ''))
                    if partner_id:
                        db.add(PropDevCapitalCall(
                            tenant_id=current_user.tenant_id,
                            company_id=company.id,
                            partner_id=partner_id,
                            period=call.get('period', ''),
                            share_percent=call.get('share_percent', 0),
                            total_call_amount=call.get('total_call_amount', 0),
                            partner_share=call.get('partner_share', 0),
                            old_dues=call.get('old_dues', 0),
                            total_due=call.get('total_due', 0),
                            amount_received=call.get('amount_received', 0),
                            status=call.get('status', 'Outstanding'),
                        ))

                # Add expenses
                for expense in expense_data.get(company_name, []):
                    db.add(PropDevExpense(
                        tenant_id=current_user.tenant_id,
                        company_id=company.id,
                        expense_date=expense.get('expense_date'),
                        expense_type=expense.get('expense_type', ''),
                        category=expense.get('category', ''),
                        vendor=expense.get('vendor', ''),
                        invoice_no=expense.get('invoice_no'),
                        amount=expense.get('amount', 0),
                        status=expense.get('status', 'Paid'),
                    ))

                db.commit()
                print(f"[IMPORT] Successfully created {company_name}")

                created_companies.append({
                    'id': str(company.id),
                    'name': company_name,
                    'property': summary.get('property_name', ''),
                    'total_lots': summary.get('total_lots', 0),
                })

            except Exception as e:
                print(f"[ERROR] Failed to create {company_name}: {str(e)}")
                db.rollback()
                continue

        print(f"[IMPORT] Complete! Created {len(created_companies)} companies")

        return {
            'status': 'success',
            'message': f'Successfully imported {len(created_companies)} companies',
            'companies_count': len(created_companies),
            'companies': created_companies,
        }

    except Exception as e:
        print(f"[IMPORT] Fatal error: {str(e)}")
        db.rollback()
        raise HTTPException(status_code=400, detail=f"Import failed: {str(e)}")
