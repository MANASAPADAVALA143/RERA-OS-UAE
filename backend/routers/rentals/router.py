"""Rentals module router — all endpoints under /api/rentals/"""
from __future__ import annotations

import csv
import io
import os
import uuid
from collections import defaultdict
from datetime import date, datetime

from fastapi import APIRouter, Depends, File, Form, HTTPException, Query, UploadFile
from fastapi.responses import StreamingResponse
from sqlalchemy import func, or_
from sqlalchemy.orm import Session

from database import SessionLocal
from middleware.auth import CurrentUser, get_current_user, require_write_access
from models.rentals.models import (
    RentalCollection,
    RentalCompany,
    RentalExpense,
    RentalExpenseCategory,
    RentalInvoice,
    RentalLease,
    RentalOwnership,
    RentalPartnerRole,
    RentalProp,
    RentalTenant,
    RentalUnit,
)
from models.rentals.maintenance import MaintenanceRequest
from models.rentals.unit_inspection import UnitInspection, UnitInspectionPhoto, UnitInspectionChecklistItem
from services.rental_calculations import (
    arrears_aging,
    company_summary,
    days_vacant,
    distribute_to_partners,
    income_trend,
    lease_expiry_pipeline,
    unit_arrears,
)

router = APIRouter(prefix="/api/rentals", tags=["rentals"])

MONTH_OPTIONS = [
    'Jan-2026', 'Feb-2026', 'Mar-2026', 'Apr-2026', 'May-2026', 'Jun-2026',
    'Jul-2026', 'Aug-2026', 'Sep-2026', 'Oct-2026', 'Nov-2026', 'Dec-2026',
]


def get_db():
    db = SessionLocal()
    try:
        yield db
    finally:
        db.close()


def _flatten_pl_items(items: list) -> list:
    """Recursively return leaf P&L items (nodes whose children list is empty)."""
    result = []
    for item in (items or []):
        children = item.get("children") or []
        if children:
            result.extend(_flatten_pl_items(children))
        else:
            result.append(item)
    return result


# ── P&L regex patterns (compiled once at import time) ────────────────────────

import re as _re
from models.rentals.models import RentalFinancialUpload as _RFU

_INCOME_RE = _re.compile(
    r"rental\s+income|rent\s+income|rent\s*[-–]|other\s+income", _re.IGNORECASE
)
_EXP_SKIP = _re.compile(
    r"^(total|subtotal|net\s|gross\s|\bincome\b|revenue|rental\s+income"
    r"|rent\s+income|rent\s*-|operating\s+income|net\s+income|net\s+loss)",
    _re.IGNORECASE,
)
_ONE_TIME = _re.compile(
    r"sec\s*481|481\s*\(a\)|accounting\s*method\s*adjustment", _re.IGNORECASE
)


def _apply_collected_fallback(
    summ: dict,
    co,
    month_abbrev: str,
    db,
    tid,
) -> None:
    """
    Apply Excel-sync → P&L fallback for collected/NOI in-place.

    Called by list_companies, company_dashboard, and get_portfolio_summary
    so the same two-source priority is enforced everywhere:
      Source B: Excel-synced monthly_rent_data / collected_this_month column
      Source C: Latest P&L upload for the company
    """
    # ── Source B: Excel sync ──────────────────────────────────────────────────
    if summ["collected_this_month"] == 0.0:
        monthly_data: dict = co.monthly_rent_data or {}
        synced = float(monthly_data.get(month_abbrev, 0.0))
        if synced == 0.0 and co.last_sync_month and co.collected_this_month:
            try:
                if (
                    datetime.strptime(co.last_sync_month, "%b-%Y").strftime("%Y-%m")
                    == datetime.strptime(month_abbrev, "%b-%Y").strftime("%Y-%m")
                ):
                    synced = float(co.collected_this_month)
            except ValueError:
                pass
        if synced > 0.0:
            summ["collected_this_month"] = synced
            summ["noi_this_month"] = round(synced - summ["total_expense_this_month"], 2)
            summ["collected_source"] = "excel_sync"

    # ── Source C: P&L upload ──────────────────────────────────────────────────
    if summ["collected_this_month"] == 0.0:
        upload = (
            db.query(_RFU)
            .filter(_RFU.tenant_id == tid, _RFU.company_id == co.id)
            .order_by(_RFU.uploaded_at.desc())
            .first()
        )
        if upload and upload.pl_data:
            mk_space = month_abbrev.replace("-", " ")
            mk_dash  = month_abbrev
            pl_income  = 0.0
            pl_expense = 0.0
            for item in _flatten_pl_items(upload.pl_data):
                if item.get("isSectionHeader") or item.get("isTotal"):
                    continue
                label = str(item.get("label", ""))
                mv    = item.get("monthlyValues") or {}
                val   = abs(float(mv.get(mk_space, mv.get(mk_dash, 0)) or 0))
                if val == 0:
                    continue
                if _INCOME_RE.match(label):
                    pl_income += val
                elif not _EXP_SKIP.match(label.strip()) and not _ONE_TIME.search(label):
                    pl_expense += val
            if pl_income > 0.0:
                summ["collected_this_month"] = pl_income
                summ["collected_source"]     = "pl_fallback"
                if pl_expense > 0.0 and summ["total_expense_this_month"] == 0.0:
                    summ["total_expense_this_month"] = pl_expense
                summ["noi_this_month"] = round(
                    summ["collected_this_month"] - summ["total_expense_this_month"], 2
                )
                if summ["billed_this_month"] == 0.0 and summ["gross_potential_rent"] > 0.0:
                    summ["billed_this_month"] = summ["gross_potential_rent"]
                summ["arrears_total"] = round(
                    max(0.0, summ["billed_this_month"] - summ["collected_this_month"]), 2
                )

    # ── Gross potential / vacancy loss column fallbacks ───────────────────────
    if summ["gross_potential_rent"] == 0.0 and co.gross_potential_rent:
        summ["gross_potential_rent"] = float(co.gross_potential_rent)
    if summ["vacancy_loss"] == 0.0 and co.vacancy_loss:
        summ["vacancy_loss"] = float(co.vacancy_loss)


def _pl_expense_breakdown(co, month_abbrev: str, db, tid) -> list[dict]:
    """
    Build expense-by-category list from the P&L upload for one company/month.
    Returns [] if no upload or no expense items for that month.
    """
    upload = (
        db.query(_RFU)
        .filter(_RFU.tenant_id == tid, _RFU.company_id == co.id)
        .order_by(_RFU.uploaded_at.desc())
        .first()
    )
    if not (upload and upload.pl_data):
        return []
    mk_space = month_abbrev.replace("-", " ")
    mk_dash  = month_abbrev
    exp_by_cat: dict[str, float] = defaultdict(float)
    for item in _flatten_pl_items(upload.pl_data):
        if item.get("isSectionHeader") or item.get("isTotal"):
            continue
        label = str(item.get("label", ""))
        if _INCOME_RE.match(label) or _EXP_SKIP.match(label.strip()) or _ONE_TIME.search(label):
            continue
        mv  = item.get("monthlyValues") or {}
        val = abs(float(mv.get(mk_space, mv.get(mk_dash, 0)) or 0))
        if val > 0:
            exp_by_cat[label] += val
    return [{"category": k, "amount": round(v, 2)} for k, v in exp_by_cat.items()]


# ── helpers ──────────────────────────────────────────────────────────────────

def _inv_dict(inv: RentalInvoice) -> dict:
    return {
        "id": str(inv.id),
        "unit_id": str(inv.unit_id),
        "lease_id": str(inv.lease_id),
        "billing_period": inv.billing_period.isoformat() if inv.billing_period else None,
        "amount_billed": float(inv.amount_billed),
        "collections": [
            {
                "id": str(c.id),
                "amount_collected": float(c.amount_collected),
                "collected_date": c.collected_date.isoformat() if c.collected_date else None,
            }
            for c in inv.collections
        ],
    }


def _unit_dict(u: RentalUnit, inv_list: list[dict] | None = None, today: date | None = None) -> dict:
    today = today or date.today()
    arrears = unit_arrears(inv_list or [])
    tenant = next((t for t in u.r_tenants if t.is_current), None)
    active_lease = next(
        (l for l in sorted(u.leases, key=lambda x: x.lease_end, reverse=True) if l.status in ("active", "notice_given")),
        None,
    )
    return {
        "id": str(u.id),
        "company_id": str(u.company_id),
        "property_id": str(u.property_id),
        "company_name": u.company.company_name if u.company else "",
        "property_name": u.property.property_name if u.property else "",
        "unit_number": u.unit_number,
        "status": u.status.value,
        "monthly_rent": float(u.monthly_rent),
        "status_changed_at": u.status_changed_at.isoformat() if u.status_changed_at else None,
        "days_vacant": days_vacant(u.status.value, u.status_changed_at, today),
        "tenant_name": tenant.tenant_name if tenant else None,
        "tenant_email": tenant.tenant_email if tenant else None,
        "lease_end": active_lease.lease_end.isoformat() if active_lease else None,
        "lease_status": active_lease.status.value if active_lease else None,
        "arrears": arrears,
        "rent_history": u.rent_history or {},
        "vacancy_loss": float(u.vacancy_loss) if u.vacancy_loss is not None else None,
    }


def _lease_dict(l: RentalLease) -> dict:
    today = date.today()
    days_left = (l.lease_end - today).days if l.lease_end else None
    return {
        "id": str(l.id),
        "unit_id": str(l.unit_id),
        "unit_number": l.unit.unit_number if l.unit else "",
        "company_name": l.unit.company.company_name if l.unit and l.unit.company else "",
        "company_id": str(l.unit.company_id) if l.unit else "",
        "property_name": l.unit.property.property_name if l.unit and l.unit.property else "",
        "tenant_name": l.rtenant.tenant_name if l.rtenant else None,
        "lease_start": l.lease_start.isoformat() if l.lease_start else None,
        "lease_end": l.lease_end.isoformat() if l.lease_end else None,
        "days_until_expiry": days_left,
        "status": l.status.value,
        "escalation_pct_annual": float(l.escalation_pct_annual) if l.escalation_pct_annual else None,
        "deposit_amount": float(l.deposit_amount) if l.deposit_amount else None,
        "notice_period_days": l.notice_period_days,
        "lock_in_end_date": l.lock_in_end_date.isoformat() if l.lock_in_end_date else None,
    }


def _expense_dict(e: RentalExpense) -> dict:
    return {
        "id": str(e.id),
        "company_id": str(e.company_id),
        "property_id": str(e.property_id),
        "company_name": e.company.company_name if e.company else "",
        "property_name": e.property.property_name if e.property else "",
        "expense_date": e.expense_date.isoformat() if e.expense_date else None,
        "category": e.category.value,
        "amount": float(e.amount),
        "description": e.description,
    }


def _load_company_data(company_id: uuid.UUID, tid: uuid.UUID, db: Session) -> tuple:
    """Returns (units, all_invoices_with_collections, expenses) for a company."""
    units = (
        db.query(RentalUnit)
        .filter(RentalUnit.tenant_id == tid, RentalUnit.company_id == company_id)
        .all()
    )
    unit_ids = [u.id for u in units]
    invoices = (
        db.query(RentalInvoice)
        .filter(RentalInvoice.tenant_id == tid, RentalInvoice.unit_id.in_(unit_ids))
        .all()
    ) if unit_ids else []
    inv_dicts = [_inv_dict(i) for i in invoices]
    expenses = (
        db.query(RentalExpense)
        .filter(RentalExpense.tenant_id == tid, RentalExpense.company_id == company_id)
        .all()
    )
    exp_dicts = [_expense_dict(e) for e in expenses]
    return units, inv_dicts, exp_dicts


def _registry_unit_counts(units: list, month: str | None = None) -> tuple[int, int]:
    """Occupied / total — when month given, use that month's rent_history."""
    total = len(units)
    if month:
        occupied = sum(
            1 for u in units
            if float((u.rent_history or {}).get(month, 0) or 0) > 0
        )
    else:
        occupied = sum(1 for u in units if u.status == "occupied")
    return occupied, total


def _reconcile_unit_status_for_month(units: list, month: str) -> None:
    """Align stored status/monthly_rent with the sync target month in rent_history."""
    for u in units:
        h = u.rent_history or {}
        if month not in h:
            continue
        amt = float(h.get(month, 0) or 0)
        u.status = "vacant" if amt == 0 else "occupied"
        if amt > 0:
            u.monthly_rent = amt


def _sync_unit_from_preview(ex, unit: dict, target_month: str | None = None) -> bool:
    """Apply Excel preview onto an existing registry unit (always overwrite stale history)."""
    changed = False
    hist = unit.get("history") or {}
    if hist:
        ex.rent_history = hist
        changed = True
    status = unit.get("status")
    if status:
        ex.status = status
        changed = True
    monthly = float(unit.get("monthly_rent", 0))
    if monthly > 0:
        ex.monthly_rent = monthly
        changed = True
    if target_month and hist and target_month in hist:
        amt = float(hist.get(target_month, 0) or 0)
        ex.status = "vacant" if amt == 0 else "occupied"
        if amt > 0:
            ex.monthly_rent = amt
        elif amt == 0:
            ex.vacancy_loss = 0.0
        changed = True
    return changed


def _apply_registry_unit_counts(company, units: list, month: str | None = None) -> None:
    """Prefer registry unit rows over Excel physical-unit inflation for occupancy."""
    if units:
        occupied, total = _registry_unit_counts(units, month)
        company.occupied_units = occupied
        company.total_units = total


def _registry_monthly_totals(units: list) -> dict[str, float]:
    """Sum rent_history across registry units — excludes Excel summary rows."""
    totals: dict[str, float] = {}
    for u in units:
        for m, v in (u.rent_history or {}).items():
            totals[m] = totals.get(m, 0.0) + float(v or 0)
    return {m: round(v, 2) for m, v in totals.items()}


def _registry_vacancy_loss(units: list) -> float:
    return round(
        sum(float(u.vacancy_loss or 0) for u in units if u.status == "vacant"),
        2,
    )


def _infer_sync_month(units: list, prefer: str | None = None) -> str | None:
    """Pick Mon-YYYY from rent_history; fall back to prefer when only monthly_rent exists."""
    months_with_data: set[str] = set()
    for u in units:
        for m, v in (u.rent_history or {}).items():
            if float(v or 0) > 0:
                months_with_data.add(m)
    if prefer and prefer in months_with_data:
        return prefer
    for m in reversed(MONTH_OPTIONS):
        if m in months_with_data:
            return m
    if prefer and any(
        float(u.monthly_rent or 0) > 0 or (u.status or "").lower() == "occupied"
        for u in units
    ):
        return prefer
    return None


def _registry_collected_for_month(units: list, month: str) -> float:
    """Sum collected rent for a month — rent_history first, monthly_rent fallback."""
    total = 0.0
    for u in units:
        h = u.rent_history or {}
        if month in h:
            total += float(h.get(month, 0) or 0)
        elif (u.status or "").lower() == "occupied":
            total += float(u.monthly_rent or 0)
    return round(total, 2)


def _heal_company_sync_fields(company, units: list, prefer_month: str | None = None) -> bool:
    """Backfill last_sync_month / collected from registry when Load Portfolio skipped sync."""
    if not units:
        return False
    changed = False
    month = company.last_sync_month or _infer_sync_month(units, prefer_month)
    if month and not company.last_sync_month:
        company.last_sync_month = month
        changed = True
    if not month:
        return changed
    try:
        if any(u.rent_history for u in units):
            _reconcile_unit_status_for_month(units, month)
        occ, total = _registry_unit_counts(units, month)
        if company.total_units != total or company.occupied_units != occ:
            company.occupied_units = occ
            company.total_units = total
            changed = True
        reg_totals = _registry_monthly_totals(units)
        if reg_totals:
            if company.monthly_rent_data != reg_totals:
                company.monthly_rent_data = reg_totals
                changed = True
            gpr = max(reg_totals.values())
            if company.gross_potential_rent != gpr:
                company.gross_potential_rent = gpr
                changed = True
        reg_collected = _registry_collected_for_month(units, month)
        reg_vac = _registry_vacancy_loss(units)
        if company.collected_this_month != reg_collected:
            company.collected_this_month = reg_collected
            changed = True
        if company.vacancy_loss != reg_vac:
            company.vacancy_loss = reg_vac
            changed = True
    except Exception as exc:
        import logging
        logging.getLogger(__name__).warning(
            "heal_company_sync_fields failed for %s: %s",
            getattr(company, "company_name", "?"), exc,
        )
    return changed


def _company_sync_collected(co, month: str | None = None) -> float | None:
    """Prefer monthly_rent_data rollup — stays aligned with registry unit sums."""
    mrd = co.monthly_rent_data or {}
    m = month or co.last_sync_month
    if m and m in mrd:
        return float(mrd[m])
    if co.collected_this_month is not None:
        return float(co.collected_this_month)
    return None


def _apply_registry_financials(company, units: list, target_month: str | None = None) -> None:
    """
    Collected / monthly totals from registry unit rows only.
    Prevents Excel footer rows (Rents, Security Deposit) from inflating company totals.
    """
    if not units:
        return
    month = target_month or company.last_sync_month
    totals = _registry_monthly_totals(units)
    if totals:
        company.monthly_rent_data = totals
        company.gross_potential_rent = max(totals.values())
    if month:
        company.collected_this_month = _registry_collected_for_month(units, month)
    company.vacancy_loss = _registry_vacancy_loss(units)


# ── portfolio summary ─────────────────────────────────────────────────────────

@router.get("/portfolio-summary")
def get_portfolio_summary(
    month: str = Query(None),  # YYYY-MM or Mon-YYYY (e.g. "2026-06" or "Jun-2026")
    current_user: CurrentUser = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    import logging
    from database import engine as _engine
    _log = logging.getLogger(__name__)

    tid = current_user.tenant_id
    today = date.today()

    # Resolve selected month to YYYY-MM
    selected_month = today.strftime("%Y-%m")
    if month:
        m = month.strip()
        if len(m) == 7 and m[4] == "-" and m[:4].isdigit():
            selected_month = m
        else:
            try:
                selected_month = datetime.strptime(m, "%b-%Y").strftime("%Y-%m")
            except ValueError:
                pass

    # Mon-YYYY form used as key in monthly_rent_data JSON
    try:
        month_abbrev = datetime.strptime(selected_month, "%Y-%m").strftime("%b-%Y")
    except ValueError:
        month_abbrev = ""

    try:
        companies = db.query(RentalCompany).filter(RentalCompany.tenant_id == tid).all()
    except Exception as exc:
        _log.warning("portfolio-summary query failed (%s) — applying patches and retrying", exc)
        try:
            from services.schema_patches import apply_schema_patches
            apply_schema_patches(_engine)
            db.expire_all()
            companies = db.query(RentalCompany).filter(RentalCompany.tenant_id == tid).all()
        except Exception:
            companies = []

    all_units_dicts: list[dict] = []
    all_inv_dicts: list[dict] = []
    all_exp_dicts: list[dict] = []
    by_company: list[dict] = []

    all_leases_raw: list[RentalLease] = []

    for co in companies:
        units, inv_dicts, exp_dicts = _load_company_data(co.id, tid, db)
        inv_by_unit: dict[str, list[dict]] = defaultdict(list)
        for inv in inv_dicts:
            inv_by_unit[inv["unit_id"]].append(inv)

        unit_dicts = [_unit_dict(u, inv_by_unit.get(str(u.id), []), today) for u in units]
        summ = company_summary(unit_dicts, inv_dicts, exp_dicts, today, cur_month=selected_month)

        # Apply Excel-sync → P&L fallback (shared helper — same logic as company_dashboard)
        _apply_collected_fallback(summ, co, month_abbrev, db, tid)
        # ─────────────────────────────────────────────────────────────────────

        all_units_dicts.extend(unit_dicts)
        all_inv_dicts.extend(inv_dicts)
        all_exp_dicts.extend(exp_dicts)

        leases = db.query(RentalLease).filter(
            RentalLease.tenant_id == tid,
            RentalLease.unit_id.in_([u.id for u in units]),
        ).all()
        all_leases_raw.extend(leases)

        by_company.append({
            "company_id": str(co.id),
            "company_name": co.company_name,
            **summ,
        })

    portfolio = company_summary(all_units_dicts, all_inv_dicts, all_exp_dicts, today, cur_month=selected_month)

    # Roll up corrected per-company values into portfolio totals
    if portfolio["collected_this_month"] == 0.0:
        total_col = sum(c["collected_this_month"] for c in by_company)
        if total_col > 0:
            portfolio["collected_this_month"] = total_col
    if portfolio["total_expense_this_month"] == 0.0:
        total_exp = sum(c["total_expense_this_month"] for c in by_company)
        if total_exp > 0:
            portfolio["total_expense_this_month"] = total_exp
    # Recompute NOI from rolled-up collected and expenses
    if portfolio["noi_this_month"] == 0.0 and portfolio["collected_this_month"] > 0.0:
        portfolio["noi_this_month"] = round(
            portfolio["collected_this_month"] - portfolio["total_expense_this_month"], 2
        )
    if portfolio["gross_potential_rent"] == 0.0:
        portfolio["gross_potential_rent"] = round(sum(c["gross_potential_rent"] for c in by_company), 2)
    if portfolio["vacancy_loss"] == 0.0:
        portfolio["vacancy_loss"] = round(sum(c["vacancy_loss"] for c in by_company), 2)
    # Roll up arrears when invoice-level arrears is zero (P&L fallback companies)
    if portfolio["arrears_total"] == 0.0:
        total_arr = sum(c.get("arrears_total", 0.0) for c in by_company)
        if total_arr > 0:
            portfolio["arrears_total"] = round(total_arr, 2)
    # billed rollup — needed for Tile 1 sub-label
    if portfolio["billed_this_month"] == 0.0:
        total_billed = sum(c.get("billed_this_month", 0.0) for c in by_company)
        if total_billed > 0:
            portfolio["billed_this_month"] = round(total_billed, 2)

    # Portfolio-level data-source flag for frontend badge
    if any(c.get("collected_source") == "pl_fallback" for c in by_company):
        portfolio_collected_source = "pl_fallback"
    elif any(c.get("collected_source") == "excel_sync" for c in by_company):
        portfolio_collected_source = "excel_sync"
    else:
        portfolio_collected_source = "rent_receivable"

    aging = arrears_aging(all_inv_dicts, today)

    # If invoice-based aging is empty, fall back to latest QB AR snapshot
    if all(v == 0.0 for v in aging.values()):
        try:
            from models.rentals.qb_ar_aging import QBArAgingRow as _QBRow, QBArAgingSnapshot as _QBSnap
            latest_snap = (
                db.query(_QBSnap)
                .filter(_QBSnap.tenant_id == tid)
                .order_by(_QBSnap.as_of_date.desc())
                .first()
            )
            if latest_snap:
                qb_rows = db.query(_QBRow).filter(_QBRow.snapshot_id == latest_snap.id).all()
                aging = {
                    "current":  round(sum(float(r.current_amount) for r in qb_rows), 2),
                    "1_30":     round(sum(float(r.days_1_30)      for r in qb_rows), 2),
                    "31_60":    round(sum(float(r.days_31_60)     for r in qb_rows), 2),
                    "61_90":    round(sum(float(r.days_61_90)     for r in qb_rows), 2),
                    "90_plus":  round(sum(float(r.days_91_plus)   for r in qb_rows), 2),
                }
        except Exception:
            pass  # never break the portfolio summary over a QB import failure

    trend = income_trend(all_inv_dicts, all_exp_dicts, months=6)
    expiry = lease_expiry_pipeline([_lease_dict(l) for l in all_leases_raw], today, window_days=90)

    # Partner share payable (limited/silent partners' cut of NOI)
    ownership_rows = db.query(RentalOwnership).filter(RentalOwnership.tenant_id == tid).all()
    has_partner_data = len(ownership_rows) > 0
    noi_val = portfolio["noi_this_month"]
    partner_share_payable = round(sum(
        noi_val * float(o.ownership_pct)
        for o in ownership_rows
        if o.role.value in ("limited_partner", "silent_partner")
    ), 2) if noi_val > 0 else 0.0

    # attention_now
    attention: list[dict] = []
    vacant_count = sum(1 for u in all_units_dicts if u["status"] == "vacant")
    if vacant_count:
        co_with_vacant = len({u["company_id"] for u in all_units_dicts if u["status"] == "vacant"})
        attention.append({"type": "vacant", "message": f"{vacant_count} vacant unit(s) across {co_with_vacant} company(ies)", "severity": "warning"})

    expiring_60 = [l for l in expiry if l.get("days_until_expiry", 999) <= 60]
    if expiring_60:
        attention.append({"type": "lease_expiry", "message": f"{len(expiring_60)} lease(s) expire within 60 days", "severity": "attention"})

    aging_31_plus = sum(v for k, v in aging.items() if k != "current")
    if aging_31_plus > 0:
        attention.append({"type": "arrears_aging", "message": f"${aging_31_plus:,.0f} in arrears older than 30 days", "severity": "warning"})

    low_occ = [c for c in by_company if c["occupancy_pct"] < 0.75]
    if low_occ:
        attention.append({"type": "low_occupancy", "message": f"{len(low_occ)} company(ies) below 75% occupancy", "severity": "attention"})

    return {
        **portfolio,
        "partner_share_payable": partner_share_payable,
        "has_partner_data": has_partner_data,
        "collected_source": portfolio_collected_source,
        "by_company": by_company,
        "arrears_aging": aging,
        "income_trend": trend,
        "lease_expiry_pipeline": expiry,
        "attention_now": attention,
    }


# ── companies ─────────────────────────────────────────────────────────────────

@router.get("/companies/count")
def count_companies(
    current_user: CurrentUser = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    n = db.query(RentalCompany).filter(RentalCompany.tenant_id == current_user.tenant_id).count()
    return {"count": n}


@router.get("/companies")
def list_companies(
    fmt: str = Query(None, alias="format"),
    month: str = Query(None, description="YYYY-MM; defaults to current month"),
    current_user: CurrentUser = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    import logging
    from database import engine as _engine
    _log = logging.getLogger(__name__)

    tid = current_user.tenant_id
    today = date.today()

    selected_month = today.strftime("%Y-%m")
    if month:
        m = month.strip()
        if len(m) == 7 and m[4] == "-" and m[:4].isdigit():
            selected_month = m
        else:
            try:
                selected_month = datetime.strptime(m, "%b-%Y").strftime("%Y-%m")
            except ValueError:
                pass

    # Self-healing: if the query fails (missing columns), apply schema patches
    # and retry once before giving up.
    try:
        companies = db.query(RentalCompany).filter(RentalCompany.tenant_id == tid).all()
    except Exception as exc:
        _log.warning("list_companies initial query failed (%s) — applying schema patches and retrying", exc)
        try:
            from services.schema_patches import apply_schema_patches
            apply_schema_patches(_engine)
            db.expire_all()
            companies = db.query(RentalCompany).filter(RentalCompany.tenant_id == tid).all()
        except Exception as exc2:
            _log.error("list_companies retry also failed: %s", exc2)
            raise HTTPException(
                status_code=500,
                detail="Could not load companies from database. Check server logs.",
            ) from exc2
    month_abbrev = datetime.strptime(selected_month, "%Y-%m").strftime("%b-%Y")
    cur_month    = selected_month
    result = []
    counts_healed = False
    for co in companies:
        try:
            units, inv_dicts, exp_dicts = _load_company_data(co.id, tid, db)
            if units:
                if _heal_company_sync_fields(co, units):
                    counts_healed = True
            inv_by_unit: dict[str, list[dict]] = defaultdict(list)
            for inv in inv_dicts:
                inv_by_unit[inv["unit_id"]].append(inv)
            unit_dicts = [_unit_dict(u, inv_by_unit.get(str(u.id), []), today) for u in units]
            summ = company_summary(unit_dicts, inv_dicts, exp_dicts, today, cur_month=cur_month)
            _apply_collected_fallback(summ, co, month_abbrev, db, tid)
            props = db.query(RentalProp).filter(RentalProp.company_id == co.id).all()
            result.append({
                "id": str(co.id),
                "company_name": co.company_name,
                "property_name": props[0].property_name if props else "",
                "property_count": len(props),
                **summ,
                "sync_collected": _company_sync_collected(co),
                "sync_vacancy_loss": float(co.vacancy_loss) if co.vacancy_loss is not None else None,
                "sync_gross_potential": float(co.gross_potential_rent) if co.gross_potential_rent is not None else None,
                "sync_occupied_units": co.occupied_units,
                "sync_total_units": co.total_units,
                "last_sync_month": co.last_sync_month,
                "monthly_rent_data": co.monthly_rent_data,
            })
        except Exception as row_exc:
            _log.error("list_companies row failed for %s: %s", co.company_name, row_exc)
            result.append({
                "id": str(co.id),
                "company_name": co.company_name,
                "property_name": "",
                "property_count": 0,
                "total_units": co.total_units or 0,
                "occupied_units": co.occupied_units or 0,
                "vacant_units": 0,
                "occupancy_pct": 0,
                "sync_collected": _company_sync_collected(co),
                "sync_vacancy_loss": float(co.vacancy_loss) if co.vacancy_loss is not None else None,
                "sync_gross_potential": float(co.gross_potential_rent) if co.gross_potential_rent is not None else None,
                "sync_occupied_units": co.occupied_units,
                "sync_total_units": co.total_units,
                "last_sync_month": co.last_sync_month,
                "monthly_rent_data": co.monthly_rent_data,
            })
    if counts_healed:
        try:
            db.commit()
        except Exception as commit_exc:
            db.rollback()
            _log.warning("list_companies heal commit failed: %s", commit_exc)
    if fmt == "csv":
        output = io.StringIO()
        if result:
            writer = csv.DictWriter(output, fieldnames=result[0].keys())
            writer.writeheader()
            writer.writerows(result)
        return StreamingResponse(
            iter([output.getvalue()]),
            media_type="text/csv",
            headers={"Content-Disposition": "attachment; filename=occupancy_report.csv"},
        )
    return result


@router.post("/companies", status_code=201)
def create_company(
    body: dict,
    current_user: CurrentUser = Depends(require_write_access()),
    db: Session = Depends(get_db),
):
    co = RentalCompany(
        tenant_id=current_user.tenant_id,
        company_name=body["company_name"],
        created_by=current_user.email,
    )
    db.add(co)
    db.commit()
    db.refresh(co)
    return {"id": str(co.id), "company_name": co.company_name}


@router.patch("/companies/{company_id}/status")
def toggle_company_status(
    company_id: uuid.UUID,
    body: dict,
    current_user: CurrentUser = Depends(require_write_access()),
    db: Session = Depends(get_db),
):
    co = db.query(RentalCompany).filter(
        RentalCompany.id == company_id,
        RentalCompany.tenant_id == current_user.tenant_id,
    ).first()
    if not co:
        raise HTTPException(404, "Company not found")
    co.status = body.get("status", "active")
    db.commit()
    return {"id": str(co.id), "status": co.status}


@router.put("/companies/{company_id}")
def update_company(
    company_id: uuid.UUID,
    body: dict,
    current_user: CurrentUser = Depends(require_write_access()),
    db: Session = Depends(get_db),
):
    co = db.query(RentalCompany).filter(
        RentalCompany.id == company_id,
        RentalCompany.tenant_id == current_user.tenant_id,
    ).first()
    if not co:
        raise HTTPException(404, "Company not found")
    if "company_name" in body:
        co.company_name = body["company_name"]
    db.commit()
    db.refresh(co)
    return {"id": str(co.id), "company_name": co.company_name}


@router.delete("/companies/{company_id}", status_code=204)
def delete_company(
    company_id: uuid.UUID,
    current_user: CurrentUser = Depends(require_write_access()),
    db: Session = Depends(get_db),
):
    co = db.query(RentalCompany).filter(
        RentalCompany.id == company_id,
        RentalCompany.tenant_id == current_user.tenant_id,
    ).first()
    if not co:
        raise HTTPException(404, "Company not found")
    db.delete(co)
    db.commit()


@router.delete("/companies", status_code=200)
def delete_all_companies(
    current_user: CurrentUser = Depends(require_write_access()),
    db: Session = Depends(get_db),
):
    """Delete ALL companies and their cascading data for this tenant."""
    tid = current_user.tenant_id
    companies = db.query(RentalCompany).filter(RentalCompany.tenant_id == tid).all()
    for co in companies:
        _wipe_company_units_and_suites(co.id, tid, db)
    count = len(companies)
    db.query(RentalOwnership).filter(RentalOwnership.tenant_id == tid).delete(synchronize_session=False)
    db.query(RentalCompany).filter(RentalCompany.tenant_id == tid).delete(synchronize_session=False)
    db.commit()
    return {"deleted": count, "message": f"Deleted {count} companies and all associated data."}


@router.get("/companies/{company_id}/dashboard")
def company_dashboard(
    company_id: uuid.UUID,
    month: str = Query(None, description="YYYY-MM; defaults to current month"),
    current_user: CurrentUser = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    tid = current_user.tenant_id
    today = date.today()
    if month:
        try:
            selected = date.fromisoformat(f"{month}-01")
        except ValueError:
            raise HTTPException(400, "month must be YYYY-MM format")
    else:
        selected = today.replace(day=1)

    co = db.query(RentalCompany).filter(RentalCompany.id == company_id, RentalCompany.tenant_id == tid).first()
    if not co:
        raise HTTPException(404, "Company not found")

    month_abbrev = selected.strftime("%b-%Y")  # e.g. "Jul-2026"
    cur_month    = selected.strftime("%Y-%m")  # e.g. "2026-07"

    units, inv_dicts, exp_dicts = _load_company_data(co.id, tid, db)
    inv_by_unit: dict[str, list[dict]] = defaultdict(list)
    for inv in inv_dicts:
        inv_by_unit[inv["unit_id"]].append(inv)
    unit_dicts = [_unit_dict(u, inv_by_unit.get(str(u.id), []), today) for u in units]
    summ = company_summary(unit_dicts, inv_dicts, exp_dicts, today, cur_month=cur_month)

    # Apply Excel-sync → P&L fallback so KPIs match Portfolio Overview
    _apply_collected_fallback(summ, co, month_abbrev, db, tid)

    # Income trend — from invoice/collection records; fallback to synced monthly data
    trend = income_trend(inv_dicts, exp_dicts, months=6)
    if not trend and co.monthly_rent_data:
        # Build a synthetic 6-month trend from synced collected amounts.
        # Sort chronologically using MONTH_OPTIONS order (not alphabetically).
        def _month_key(m: str) -> int:
            try:
                return MONTH_OPTIONS.index(m)
            except ValueError:
                return 999
        sorted_entries = sorted(
            ((m, float(v)) for m, v in co.monthly_rent_data.items() if float(v) > 0),
            key=lambda x: _month_key(x[0]),
        )[-6:]
        trend = [
            {"month": m, "billed": 0.0, "collected": v, "expense": 0.0, "noi": v}
            for m, v in sorted_entries
        ]

    # Expense breakdown — from RentalExpense records; fallback to P&L upload
    exp_by_cat: dict[str, float] = defaultdict(float)
    for e in exp_dicts:
        exp_by_cat[e["category"]] += e["amount"]
    expense_breakdown = [{"category": k, "amount": round(v, 2)} for k, v in exp_by_cat.items()]
    if not expense_breakdown:
        expense_breakdown = _pl_expense_breakdown(co, month_abbrev, db, tid)

    ownership_rows = db.query(RentalOwnership).filter(
        RentalOwnership.tenant_id == tid, RentalOwnership.company_id == co.id,
    ).all()
    own_dicts = [
        {"partner_name": o.partner_name, "ownership_pct": float(o.ownership_pct), "role": o.role.value}
        for o in ownership_rows
    ]
    partner_distribution = distribute_to_partners(summ["noi_this_month"], own_dicts)

    props = db.query(RentalProp).filter(RentalProp.company_id == co.id).all()

    return {
        "id": str(co.id),
        "company_name": co.company_name,
        "property_name": props[0].property_name if props else "",
        "property_count": len(props),
        **summ,
        "income_trend": trend,
        "expense_breakdown": expense_breakdown,
        "units": unit_dicts,
        "ownership": own_dicts,
        "partner_distribution": partner_distribution,
    }


# ── suites (properties) ───────────────────────────────────────────────────────

@router.get("/suites")
def list_suites(
    company_id: str = Query(None),
    current_user: CurrentUser = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    tid = current_user.tenant_id
    q = db.query(RentalProp).filter(RentalProp.tenant_id == tid)
    if company_id:
        q = q.filter(RentalProp.company_id == uuid.UUID(company_id))
    suites = q.order_by(RentalProp.property_name).all()
    return [
        {
            "id": str(s.id),
            "company_id": str(s.company_id),
            "property_name": s.property_name,
            "address": s.address,
            "property_type": s.property_type,
            "unit_count": len(s.units),
        }
        for s in suites
    ]


@router.post("/suites", status_code=201)
def create_suite(
    body: dict,
    current_user: CurrentUser = Depends(require_write_access()),
    db: Session = Depends(get_db),
):
    cid = uuid.UUID(body["company_id"])
    co = db.query(RentalCompany).filter(
        RentalCompany.id == cid,
        RentalCompany.tenant_id == current_user.tenant_id,
    ).first()
    if not co:
        raise HTTPException(404, "Company not found")
    s = RentalProp(
        tenant_id=current_user.tenant_id,
        company_id=cid,
        property_name=body["property_name"],
        address=body.get("address"),
        property_type=body.get("property_type"),
    )
    db.add(s)
    db.commit()
    db.refresh(s)
    return {"id": str(s.id), "property_name": s.property_name, "company_id": str(s.company_id)}


@router.put("/suites/{suite_id}")
def update_suite(
    suite_id: uuid.UUID,
    body: dict,
    current_user: CurrentUser = Depends(require_write_access()),
    db: Session = Depends(get_db),
):
    s = db.query(RentalProp).filter(
        RentalProp.id == suite_id,
        RentalProp.tenant_id == current_user.tenant_id,
    ).first()
    if not s:
        raise HTTPException(404, "Suite not found")
    if "property_name" in body and body["property_name"]:
        s.property_name = body["property_name"]
    if "address" in body:
        s.address = body["address"] or None
    if "property_type" in body:
        s.property_type = body["property_type"] or None
    db.commit()
    db.refresh(s)
    return {"id": str(s.id), "property_name": s.property_name}


@router.delete("/suites/{suite_id}", status_code=204)
def delete_suite(
    suite_id: uuid.UUID,
    current_user: CurrentUser = Depends(require_write_access()),
    db: Session = Depends(get_db),
):
    s = db.query(RentalProp).filter(
        RentalProp.id == suite_id,
        RentalProp.tenant_id == current_user.tenant_id,
    ).first()
    if not s:
        raise HTTPException(404, "Suite not found")

    unit_ids = [u.id for u in s.units]
    if unit_ids:
        # maintenance requests → must go before units
        db.query(MaintenanceRequest).filter(MaintenanceRequest.unit_id.in_(unit_ids)).delete(synchronize_session=False)
        # unit inspections chain
        insp_ids = [
            i.id for i in db.query(UnitInspection).filter(UnitInspection.unit_id.in_(unit_ids)).all()
        ]
        if insp_ids:
            db.query(UnitInspectionPhoto).filter(UnitInspectionPhoto.inspection_id.in_(insp_ids)).delete(synchronize_session=False)
            db.query(UnitInspectionChecklistItem).filter(UnitInspectionChecklistItem.inspection_id.in_(insp_ids)).delete(synchronize_session=False)
        db.query(UnitInspection).filter(UnitInspection.unit_id.in_(unit_ids)).delete(synchronize_session=False)
        # invoice chain
        inv_ids = [i.id for i in db.query(RentalInvoice).filter(RentalInvoice.unit_id.in_(unit_ids)).all()]
        if inv_ids:
            db.query(RentalCollection).filter(RentalCollection.invoice_id.in_(inv_ids)).delete(synchronize_session=False)
        db.query(RentalInvoice).filter(RentalInvoice.unit_id.in_(unit_ids)).delete(synchronize_session=False)
        db.query(RentalLease).filter(RentalLease.unit_id.in_(unit_ids)).delete(synchronize_session=False)
        db.query(RentalTenant).filter(RentalTenant.unit_id.in_(unit_ids)).delete(synchronize_session=False)
    db.query(RentalExpense).filter(RentalExpense.property_id == suite_id).delete(synchronize_session=False)
    db.query(RentalUnit).filter(RentalUnit.property_id == suite_id).delete(synchronize_session=False)
    db.query(RentalProp).filter(RentalProp.id == suite_id).delete(synchronize_session=False)
    db.commit()


# ── units ─────────────────────────────────────────────────────────────────────

@router.get("/units")
def list_units(
    company_id: str = Query(None),
    property_id: str = Query(None),
    status: str = Query(None),
    fmt: str = Query(None, alias="format"),
    current_user: CurrentUser = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    tid = current_user.tenant_id
    today = date.today()
    q = db.query(RentalUnit).filter(RentalUnit.tenant_id == tid)
    if company_id:
        try:
            q = q.filter(RentalUnit.company_id == uuid.UUID(company_id))
        except ValueError:
            pass
    if property_id:
        try:
            q = q.filter(RentalUnit.property_id == uuid.UUID(property_id))
        except ValueError:
            pass
    if status:
        q = q.filter(RentalUnit.status == status)
    units = q.all()

    unit_ids = [u.id for u in units]
    invoices = db.query(RentalInvoice).filter(
        RentalInvoice.tenant_id == tid, RentalInvoice.unit_id.in_(unit_ids)
    ).all() if unit_ids else []
    inv_by_unit: dict[str, list[dict]] = defaultdict(list)
    for inv in invoices:
        inv_by_unit[str(inv.unit_id)].append(_inv_dict(inv))

    result = [_unit_dict(u, inv_by_unit.get(str(u.id), []), today) for u in units]

    if fmt == "csv":
        output = io.StringIO()
        if result:
            fields = ["unit_number", "company_name", "property_name", "status", "tenant_name", "lease_end", "monthly_rent", "arrears", "days_vacant"]
            writer = csv.DictWriter(output, fieldnames=fields, extrasaction="ignore")
            writer.writeheader()
            writer.writerows(result)
        return StreamingResponse(
            iter([output.getvalue()]),
            media_type="text/csv",
            headers={"Content-Disposition": "attachment; filename=rent_roll.csv"},
        )
    return result


@router.post("/units", status_code=201)
def create_unit(
    body: dict,
    current_user: CurrentUser = Depends(require_write_access()),
    db: Session = Depends(get_db),
):
    u = RentalUnit(
        tenant_id=current_user.tenant_id,
        property_id=uuid.UUID(body["property_id"]),
        company_id=uuid.UUID(body["company_id"]),
        unit_number=body["unit_number"],
        status=body.get("status", "vacant"),
        monthly_rent=float(body["monthly_rent"]),
        status_changed_at=date.fromisoformat(body["status_changed_at"]) if body.get("status_changed_at") else None,
    )
    db.add(u)
    db.commit()
    db.refresh(u)
    return {"id": str(u.id)}


@router.put("/units/{unit_id}")
def update_unit(
    unit_id: uuid.UUID,
    body: dict,
    current_user: CurrentUser = Depends(require_write_access()),
    db: Session = Depends(get_db),
):
    u = db.query(RentalUnit).filter(RentalUnit.id == unit_id, RentalUnit.tenant_id == current_user.tenant_id).first()
    if not u:
        raise HTTPException(404)
    for field in ("unit_number", "status", "monthly_rent", "vacancy_loss"):
        if field in body:
            setattr(u, field, body[field])
    if "status_changed_at" in body:
        u.status_changed_at = date.fromisoformat(body["status_changed_at"]) if body["status_changed_at"] else None
    db.commit()
    return {"id": str(u.id)}


@router.delete("/units/{unit_id}", status_code=204)
def delete_unit(
    unit_id: uuid.UUID,
    current_user: CurrentUser = Depends(require_write_access()),
    db: Session = Depends(get_db),
):
    u = db.query(RentalUnit).filter(RentalUnit.id == unit_id, RentalUnit.tenant_id == current_user.tenant_id).first()
    if not u:
        raise HTTPException(404)
    db.delete(u)
    db.commit()


# ── leases ────────────────────────────────────────────────────────────────────

@router.get("/leases")
def list_leases(
    company_id: str = Query(None),
    fmt: str = Query(None, alias="format"),
    current_user: CurrentUser = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    tid = current_user.tenant_id
    today = date.today()
    q = db.query(RentalLease).filter(RentalLease.tenant_id == tid)
    if company_id:
        try:
            cid = uuid.UUID(company_id)
            unit_ids = [u.id for u in db.query(RentalUnit).filter(RentalUnit.tenant_id == tid, RentalUnit.company_id == cid).all()]
            if unit_ids:
                q = q.filter(RentalLease.unit_id.in_(unit_ids))
            else:
                return []
        except ValueError:
            pass
    leases = q.all()
    result = [_lease_dict(l) for l in leases]
    expiry_30 = sum(1 for l in result if l["days_until_expiry"] is not None and 0 <= l["days_until_expiry"] <= 30)
    expiry_60 = sum(1 for l in result if l["days_until_expiry"] is not None and 0 <= l["days_until_expiry"] <= 60)
    expiry_90 = sum(1 for l in result if l["days_until_expiry"] is not None and 0 <= l["days_until_expiry"] <= 90)

    if fmt == "csv":
        output = io.StringIO()
        if result:
            fields = ["unit_number", "company_name", "tenant_name", "lease_start", "lease_end", "days_until_expiry", "status", "deposit_amount", "escalation_pct_annual"]
            writer = csv.DictWriter(output, fieldnames=fields, extrasaction="ignore")
            writer.writeheader()
            writer.writerows(result)
        return StreamingResponse(
            iter([output.getvalue()]),
            media_type="text/csv",
            headers={"Content-Disposition": "attachment; filename=lease_expiry.csv"},
        )
    return {
        "leases": result,
        "expiry_pipeline": {"days_30": expiry_30, "days_60": expiry_60, "days_90": expiry_90},
    }


@router.post("/leases", status_code=201)
def create_lease(
    body: dict,
    current_user: CurrentUser = Depends(require_write_access()),
    db: Session = Depends(get_db),
):
    l = RentalLease(
        tenant_id=current_user.tenant_id,
        unit_id=uuid.UUID(body["unit_id"]),
        r_tenant_id=uuid.UUID(body["r_tenant_id"]) if body.get("r_tenant_id") else None,
        lease_start=date.fromisoformat(body["lease_start"]),
        lease_end=date.fromisoformat(body["lease_end"]),
        escalation_pct_annual=float(body["escalation_pct_annual"]) if body.get("escalation_pct_annual") else None,
        deposit_amount=float(body["deposit_amount"]) if body.get("deposit_amount") else None,
        notice_period_days=int(body["notice_period_days"]) if body.get("notice_period_days") else None,
        lock_in_end_date=date.fromisoformat(body["lock_in_end_date"]) if body.get("lock_in_end_date") else None,
        status=body.get("status", "active"),
    )
    db.add(l)
    db.commit()
    db.refresh(l)
    return {"id": str(l.id)}


@router.put("/leases/{lease_id}")
def update_lease(
    lease_id: uuid.UUID,
    body: dict,
    current_user: CurrentUser = Depends(require_write_access()),
    db: Session = Depends(get_db),
):
    l = db.query(RentalLease).filter(RentalLease.id == lease_id, RentalLease.tenant_id == current_user.tenant_id).first()
    if not l:
        raise HTTPException(404)
    if "status" in body:
        l.status = body["status"]
    if "lease_end" in body:
        l.lease_end = date.fromisoformat(body["lease_end"])
    db.commit()
    return {"id": str(l.id)}


# ── collections ───────────────────────────────────────────────────────────────

@router.get("/collections")
def list_collections(
    company_id: str = Query(None),
    month: str = Query(None),  # YYYY-MM, defaults to current month
    fmt: str = Query(None, alias="format"),
    current_user: CurrentUser = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    tid = current_user.tenant_id
    today = date.today()
    cur_month = month if month else today.strftime("%Y-%m")
    q = db.query(RentalInvoice).filter(RentalInvoice.tenant_id == tid)
    if company_id:
        try:
            cid = uuid.UUID(company_id)
            unit_ids = [u.id for u in db.query(RentalUnit).filter(RentalUnit.tenant_id == tid, RentalUnit.company_id == cid).all()]
            if unit_ids:
                q = q.filter(RentalInvoice.unit_id.in_(unit_ids))
            else:
                return {"items": [], "summary": {}}
        except ValueError:
            pass
    # Filter invoices to the selected billing month
    invoices = [i for i in q.all() if str(i.billing_period)[:7] == cur_month]
    inv_dicts = [_inv_dict(i) for i in invoices]

    items = []
    for inv in invoices:
        d = _inv_dict(inv)
        collected = sum(c["amount_collected"] for c in d["collections"])
        balance = max(0.0, d["amount_billed"] - collected)
        status_str = "paid" if balance == 0 else ("partial" if collected > 0 else "unpaid")
        unit = inv.unit
        items.append({
            **d,
            "unit_number": unit.unit_number if unit else "",
            "company_name": unit.company.company_name if unit and unit.company else "",
            "company_id": str(unit.company_id) if unit else "",
            "amount_collected": collected,
            "balance": balance,
            "collection_status": status_str,
        })

    total_billed = sum(i["amount_billed"] for i in items)
    total_collected = sum(i["amount_collected"] for i in items)
    total_arrears = sum(i["balance"] for i in items)
    collection_rate = total_collected / total_billed if total_billed else 0.0
    aging = arrears_aging(inv_dicts, today)

    if fmt == "csv":
        output = io.StringIO()
        if items:
            fields = ["unit_number", "company_name", "billing_period", "amount_billed", "amount_collected", "balance", "collection_status"]
            writer = csv.DictWriter(output, fieldnames=fields, extrasaction="ignore")
            writer.writeheader()
            writer.writerows(items)
        return StreamingResponse(
            iter([output.getvalue()]),
            media_type="text/csv",
            headers={"Content-Disposition": "attachment; filename=arrears_aging.csv"},
        )
    return {
        "items": items,
        "month": cur_month,
        "summary": {
            "total_billed": round(total_billed, 2),
            "total_collected": round(total_collected, 2),
            "total_arrears": round(total_arrears, 2),
            "collection_rate": round(collection_rate, 4),
        },
        "arrears_aging": aging,
    }


@router.post("/invoices", status_code=201)
def create_invoice(
    body: dict,
    current_user: CurrentUser = Depends(require_write_access()),
    db: Session = Depends(get_db),
):
    inv = RentalInvoice(
        tenant_id=current_user.tenant_id,
        unit_id=uuid.UUID(body["unit_id"]),
        lease_id=uuid.UUID(body["lease_id"]),
        billing_period=date.fromisoformat(body["billing_period"]),
        amount_billed=float(body["amount_billed"]),
    )
    db.add(inv)
    db.commit()
    db.refresh(inv)
    return {"id": str(inv.id)}


@router.post("/collections", status_code=201)
def create_collection(
    body: dict,
    current_user: CurrentUser = Depends(require_write_access()),
    db: Session = Depends(get_db),
):
    col = RentalCollection(
        tenant_id=current_user.tenant_id,
        invoice_id=uuid.UUID(body["invoice_id"]),
        amount_collected=float(body["amount_collected"]),
        collected_date=date.fromisoformat(body["collected_date"]),
    )
    db.add(col)
    db.commit()
    db.refresh(col)
    return {"id": str(col.id)}


# ── expenses ──────────────────────────────────────────────────────────────────

@router.get("/expenses")
def list_expenses(
    company_id: str = Query(None),
    category: str = Query(None),
    fmt: str = Query(None, alias="format"),
    current_user: CurrentUser = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    tid = current_user.tenant_id
    q = db.query(RentalExpense).filter(RentalExpense.tenant_id == tid)
    if company_id:
        try:
            q = q.filter(RentalExpense.company_id == uuid.UUID(company_id))
        except ValueError:
            pass
    if category:
        try:
            q = q.filter(RentalExpense.category == RentalExpenseCategory(category))
        except ValueError:
            pass
    expenses = q.order_by(RentalExpense.expense_date.desc()).all()
    result = [_expense_dict(e) for e in expenses]

    if fmt == "csv":
        output = io.StringIO()
        if result:
            writer = csv.DictWriter(output, fieldnames=result[0].keys())
            writer.writeheader()
            writer.writerows(result)
        return StreamingResponse(
            iter([output.getvalue()]),
            media_type="text/csv",
            headers={"Content-Disposition": "attachment; filename=expenses.csv"},
        )
    return result


@router.get("/expenses-summary")
def get_expenses_summary(
    company_id: str = Query(None),
    current_user: CurrentUser = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Get expenses with KPIs and category breakdown."""
    tid = current_user.tenant_id
    q = db.query(RentalExpense).filter(RentalExpense.tenant_id == tid)
    if company_id:
        try:
            q = q.filter(RentalExpense.company_id == uuid.UUID(company_id))
        except ValueError:
            pass
    expenses = q.order_by(RentalExpense.expense_date.desc()).all()
    items = [_expense_dict(e) for e in expenses]

    today = date.today()
    current_month = today.strftime("%Y-%m")

    total_this_month = sum(e.amount for e in expenses if str(e.expense_date)[:7] == current_month)
    total_all_time = sum(e.amount for e in expenses)

    # Group by category
    by_category_dict = {}
    for e in expenses:
        cat = e.category.value
        by_category_dict[cat] = by_category_dict.get(cat, 0) + e.amount

    by_category = [{"category": k, "amount": v} for k, v in sorted(by_category_dict.items(), key=lambda x: x[1], reverse=True)]
    most_expensive_category = by_category[0]["category"] if by_category else None

    return {
        "kpis": {
            "total_this_month": round(total_this_month, 2),
            "total_all_time": round(total_all_time, 2),
            "most_expensive_category": most_expensive_category,
        },
        "items": items,
        "by_category": by_category,
    }


@router.post("/expenses", status_code=201)
def create_expense(
    body: dict,
    current_user: CurrentUser = Depends(require_write_access()),
    db: Session = Depends(get_db),
):
    company_id = uuid.UUID(body["company_id"])

    # If property_id is empty/missing, get the first property for this company
    property_id = body.get("property_id") or ""
    if not property_id or property_id == "":
        prop = db.query(RentalProp).filter(
            RentalProp.tenant_id == current_user.tenant_id,
            RentalProp.company_id == company_id
        ).first()
        if prop:
            property_id = str(prop.id)
        else:
            raise HTTPException(status_code=400, detail="No properties found for this company")

    e = RentalExpense(
        tenant_id=current_user.tenant_id,
        property_id=uuid.UUID(property_id),
        company_id=company_id,
        expense_date=date.fromisoformat(body["expense_date"]),
        category=RentalExpenseCategory(body["category"]),
        amount=float(body["amount"]),
        description=body.get("description"),
        created_by=current_user.email,
    )
    db.add(e)
    db.commit()
    db.refresh(e)
    return {"id": str(e.id)}


@router.delete("/expenses/{expense_id}", status_code=204)
def delete_expense(
    expense_id: uuid.UUID,
    current_user: CurrentUser = Depends(require_write_access()),
    db: Session = Depends(get_db),
):
    e = db.query(RentalExpense).filter(RentalExpense.id == expense_id, RentalExpense.tenant_id == current_user.tenant_id).first()
    if not e:
        raise HTTPException(404)
    db.delete(e)
    db.commit()


# ── ownership ─────────────────────────────────────────────────────────────────

@router.get("/ownership")
def list_ownership(
    fmt: str = Query(None, alias="format"),
    current_user: CurrentUser = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    tid = current_user.tenant_id
    today = date.today()
    rows = db.query(RentalOwnership).filter(RentalOwnership.tenant_id == tid).all()

    # collect NOI per company
    companies = db.query(RentalCompany).filter(RentalCompany.tenant_id == tid).all()
    co_noi: dict[str, float] = {}
    for co in companies:
        units, inv_dicts, exp_dicts = _load_company_data(co.id, tid, db)
        inv_by_unit: dict[str, list[dict]] = defaultdict(list)
        for inv in inv_dicts:
            inv_by_unit[inv["unit_id"]].append(inv)
        unit_dicts = [_unit_dict(u, inv_by_unit.get(str(u.id), []), today) for u in units]
        summ = company_summary(unit_dicts, inv_dicts, exp_dicts, today)
        co_noi[str(co.id)] = summ["noi_this_month"]

    # group by partner
    by_partner: dict[str, list[dict]] = defaultdict(list)
    for row in rows:
        noi = co_noi.get(str(row.company_id), 0.0)
        by_partner[row.partner_name].append({
            "ownership_id": str(row.id),
            "company_id": str(row.company_id),
            "company_name": row.company.company_name if row.company else "",
            "ownership_pct": float(row.ownership_pct),
            "role": row.role.value,
            "noi_this_month": noi,
            "noi_share": round(noi * float(row.ownership_pct), 2),
        })

    result = []
    for partner_name, holdings in by_partner.items():
        total_share = round(sum(h["noi_share"] for h in holdings), 2)
        result.append({
            "partner_name": partner_name,
            "company_count": len(holdings),
            "total_noi_share": total_share,
            "holdings": holdings,
        })

    if fmt == "csv":
        flat = []
        for p in result:
            for h in p["holdings"]:
                flat.append({"partner_name": p["partner_name"], **h})
        output = io.StringIO()
        if flat:
            writer = csv.DictWriter(output, fieldnames=flat[0].keys())
            writer.writeheader()
            writer.writerows(flat)
        return StreamingResponse(
            iter([output.getvalue()]),
            media_type="text/csv",
            headers={"Content-Disposition": "attachment; filename=partner_distribution.csv"},
        )
    return result


@router.post("/ownership", status_code=201)
def create_ownership(
    body: dict,
    current_user: CurrentUser = Depends(require_write_access()),
    db: Session = Depends(get_db),
):
    o = RentalOwnership(
        tenant_id=current_user.tenant_id,
        company_id=uuid.UUID(body["company_id"]),
        partner_name=body["partner_name"],
        ownership_pct=float(body["ownership_pct"]),
        role=RentalPartnerRole(body.get("role", "limited_partner")),
    )
    db.add(o)
    db.commit()
    db.refresh(o)
    return {"id": str(o.id)}


@router.post("/ownership/import")
async def import_ownership(
    company_id: str = Form(...),
    file: UploadFile = File(...),
    current_user: CurrentUser = Depends(require_write_access()),
    db: Session = Depends(get_db),
):
    """Import partners from Excel file. Auto-detects header row and column positions.
    Supports flexible formats — looks for columns named Partner Name, Ownership %, Role.
    Role aliases: Managing Partner → managing_member, Investor Partner → limited_partner, etc.
    Also stores Capital Contributed and Current Equity Balance if present.
    """
    import openpyxl
    try:
        content = await file.read()
        wb = openpyxl.load_workbook(io.BytesIO(content))
        ws = wb.active

        co_id = uuid.UUID(company_id)
        imported_count = 0
        errors = []

        ROLE_MAP = {
            "managing_partner":   "managing_member",
            "managing_member":    "managing_member",
            "general_partner":    "general_partner",
            "limited_partner":    "limited_partner",
            "partner":            "limited_partner",
            "investor_partner":   "limited_partner",
            "passive_investor":   "passive_investor",
            "silent_partner":     "silent_partner",
            "management_entity":  "managing_member",
        }
        VALID_ROLES = set(ROLE_MAP.values())

        # ── Auto-detect header row by scanning for "partner name" keyword ──
        header_row_idx = None
        col_map: dict[str, int] = {}
        all_rows = list(ws.iter_rows(values_only=True))

        for i, row in enumerate(all_rows):
            cells = [str(c or "").strip().lower() for c in row]
            if any("partner name" in c or c == "partner" for c in cells):
                header_row_idx = i
                for j, cell in enumerate(cells):
                    if "partner name" in cell or cell == "partner":
                        col_map["name"] = j
                    elif "ownership" in cell or "own %" in cell or cell == "own%":
                        col_map["pct"] = j
                    elif "role" in cell:
                        col_map["role"] = j
                    elif "capital contributed" in cell or "capital in" in cell:
                        col_map["capital"] = j
                    elif "equity balance" in cell or "current equity" in cell:
                        col_map["equity"] = j
                break

        # Fallback to positional (col 0=name, 1=pct, 2=role) if no header found
        if header_row_idx is None:
            col_map = {"name": 0, "pct": 1, "role": 2}
            data_rows = all_rows[1:]  # skip first row
        else:
            data_rows = all_rows[header_row_idx + 1:]

        name_col   = col_map.get("name", 0)
        pct_col    = col_map.get("pct", 1)
        role_col   = col_map.get("role", 2)
        cap_col    = col_map.get("capital")
        equity_col = col_map.get("equity")

        for row_idx, row in enumerate(data_rows, start=1):
            try:
                raw_name = row[name_col] if len(row) > name_col else None
                if not raw_name:
                    continue
                partner_name = str(raw_name).strip()
                # Skip totals row
                if partner_name.upper() in ("TOTAL", "GRAND TOTAL", "SUM"):
                    continue

                raw_pct = row[pct_col] if len(row) > pct_col else 0
                try:
                    ownership_pct = float(raw_pct or 0)
                except (ValueError, TypeError):
                    ownership_pct = 0.0

                raw_role = str(row[role_col] if len(row) > role_col else "").strip().lower().replace(" ", "_") if role_col is not None else ""
                role = ROLE_MAP.get(raw_role, "limited_partner")
                if role not in VALID_ROLES:
                    role = "limited_partner"

                capital_contributed = None
                if cap_col is not None and len(row) > cap_col:
                    try:
                        capital_contributed = float(row[cap_col] or 0)
                    except (ValueError, TypeError):
                        capital_contributed = None

                equity_balance = None
                if equity_col is not None and len(row) > equity_col:
                    try:
                        equity_balance = float(row[equity_col] or 0)
                    except (ValueError, TypeError):
                        equity_balance = None

                existing = db.query(RentalOwnership).filter(
                    RentalOwnership.tenant_id == current_user.tenant_id,
                    RentalOwnership.company_id == co_id,
                    RentalOwnership.partner_name == partner_name,
                ).first()

                if not existing:
                    kwargs: dict = dict(
                        tenant_id=current_user.tenant_id,
                        company_id=co_id,
                        partner_name=partner_name,
                        ownership_pct=ownership_pct,
                        role=RentalPartnerRole(role),
                    )
                    if capital_contributed is not None and hasattr(RentalOwnership, "capital_contributed"):
                        kwargs["capital_contributed"] = capital_contributed
                    if equity_balance is not None and hasattr(RentalOwnership, "equity_balance"):
                        kwargs["equity_balance"] = equity_balance
                    db.add(RentalOwnership(**kwargs))
                    imported_count += 1
                else:
                    existing.ownership_pct = ownership_pct
                    existing.role = RentalPartnerRole(role)
                    if capital_contributed is not None and hasattr(existing, "capital_contributed"):
                        existing.capital_contributed = capital_contributed
                    if equity_balance is not None and hasattr(existing, "equity_balance"):
                        existing.equity_balance = equity_balance

            except Exception as e:
                errors.append(f"Row {row_idx}: {str(e)}")

        db.commit()
        return {
            "status": "imported",
            "imported_count": imported_count,
            "errors": errors,
        }
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Import failed: {str(e)}")


# ── vacancy ───────────────────────────────────────────────────────────────────

@router.get("/vacancy")
def vacancy_summary(
    fmt: str = Query(None, alias="format"),
    current_user: CurrentUser = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    tid = current_user.tenant_id
    today = date.today()
    units = db.query(RentalUnit).filter(
        RentalUnit.tenant_id == tid, RentalUnit.status == "vacant"
    ).all()

    result = []
    for u in units:
        dv = days_vacant(u.status.value, u.status_changed_at, today)
        result.append({
            "id": str(u.id),
            "unit_number": u.unit_number,
            "company_name": u.company.company_name if u.company else "",
            "company_id": str(u.company_id),
            "property_name": u.property.property_name if u.property else "",
            "monthly_rent": float(u.monthly_rent),
            "status_changed_at": u.status_changed_at.isoformat() if u.status_changed_at else None,
            "days_vacant": dv,
        })

    by_company: dict[str, float] = defaultdict(float)
    for r in result:
        by_company[r["company_name"]] += r["monthly_rent"]

    total_loss = sum(r["monthly_rent"] for r in result)
    avg_days = round(sum(r["days_vacant"] or 0 for r in result) / len(result), 1) if result else 0

    if fmt == "csv":
        output = io.StringIO()
        if result:
            writer = csv.DictWriter(output, fieldnames=result[0].keys())
            writer.writeheader()
            writer.writerows(result)
        return StreamingResponse(
            iter([output.getvalue()]),
            media_type="text/csv",
            headers={"Content-Disposition": "attachment; filename=vacancy_loss.csv"},
        )
    return {
        "vacant_units": result,
        "summary": {
            "count": len(result),
            "total_vacancy_loss": round(total_loss, 2),
            "avg_days_vacant": avg_days,
        },
        "loss_by_company": [{"company_name": k, "vacancy_loss": round(v, 2)} for k, v in by_company.items()],
    }


# ── rent receivable upload ────────────────────────────────────────────────────

@router.post("/upload-rent-receivable/preview")
async def preview_rent_receivable(
    file: UploadFile = File(...),
    target_month: str = Form(default='Jun-2026'),
    current_user: CurrentUser = Depends(get_current_user),
):
    """Step 1: Parse EstateCFO_Rent_Template_ByCompany.xlsx and return preview WITHOUT saving."""
    import tempfile
    from services.rent_receivable_parser import parse_rent_receivable_file

    if target_month not in MONTH_OPTIONS:
        raise HTTPException(status_code=400, detail=f"Invalid month. Choose from: {MONTH_OPTIONS}")

    contents = await file.read()
    with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
        tmp.write(contents)
        tmp_path = tmp.name

    try:
        parsed = parse_rent_receivable_file(tmp_path, target_month=target_month)
        return {
            'target_month': target_month,
            'portfolio': parsed['portfolio'],
            'companies': {
                co: {
                    'company': data['company'],
                    'total_units': data['total_physical_units'],
                    'occupied': data['occupied_count'],
                    'vacant': data['vacant_count'],
                    'occupancy_rate': data['occupancy_rate'],
                    'collected': data['collected'],
                    'gross_potential': data['gross_potential'],
                    'vacancy_loss': data['vacancy_loss'],
                    'vacant_units': data['vacant_units'],
                    'monthly_totals': data['monthly_totals'],
                }
                for co, data in parsed['companies'].items()
            },
            'temp_file_id': tmp_path,
        }
    except Exception as e:
        if os.path.exists(tmp_path):
            os.unlink(tmp_path)
        raise HTTPException(status_code=400, detail=f"Parse error: {str(e)}")


@router.post("/upload-rent-receivable/confirm")
async def confirm_rent_receivable(
    payload: dict,
    db: Session = Depends(get_db),
    current_user: CurrentUser = Depends(get_current_user),
):
    """Step 2: Client confirmed preview — save parsed data to all DB tables."""
    from services.rent_receivable_parser import parse_rent_receivable_file

    tmp_path = payload.get('temp_file_id')
    target_month = payload.get('target_month', 'Jun-2026')

    if not tmp_path or not os.path.exists(tmp_path):
        raise HTTPException(status_code=400, detail="Session expired — please upload again")

    tid = current_user.tenant_id

    try:
        parsed = parse_rent_receivable_file(tmp_path, target_month=target_month)
        updated = []

        for co_name, data in parsed['companies'].items():
            company = db.query(RentalCompany).filter(
                RentalCompany.tenant_id == tid,
                or_(
                    func.trim(func.lower(RentalCompany.company_name)) == co_name.lower().strip(),
                    RentalCompany.company_name.ilike(f'%{co_name.strip()}%'),
                )
            ).first()

            if not company:
                continue

            company.last_sync_month = target_month
            company.last_sync_date = datetime.utcnow()

            # ── Update each unit status and rent ───────────────────────────
            from services.rent_receivable_parser import expand_unit_match_names, scale_amount_map

            for unit_data in data['units']:
                raw_name = unit_data['name'].strip()
                is_vacant: bool = unit_data['is_vacant']
                current_amount: float = unit_data['current_amount']
                unit_vac_loss: float = unit_data['vacancy_loss']
                history: dict = unit_data['history']

                matched_units = []
                seen_unit_ids: set = set()
                for part in expand_unit_match_names(raw_name):
                    candidates = [part]
                    if not part.lower().startswith('unit'):
                        candidates.append(f'unit {part.lower()}')
                    for cand in candidates:
                        unit = db.query(RentalUnit).filter(
                            RentalUnit.company_id == company.id,
                            func.lower(func.trim(RentalUnit.unit_number)) == cand.lower().strip(),
                        ).first()
                        if unit and unit.id not in seen_unit_ids:
                            seen_unit_ids.add(unit.id)
                            matched_units.append(unit)

                if not matched_units:
                    continue

                n = len(matched_units)
                hist = scale_amount_map(history, float(n))
                per_rent = round(current_amount / n, 2) if n > 1 else current_amount
                per_vac = round(unit_vac_loss / n, 2) if n > 1 else unit_vac_loss
                rent = per_rent if not is_vacant else per_vac

                for unit in matched_units:
                    unit.status = 'vacant' if is_vacant else 'occupied'
                    unit.monthly_rent = rent if not is_vacant else (per_vac or unit.monthly_rent)
                    unit.rent_history = hist
                    unit.vacancy_loss = per_vac

            registry_units = db.query(RentalUnit).filter(
                RentalUnit.company_id == company.id,
            ).all()
            if registry_units:
                _reconcile_unit_status_for_month(registry_units, target_month)
                _apply_registry_unit_counts(company, registry_units, target_month)
                _apply_registry_financials(company, registry_units, target_month)
            else:
                company.collected_this_month = data['collected']
                company.vacancy_loss = data['vacancy_loss']
                company.gross_potential_rent = data['gross_potential']
                company.monthly_rent_data = data['monthly_totals']
                company.occupied_units = data['occupied_count']
                company.total_units = data['total_physical_units']

            db.commit()
            updated.append(co_name)

        return {
            'status': 'success',
            'target_month': target_month,
            'updated_companies': updated,
            'portfolio': parsed['portfolio'],
            'message': f"Updated {len(updated)} companies for {target_month}",
        }

    finally:
        if os.path.exists(tmp_path):
            os.unlink(tmp_path)


# ── portfolio import from Excel template ──────────────────────────────────────

@router.post("/import-portfolio")
async def import_portfolio(
    file: UploadFile = File(...),
    current_user: CurrentUser = Depends(require_write_access()),
    db: Session = Depends(get_db),
):
    """
    Import companies, suites, and units from the EstateCFO Rent Template Excel.
    Each sheet (except SUMMARY) is treated as one company.
    Units with '(SXXX)' in the name are grouped under a suite named SXXX.
    Existing companies by name are skipped (idempotent).
    """
    import re
    import tempfile
    import openpyxl

    contents = await file.read()
    tmp_fd, tmp_path = tempfile.mkstemp(suffix=".xlsx")
    try:
        with os.fdopen(tmp_fd, "wb") as f:
            f.write(contents)

        wb = openpyxl.load_workbook(tmp_path, data_only=True)
        tid = current_user.tenant_id

        # Load existing company names to skip duplicates
        existing_companies = {
            co.company_name.strip().lower(): co
            for co in db.query(RentalCompany).filter(RentalCompany.tenant_id == tid).all()
        }

        created_companies = 0
        created_suites = 0
        created_units = 0
        skipped_companies: list[str] = []

        for sheet_name in wb.sheetnames:
            if sheet_name.strip().upper() == "SUMMARY":
                continue

            ws = wb[sheet_name]
            company_name = sheet_name.strip()

            # Get or create company
            if company_name.lower() in existing_companies:
                co = existing_companies[company_name.lower()]
                skipped_companies.append(company_name)
            else:
                co = RentalCompany(
                    tenant_id=tid,
                    company_name=company_name,
                    created_by=current_user.email,
                )
                db.add(co)
                db.flush()
                existing_companies[company_name.lower()] = co
                created_companies += 1

            # Read sheet rows — row 3 (1-indexed) is the header, data starts row 4
            # Col A = Unit Name, Cols B-M = Jan-Dec 2026
            rows = list(ws.iter_rows(values_only=True))
            if len(rows) < 4:
                continue

            # Determine which row is the header (contains 'Unit Name')
            header_row_idx = 2  # default index 2 (3rd row, 1-indexed row 3)
            for i, row in enumerate(rows[:5]):
                if row and str(row[0] or "").strip().lower() == "unit name":
                    header_row_idx = i
                    break

            data_rows = rows[header_row_idx + 1:]

            # Group units by suite
            suites: dict[str, list[dict]] = {}   # suite_name → [unit_info]
            no_suite: list[dict] = []

            for row in data_rows:
                if not row or not row[0]:
                    continue
                unit_name = str(row[0]).strip()
                if not unit_name or unit_name.upper().startswith("TOTAL"):
                    continue

                # Monthly payment values (cols 1-12)
                monthly = []
                for v in row[1:13]:
                    try:
                        monthly.append(float(v) if v is not None else 0.0)
                    except (TypeError, ValueError):
                        monthly.append(0.0)

                non_zero = [v for v in monthly if v > 0]
                monthly_rent = max(non_zero) if non_zero else 0.0

                # Determine status: occupied if recent months (Apr=idx3, May=idx4) have payment
                recent = monthly[3:6]  # Apr, May, Jun
                status = "occupied" if any(v > 0 for v in recent) else "vacant"

                # Detect suite tag like (S789)
                suite_match = re.search(r'\(S(\d+)\)', unit_name)
                unit_info = {
                    "unit_number": unit_name,
                    "monthly_rent": monthly_rent,
                    "status": status,
                }
                if suite_match:
                    suite_key = f"Suite {suite_match.group(1)}"
                    suites.setdefault(suite_key, []).append(unit_info)
                else:
                    no_suite.append(unit_info)

            # Create suites and their units
            def _make_suite(suite_name: str) -> RentalProp:
                prop = RentalProp(
                    tenant_id=tid,
                    company_id=co.id,
                    property_name=suite_name,
                )
                db.add(prop)
                db.flush()
                return prop

            def _make_unit(prop: RentalProp, info: dict):
                unit = RentalUnit(
                    tenant_id=tid,
                    property_id=prop.id,
                    company_id=co.id,
                    unit_number=info["unit_number"],
                    status=info["status"],
                    monthly_rent=info["monthly_rent"],
                )
                db.add(unit)

            # Named suites (e.g. Suite 789)
            for suite_name, units_list in suites.items():
                prop = _make_suite(suite_name)
                created_suites += 1
                for info in units_list:
                    _make_unit(prop, info)
                    created_units += 1

            # Units without a suite go under a default suite named after the company
            if no_suite:
                prop = _make_suite(company_name)
                created_suites += 1
                for info in no_suite:
                    _make_unit(prop, info)
                    created_units += 1

        db.commit()
        return {
            "status": "success",
            "created_companies": created_companies,
            "created_suites": created_suites,
            "created_units": created_units,
            "skipped_companies": skipped_companies,
            "message": (
                f"Imported {created_companies} companies, "
                f"{created_suites} suites, {created_units} units. "
                f"Skipped {len(skipped_companies)} existing."
            ),
        }

    finally:
        if os.path.exists(tmp_path):
            os.unlink(tmp_path)

# â”€â”€ hardcoded portfolio seed (no file upload needed) â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€

# ── portfolio import: preview + confirm (with review gate) ────────────────────

@router.post("/import-portfolio/preview")
async def preview_portfolio_import(
    file: UploadFile = File(...),
    target_month: str = Form(default="Jun-2026"),
    current_user: CurrentUser = Depends(require_write_access()),
    db: Session = Depends(get_db),
):
    """
    Parse uploaded Rent Receivable Excel and return a diff-style preview
    (companies/units to create vs match vs skip). Nothing is written to the DB.
    """
    import tempfile
    from services.rent_receivable_parser import parse_rent_receivable_file

    # Normalise month: accept "Jun-2026" or "2026-06"
    if target_month and len(target_month) == 7 and target_month[4] == "-":
        from datetime import datetime as _ddt
        try:
            target_month = _ddt.strptime(target_month, "%Y-%m").strftime("%b-%Y")
        except ValueError:
            pass

    contents = await file.read()
    tmp_fd, tmp_path = tempfile.mkstemp(suffix=".xlsx")
    try:
        with os.fdopen(tmp_fd, "wb") as fh:
            fh.write(contents)
        try:
            parsed = parse_rent_receivable_file(tmp_path, target_month=target_month)
        except Exception as exc:
            raise HTTPException(status_code=400, detail=f"Failed to parse file: {exc}")
    finally:
        if os.path.exists(tmp_path):
            os.unlink(tmp_path)

    tid = current_user.tenant_id
    MONTHS_ORDER = [
        "Jan-2026", "Feb-2026", "Mar-2026", "Apr-2026", "May-2026", "Jun-2026",
        "Jul-2026", "Aug-2026", "Sep-2026", "Oct-2026", "Nov-2026", "Dec-2026",
    ]

    existing_cos = db.query(RentalCompany).filter(RentalCompany.tenant_id == tid).all()
    existing_co_map = {c.company_name.strip().lower(): c for c in existing_cos}

    companies_preview = []

    for co_name, data in parsed["companies"].items():
        norm = co_name.strip().lower()
        match_co = existing_co_map.get(norm)
        if not match_co:
            for k, v in existing_co_map.items():
                if norm in k or k in norm:
                    match_co = v
                    break

        co_action = "match" if match_co else "create"

        ex_unit_map: dict = {}
        if match_co:
            ex_units = db.query(RentalUnit).filter(
                RentalUnit.company_id == match_co.id,
                RentalUnit.tenant_id == tid,
            ).all()
            ex_unit_map = {u.unit_number.strip().lower(): u for u in ex_units}

        units_preview = []
        for unit_data in data["units"]:
            raw_name = unit_data["name"].strip()
            history: dict = unit_data["history"]
            is_vacant: bool = unit_data["is_vacant"]
            current_amount: float = unit_data["current_amount"]
            vacancy_loss: float = unit_data["vacancy_loss"]

            # Keep combined names (e.g. "Unit E, F") as ONE unit — do not split
            best_rent = current_amount if (not is_vacant and current_amount > 0) else vacancy_loss
            if best_rent == 0 and history:
                sorted_nz = [m for m in MONTHS_ORDER if history.get(m, 0) > 0]
                if sorted_nz:
                    best_rent = history[sorted_nz[-1]]

            norm_unit = raw_name.lower()
            ex_unit = ex_unit_map.get(norm_unit)

            if ex_unit:
                unit_action = (
                    "update_rent" if (ex_unit.monthly_rent == 0 and best_rent > 0) else "skip"
                )
            else:
                unit_action = "create"

            units_preview.append({
                    "label": raw_name,
                    "unit_name": raw_name,
                    "suite_name": unit_data.get("suite", ""),
                    "action": unit_action,
                    "monthly_rent": round(best_rent, 2),
                    "status": "vacant" if is_vacant else "occupied",
                    "history": history,
                    "match_unit_id": str(ex_unit.id) if ex_unit else None,
                    "match_unit_rent": ex_unit.monthly_rent if ex_unit else None,
                })

        companies_preview.append({
            "excel_name": co_name,
            "display_name": match_co.company_name if match_co else co_name,
            "action": co_action,
            "match_id": str(match_co.id) if match_co else None,
            "units": units_preview,
            "total_units": data["total_physical_units"],
            "occupied": data["occupied_count"],
            "vacant": data["vacant_count"],
            "target_month": data.get("target_month", ""),
        })

    skipped = parsed["portfolio"].get("skipped", [])
    summary = {
        "companies_to_create": sum(1 for c in companies_preview if c["action"] == "create"),
        "companies_to_match": sum(1 for c in companies_preview if c["action"] == "match"),
        "units_to_create": sum(
            sum(1 for u in c["units"] if u["action"] == "create") for c in companies_preview
        ),
        "units_to_skip": sum(
            sum(1 for u in c["units"] if u["action"] in ("skip", "update_rent"))
            for c in companies_preview
        ),
    }

    return {
        "companies": companies_preview,
        "skipped": skipped,
        "summary": summary,
        "_sheets_found": list(parsed["companies"].keys()),
    }


def _wipe_company_units_and_suites(company_id: uuid.UUID, tid: uuid.UUID, db: Session):
    """Delete all units (+ child records) and suites for a company, keeping the company row."""
    unit_ids = [
        u.id for u in db.query(RentalUnit).filter(
            RentalUnit.company_id == company_id, RentalUnit.tenant_id == tid
        ).all()
    ]
    if unit_ids:
        db.query(MaintenanceRequest).filter(MaintenanceRequest.unit_id.in_(unit_ids)).delete(synchronize_session=False)
        insp_ids = [
            i.id for i in db.query(UnitInspection).filter(UnitInspection.unit_id.in_(unit_ids)).all()
        ]
        if insp_ids:
            db.query(UnitInspectionPhoto).filter(UnitInspectionPhoto.inspection_id.in_(insp_ids)).delete(synchronize_session=False)
            db.query(UnitInspectionChecklistItem).filter(UnitInspectionChecklistItem.inspection_id.in_(insp_ids)).delete(synchronize_session=False)
        db.query(UnitInspection).filter(UnitInspection.unit_id.in_(unit_ids)).delete(synchronize_session=False)
        inv_ids = [i.id for i in db.query(RentalInvoice).filter(RentalInvoice.unit_id.in_(unit_ids)).all()]
        if inv_ids:
            db.query(RentalCollection).filter(RentalCollection.invoice_id.in_(inv_ids)).delete(synchronize_session=False)
        db.query(RentalInvoice).filter(RentalInvoice.unit_id.in_(unit_ids)).delete(synchronize_session=False)
        db.query(RentalLease).filter(RentalLease.unit_id.in_(unit_ids)).delete(synchronize_session=False)
        db.query(RentalTenant).filter(RentalTenant.unit_id.in_(unit_ids)).delete(synchronize_session=False)
        db.query(RentalUnit).filter(RentalUnit.company_id == company_id, RentalUnit.tenant_id == tid).delete(synchronize_session=False)
    # Delete all suites
    suite_ids = [s.id for s in db.query(RentalProp).filter(RentalProp.company_id == company_id, RentalProp.tenant_id == tid).all()]
    if suite_ids:
        db.query(RentalExpense).filter(RentalExpense.property_id.in_(suite_ids)).delete(synchronize_session=False)
        db.query(RentalProp).filter(RentalProp.company_id == company_id, RentalProp.tenant_id == tid).delete(synchronize_session=False)
    db.flush()


@router.post("/import-portfolio/confirm")
async def confirm_portfolio_import(
    payload: dict,
    current_user: CurrentUser = Depends(require_write_access()),
    db: Session = Depends(get_db),
):
    """
    Commit previewed import.
    force_replace=true: for matched companies, wipe existing units/suites and recreate from Excel.
    """
    tid = current_user.tenant_id
    companies_data = payload.get("companies", [])
    force_replace: bool = bool(payload.get("force_replace", False))
    target_month: str = payload.get("target_month") or "Jun-2026"

    created_companies = 0
    replaced_companies = 0
    created_suites = 0
    created_units = 0
    updated_units = 0
    errors: list[str] = []

    try:
        for co in companies_data:
            try:
                co_action = co.get("action")
                suite_cache: dict[str, uuid.UUID] = {}

                if co_action == "create":
                    new_co = RentalCompany(
                        tenant_id=tid,
                        company_name=co["excel_name"].strip(),
                        status="active",
                        created_by=current_user.email,
                    )
                    db.add(new_co)
                    db.flush()
                    company_id = new_co.id
                    created_companies += 1

                elif co_action == "match" and co.get("match_id"):
                    company_id = uuid.UUID(co["match_id"])
                    if force_replace:
                        # Wipe existing suites/units so we can recreate cleanly
                        _wipe_company_units_and_suites(company_id, tid, db)
                        replaced_companies += 1
                        # suite_cache starts empty — all suites will be created fresh
                    else:
                        # Soft mode: reuse existing suites by name
                        existing_suites = db.query(RentalProp).filter(
                            RentalProp.company_id == company_id,
                            RentalProp.tenant_id == tid,
                        ).all()
                        suite_cache = {s.property_name.strip().lower(): s.id for s in existing_suites}

                else:
                    continue

                def _get_or_create_suite(suite_name: str) -> uuid.UUID:
                    nonlocal created_suites
                    prop_name = suite_name.strip() if suite_name.strip() else co.get("display_name", co["excel_name"]).strip()
                    key = prop_name.lower()
                    if key not in suite_cache:
                        new_suite = RentalProp(
                            tenant_id=tid,
                            company_id=company_id,
                            property_name=prop_name,
                        )
                        db.add(new_suite)
                        db.flush()
                        suite_cache[key] = new_suite.id
                        created_suites += 1
                    return suite_cache[key]

                for unit in co.get("units", []):
                    unit_action = unit.get("action")
                    property_id = _get_or_create_suite(unit.get("suite_name", ""))

                    if unit_action == "create" or force_replace:
                        db.add(RentalUnit(
                            tenant_id=tid,
                            property_id=property_id,
                            company_id=company_id,
                            unit_number=unit["unit_name"],
                            status=unit["status"],
                            monthly_rent=float(unit.get("monthly_rent", 0)),
                            rent_history=unit.get("history", {}),
                        ))
                        created_units += 1
                    elif unit_action in ("update_rent", "skip") and unit.get("match_unit_id"):
                        ex = (
                            db.query(RentalUnit)
                            .filter(
                                RentalUnit.id == uuid.UUID(unit["match_unit_id"]),
                                RentalUnit.tenant_id == tid,
                            )
                            .first()
                        )
                        co_month = co.get("target_month") or target_month
                        if ex and _sync_unit_from_preview(ex, unit, co_month):
                            updated_units += 1

                co_month = co.get("target_month") or target_month
                company_row = db.query(RentalCompany).filter(
                    RentalCompany.id == company_id,
                    RentalCompany.tenant_id == tid,
                ).first()
                if company_row:
                    registry_units = db.query(RentalUnit).filter(
                        RentalUnit.company_id == company_id,
                        RentalUnit.tenant_id == tid,
                    ).all()
                    _heal_company_sync_fields(company_row, registry_units, prefer_month=co_month)

            except Exception as row_err:
                errors.append(f"{co.get('excel_name', '?')}: {row_err}")

        db.commit()

    except Exception as exc:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Import failed: {exc}")

    total_new = created_companies + replaced_companies
    return {
        "status": "success",
        "created_companies": created_companies,
        "replaced_companies": replaced_companies,
        "created_suites": created_suites,
        "created_units": created_units,
        "updated_units": updated_units,
        "errors": errors,
        "message": (
            f"Imported {total_new} companies ({replaced_companies} replaced), "
            f"{created_suites} suites, {created_units} units."
        ),
    }


PORTFOLIO_DATA = [
    {"company": "ABC LLC", "suites": [
        {"name": "Suite 789", "units": [
            {"unit_number": "Unit A,B,C (S789)", "monthly_rent": 2050, "status": "occupied"},
            {"unit_number": "Unit D,E,F (S789)", "monthly_rent": 1725, "status": "occupied"},
            {"unit_number": "Unit G (S789)",     "monthly_rent": 775,  "status": "occupied"},
            {"unit_number": "Unit H (S789)",     "monthly_rent": 830,  "status": "occupied"},
            {"unit_number": "Unit I (S789)",     "monthly_rent": 800,  "status": "occupied"},
            {"unit_number": "Unit J,K,L (S789)", "monthly_rent": 1675, "status": "occupied"},
            {"unit_number": "Unit M (S789)",     "monthly_rent": 1750, "status": "occupied"},
            {"unit_number": "Unit N (S789)",     "monthly_rent": 0,    "status": "vacant"},
            {"unit_number": "Unit O (S789)",     "monthly_rent": 0,    "status": "vacant"},
            {"unit_number": "Unit P (S789)",     "monthly_rent": 800,  "status": "occupied"},
            {"unit_number": "Unit Q (S789)",     "monthly_rent": 730,  "status": "occupied"},
            {"unit_number": "Unit R & S (S789)", "monthly_rent": 1800, "status": "occupied"},
            {"unit_number": "Unit T (S789)",     "monthly_rent": 830,  "status": "occupied"},
            {"unit_number": "Unit U (S789)",     "monthly_rent": 700,  "status": "occupied"},
            {"unit_number": "Unit V (S789)",     "monthly_rent": 800,  "status": "occupied"},
            {"unit_number": "Unit W (S789)",     "monthly_rent": 800,  "status": "occupied"},
        ]},
        {"name": "ABC LLC", "units": [
            {"unit_number": "Unit A",     "monthly_rent": 850,  "status": "occupied"},
            {"unit_number": "Unit B",     "monthly_rent": 700,  "status": "occupied"},
            {"unit_number": "Unit C",     "monthly_rent": 925,  "status": "occupied"},
            {"unit_number": "Unit D",     "monthly_rent": 3100, "status": "vacant"},
            {"unit_number": "Unit E,F,G", "monthly_rent": 3100, "status": "occupied"},
            {"unit_number": "Unit 401",   "monthly_rent": 4158, "status": "occupied"},
            {"unit_number": "Unit 402",   "monthly_rent": 2000, "status": "occupied"},
        ]},
    ]},
    {"company": "TOWN Houses", "suites": [{"name": "TOWN Houses", "units": [
        {"unit_number": "NHJ LLC - Unit A", "monthly_rent": 0,    "status": "vacant"},
        {"unit_number": "LOP LLC - Unit B", "monthly_rent": 2401, "status": "occupied"},
        {"unit_number": "NHJ LLC - Unit C", "monthly_rent": 2000, "status": "occupied"},
        {"unit_number": "JKL LLC - Unit D", "monthly_rent": 6500, "status": "occupied"},
        {"unit_number": "LOP LLC - Unit E", "monthly_rent": 2791, "status": "occupied"},
        {"unit_number": "LOP LLC - Unit F", "monthly_rent": 1750, "status": "occupied"},
        {"unit_number": "LOP LLC - Unit G", "monthly_rent": 1875, "status": "occupied"},
        {"unit_number": "LOP LLC - Unit H", "monthly_rent": 4000, "status": "occupied"},
        {"unit_number": "LPO LLC - Unit I", "monthly_rent": 7242, "status": "occupied"},
        {"unit_number": "LPO LLC - Unit J", "monthly_rent": 3500, "status": "occupied"},
        {"unit_number": "PPP LLC - Unit K", "monthly_rent": 3200, "status": "occupied"},
        {"unit_number": "ABC LLC - Unit L", "monthly_rent": 0,    "status": "vacant"},
    ]}]},
    {"company": "BNC LLC", "suites": [{"name": "BNC LLC", "units": [
        {"unit_number": "Unit A",     "monthly_rent": 800,  "status": "occupied"},
        {"unit_number": "Unit B,C",   "monthly_rent": 0,    "status": "vacant"},
        {"unit_number": "Unit D",     "monthly_rent": 730,  "status": "occupied"},
        {"unit_number": "Unit E & F", "monthly_rent": 1575, "status": "occupied"},
        {"unit_number": "Unit G",     "monthly_rent": 875,  "status": "occupied"},
        {"unit_number": "Unit H",     "monthly_rent": 800,  "status": "occupied"},
        {"unit_number": "Unit I",     "monthly_rent": 830,  "status": "occupied"},
        {"unit_number": "Unit J",     "monthly_rent": 825,  "status": "occupied"},
        {"unit_number": "Unit K&L",   "monthly_rent": 1150, "status": "vacant"},
        {"unit_number": "Unit M",     "monthly_rent": 950,  "status": "occupied"},
    ]}]},
    {"company": "DEC LLC", "suites": [{"name": "DEC LLC", "units": [
        {"unit_number": "Unit A", "monthly_rent": 800,  "status": "occupied"},
        {"unit_number": "Unit B", "monthly_rent": 0,    "status": "vacant"},
        {"unit_number": "Unit C", "monthly_rent": 900,  "status": "occupied"},
        {"unit_number": "Unit D", "monthly_rent": 830,  "status": "occupied"},
        {"unit_number": "Unit E", "monthly_rent": 850,  "status": "occupied"},
        {"unit_number": "Unit F", "monthly_rent": 850,  "status": "occupied"},
        {"unit_number": "Unit G", "monthly_rent": 800,  "status": "occupied"},
        {"unit_number": "Unit H", "monthly_rent": 900,  "status": "occupied"},
        {"unit_number": "Unit I", "monthly_rent": 800,  "status": "occupied"},
        {"unit_number": "Unit J", "monthly_rent": 1254, "status": "vacant"},
        {"unit_number": "Unit K", "monthly_rent": 1950, "status": "occupied"},
        {"unit_number": "Unit L", "monthly_rent": 0,    "status": "vacant"},
        {"unit_number": "Unit M", "monthly_rent": 2129, "status": "vacant"},
        {"unit_number": "Unit N", "monthly_rent": 0,    "status": "vacant"},
        {"unit_number": "Unit O", "monthly_rent": 2150, "status": "occupied"},
        {"unit_number": "Unit P", "monthly_rent": 0,    "status": "vacant"},
        {"unit_number": "Unit Q", "monthly_rent": 1860, "status": "occupied"},
        {"unit_number": "Unit R", "monthly_rent": 0,    "status": "vacant"},
        {"unit_number": "Unit S", "monthly_rent": 950,  "status": "occupied"},
    ]}]},
    {"company": "XYZ LLC", "suites": [{"name": "XYZ LLC", "units": [
        {"unit_number": "Unit A", "monthly_rent": 600, "status": "occupied"},
        {"unit_number": "Unit B", "monthly_rent": 800, "status": "occupied"},
        {"unit_number": "Unit C", "monthly_rent": 900, "status": "occupied"},
        {"unit_number": "Unit D", "monthly_rent": 800, "status": "occupied"},
        {"unit_number": "Unit E", "monthly_rent": 800, "status": "occupied"},
        {"unit_number": "Unit F", "monthly_rent": 800, "status": "occupied"},
    ]}]},
    {"company": "ZYC LLC", "suites": [{"name": "ZYC LLC", "units": [
        {"unit_number": "Unit A", "monthly_rent": 900, "status": "occupied"},
        {"unit_number": "Unit B", "monthly_rent": 875, "status": "occupied"},
        {"unit_number": "Unit C", "monthly_rent": 825, "status": "occupied"},
        {"unit_number": "Unit D", "monthly_rent": 950, "status": "occupied"},
        {"unit_number": "Unit E", "monthly_rent": 825, "status": "occupied"},
        {"unit_number": "Unit F", "monthly_rent": 750, "status": "occupied"},
        {"unit_number": "Unit G", "monthly_rent": 900, "status": "occupied"},
        {"unit_number": "Unit H", "monthly_rent": 925, "status": "occupied"},
        {"unit_number": "Unit I", "monthly_rent": 800, "status": "occupied"},
        {"unit_number": "Unit J", "monthly_rent": 800, "status": "occupied"},
        {"unit_number": "Unit K", "monthly_rent": 750, "status": "occupied"},
        {"unit_number": "Unit L", "monthly_rent": 825, "status": "occupied"},
        {"unit_number": "Unit M", "monthly_rent": 900, "status": "occupied"},
        {"unit_number": "Unit N", "monthly_rent": 850, "status": "occupied"},
        {"unit_number": "Unit O", "monthly_rent": 0,   "status": "vacant"},
        {"unit_number": "Unit P", "monthly_rent": 800, "status": "occupied"},
        {"unit_number": "Unit Q", "monthly_rent": 850, "status": "occupied"},
        {"unit_number": "Unit R", "monthly_rent": 850, "status": "occupied"},
        {"unit_number": "Unit S", "monthly_rent": 800, "status": "occupied"},
        {"unit_number": "Unit T", "monthly_rent": 825, "status": "occupied"},
    ]}]},
    {"company": "ACD LLC", "suites": [{"name": "ACD LLC", "units": [
        {"unit_number": "Unit A", "monthly_rent": 1100, "status": "occupied"},
        {"unit_number": "Unit B", "monthly_rent": 830,  "status": "occupied"},
        {"unit_number": "Unit C", "monthly_rent": 800,  "status": "vacant"},
        {"unit_number": "Unit D", "monthly_rent": 900,  "status": "occupied"},
        {"unit_number": "Unit E", "monthly_rent": 775,  "status": "occupied"},
        {"unit_number": "Unit F", "monthly_rent": 650,  "status": "occupied"},
        {"unit_number": "Unit G", "monthly_rent": 800,  "status": "occupied"},
        {"unit_number": "Unit H", "monthly_rent": 775,  "status": "occupied"},
        {"unit_number": "Unit I", "monthly_rent": 800,  "status": "occupied"},
        {"unit_number": "Unit J", "monthly_rent": 650,  "status": "vacant"},
        {"unit_number": "Unit K", "monthly_rent": 850,  "status": "occupied"},
        {"unit_number": "Unit L", "monthly_rent": 1400, "status": "occupied"},
        {"unit_number": "Unit M", "monthly_rent": 0,    "status": "vacant"},
        {"unit_number": "Unit N", "monthly_rent": 900,  "status": "occupied"},
    ]}]},
    {"company": "NHJ LLC", "suites": [{"name": "NHJ LLC", "units": [
        {"unit_number": "Unit A,B,C,G", "monthly_rent": 2700, "status": "occupied"},
        {"unit_number": "Unit D",       "monthly_rent": 850,  "status": "occupied"},
        {"unit_number": "Unit E",       "monthly_rent": 800,  "status": "occupied"},
        {"unit_number": "Unit F",       "monthly_rent": 880,  "status": "occupied"},
        {"unit_number": "Unit H",       "monthly_rent": 800,  "status": "occupied"},
    ]}]},
    {"company": "FJH LLC", "suites": [{"name": "FJH LLC", "units": [
        {"unit_number": "REAR unit", "monthly_rent": 2850, "status": "occupied"},
        {"unit_number": "Unit A",    "monthly_rent": 400,  "status": "vacant"},
        {"unit_number": "Unit B",    "monthly_rent": 1400, "status": "occupied"},
        {"unit_number": "Unit C",    "monthly_rent": 0,    "status": "vacant"},
        {"unit_number": "Unit D",    "monthly_rent": 825,  "status": "occupied"},
        {"unit_number": "Unit E",    "monthly_rent": 1630, "status": "occupied"},
        {"unit_number": "Unit F",    "monthly_rent": 0,    "status": "vacant"},
        {"unit_number": "Unit G",    "monthly_rent": 800,  "status": "occupied"},
    ]}]},
    {"company": "KLI LLC", "suites": [{"name": "KLI LLC", "units": [
        {"unit_number": "Unit A", "monthly_rent": 800,  "status": "occupied"},
        {"unit_number": "Unit B", "monthly_rent": 800,  "status": "occupied"},
        {"unit_number": "Unit C", "monthly_rent": 1000, "status": "occupied"},
        {"unit_number": "Unit D", "monthly_rent": 850,  "status": "occupied"},
        {"unit_number": "Unit E", "monthly_rent": 800,  "status": "occupied"},
        {"unit_number": "Unit F", "monthly_rent": 725,  "status": "occupied"},
        {"unit_number": "Unit G", "monthly_rent": 0,    "status": "vacant"},
        {"unit_number": "Unit H", "monthly_rent": 825,  "status": "occupied"},
        {"unit_number": "Unit I", "monthly_rent": 900,  "status": "occupied"},
        {"unit_number": "Unit J", "monthly_rent": 775,  "status": "occupied"},
        {"unit_number": "Unit K", "monthly_rent": 800,  "status": "occupied"},
        {"unit_number": "Unit L", "monthly_rent": 850,  "status": "occupied"},
        {"unit_number": "Unit M", "monthly_rent": 0,    "status": "vacant"},
        {"unit_number": "Unit N", "monthly_rent": 1800, "status": "occupied"},
        {"unit_number": "Unit O", "monthly_rent": 2100, "status": "occupied"},
    ]}]},
]


@router.post("/seed-portfolio")
def seed_portfolio(
    current_user: CurrentUser = Depends(require_write_access()),
    db: Session = Depends(get_db),
):
    """One-click seed: insert all 10 portfolio companies/suites/units for current tenant."""
    tid = current_user.tenant_id
    existing = {
        co.company_name.strip().lower()
        for co in db.query(RentalCompany).filter(RentalCompany.tenant_id == tid).all()
    }
    created_co = created_su = created_un = 0
    skipped: list[str] = []

    for entry in PORTFOLIO_DATA:
        co_name = entry["company"]
        if co_name.lower() in existing:
            skipped.append(co_name)
            continue
        co = RentalCompany(tenant_id=tid, company_name=co_name, created_by=current_user.email)
        db.add(co)
        db.flush()
        created_co += 1
        for suite_def in entry["suites"]:
            prop = RentalProp(tenant_id=tid, company_id=co.id, property_name=suite_def["name"])
            db.add(prop)
            db.flush()
            created_su += 1
            for u in suite_def["units"]:
                db.add(RentalUnit(
                    tenant_id=tid, property_id=prop.id, company_id=co.id,
                    unit_number=u["unit_number"], status=u["status"],
                    monthly_rent=float(u["monthly_rent"]),
                ))
                created_un += 1

    db.commit()
    return {
        "status": "success",
        "created_companies": created_co,
        "created_suites": created_su,
        "created_units": created_un,
        "skipped": skipped,
        "message": (
            f"Seeded {created_co} companies, {created_su} suites, "
            f"{created_un} units. Skipped {len(skipped)} existing."
        ),
    }


# ── financial statements (P&L / BS / CF) ─────────────────────────────────────

def _ensure_fin_uploads_table(engine) -> None:
    from models.rentals.models import RentalFinancialUpload
    RentalFinancialUpload.__table__.create(bind=engine, checkfirst=True)
    # Add periods column if it doesn't exist yet (safe migration)
    try:
        with engine.connect() as conn:
            conn.execute(
                __import__('sqlalchemy').text(
                    "ALTER TABLE r_financial_uploads ADD COLUMN IF NOT EXISTS periods JSON"
                )
            )
            conn.commit()
    except Exception:
        pass


@router.get("/financials")
def list_financials(
    current_user: CurrentUser = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    from database import engine as _engine
    _ensure_fin_uploads_table(_engine)
    from models.rentals.models import RentalFinancialUpload
    rows = db.query(RentalFinancialUpload).filter(
        RentalFinancialUpload.tenant_id == current_user.tenant_id
    ).all()
    return [
        {
            "company_id": str(r.company_id),
            "company_name": r.company_name,
            "filename": r.filename,
            "years": r.years or [],
            "uploaded_at": r.uploaded_at.isoformat() if r.uploaded_at else None,
        }
        for r in rows
    ]


@router.get("/financials/{company_id}")
def get_financials(
    company_id: str,
    current_user: CurrentUser = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    from database import engine as _engine
    _ensure_fin_uploads_table(_engine)
    from models.rentals.models import RentalFinancialUpload
    row = db.query(RentalFinancialUpload).filter(
        RentalFinancialUpload.tenant_id == current_user.tenant_id,
        RentalFinancialUpload.company_id == uuid.UUID(company_id),
    ).first()
    if not row:
        raise HTTPException(status_code=404, detail="No financials for this company")
    return {
        "company_id": str(row.company_id),
        "company_name": row.company_name,
        "filename": row.filename,
        "date_range": row.date_range,
        "years": row.years or [],
        "periods": row.periods or [],
        "pl": row.pl_data or [],
        "bs": row.bs_data or [],
        "cf": row.cf_data or [],
        "uploaded_at": row.uploaded_at.isoformat() if row.uploaded_at else None,
    }


@router.post("/financials/save")
def save_financials(
    body: dict,
    current_user: CurrentUser = Depends(require_write_access()),
    db: Session = Depends(get_db),
):
    from database import engine as _engine
    _ensure_fin_uploads_table(_engine)
    from models.rentals.models import RentalFinancialUpload

    company_id = uuid.UUID(body["company_id"])
    existing = db.query(RentalFinancialUpload).filter(
        RentalFinancialUpload.tenant_id == current_user.tenant_id,
        RentalFinancialUpload.company_id == company_id,
    ).first()

    if existing:
        existing.company_name = body.get("company_name", existing.company_name)
        existing.filename = body.get("filename", existing.filename)
        existing.date_range = body.get("date_range", existing.date_range)
        existing.years = body.get("years", existing.years)
        existing.periods = body.get("periods", existing.periods)
        existing.pl_data = body.get("pl", existing.pl_data)
        existing.bs_data = body.get("bs", existing.bs_data)
        existing.cf_data = body.get("cf", existing.cf_data)
        existing.uploaded_by = current_user.email
    else:
        row = RentalFinancialUpload(
            tenant_id=current_user.tenant_id,
            company_id=company_id,
            company_name=body.get("company_name", ""),
            filename=body.get("filename", ""),
            date_range=body.get("date_range", ""),
            years=body.get("years", []),
            periods=body.get("periods", []),
            pl_data=body.get("pl", []),
            bs_data=body.get("bs", []),
            cf_data=body.get("cf", []),
            uploaded_by=current_user.email,
        )
        db.add(row)

    db.commit()
    return {
        "status": "saved",
        "years": body.get("years", []),
        "pl_rows": len(body.get("pl", [])),
        "bs_rows": len(body.get("bs", [])),
        "cf_rows": len(body.get("cf", [])),
    }


@router.delete("/financials/{company_id}", status_code=204)
def delete_financials(
    company_id: str,
    current_user: CurrentUser = Depends(require_write_access()),
    db: Session = Depends(get_db),
):
    from database import engine as _engine
    _ensure_fin_uploads_table(_engine)
    from models.rentals.models import RentalFinancialUpload
    row = db.query(RentalFinancialUpload).filter(
        RentalFinancialUpload.tenant_id == current_user.tenant_id,
        RentalFinancialUpload.company_id == uuid.UUID(company_id),
    ).first()
    if row:
        db.delete(row)
        db.commit()


# ── ar aging detail (month-by-month per unit) ────────────────────────────────

@router.get("/ar-aging-detail")
def get_ar_aging_detail(
    company_id: str = Query(None),
    current_user: CurrentUser = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Get AR aging by bucket with per-unit, per-month breakdown."""
    tid = current_user.tenant_id
    today = date.today()

    query = db.query(RentalUnit).filter(RentalUnit.tenant_id == tid)
    if company_id:
        try:
            cid = uuid.UUID(company_id)
            query = query.filter(RentalUnit.company_id == cid)
        except ValueError:
            pass

    units = query.all()
    unit_aging_detail = []
    portfolio_buckets = {"current": 0.0, "1_30": 0.0, "31_60": 0.0, "61_90": 0.0, "90_plus": 0.0}

    for unit in units:
        invoices = db.query(RentalInvoice).filter(
            RentalInvoice.tenant_id == tid, RentalInvoice.unit_id == unit.id
        ).order_by(RentalInvoice.billing_period.desc()).all()

        unit_buckets = {"current": 0.0, "1_30": 0.0, "31_60": 0.0, "61_90": 0.0, "90_plus": 0.0}

        for inv in invoices:
            collected = sum(float(c.amount_collected) for c in inv.collections)
            owed = max(0.0, float(inv.amount_billed) - collected)
            if owed <= 0:
                continue

            due_date = inv.billing_period.replace(day=1)
            days_past_due = (today - due_date).days

            if days_past_due <= 0:
                bucket = "current"
            elif days_past_due <= 30:
                bucket = "1_30"
            elif days_past_due <= 60:
                bucket = "31_60"
            elif days_past_due <= 90:
                bucket = "61_90"
            else:
                bucket = "90_plus"

            unit_buckets[bucket] += owed
            portfolio_buckets[bucket] += owed

            unit_aging_detail.append({
                "unit_id": str(unit.id),
                "unit_number": unit.unit_number,
                "company_name": unit.company.company_name if unit.company else "",
                "billing_month": inv.billing_period.isoformat(),
                "amount_billed": float(inv.amount_billed),
                "amount_collected": collected,
                "owed": round(owed, 2),
                "days_past_due": days_past_due,
                "bucket": bucket,
            })

        if sum(unit_buckets.values()) > 0:
            unit_aging_detail.append({
                "unit_id": str(unit.id),
                "unit_number": unit.unit_number,
                "company_name": unit.company.company_name if unit.company else "",
                "billing_month": "UNIT_TOTAL",
                "owed": round(sum(unit_buckets.values()), 2),
                "buckets": {k: round(v, 2) for k, v in unit_buckets.items()},
            })

    return {
        "portfolio_buckets": {k: round(v, 2) for k, v in portfolio_buckets.items()},
        "total_ar": round(sum(portfolio_buckets.values()), 2),
        "unit_detail": unit_aging_detail,
        "generated_at": today.isoformat(),
    }


# ── AR summary — registry-driven billed vs collected ─────────────────────────

@router.get("/ar-summary")
def get_ar_summary(
    month: str = Query(None),        # "Jan-2026" — filter to a specific month
    company_id: str = Query(None),   # UUID — filter to one company
    current_user: CurrentUser = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """
    AR summary per company per month.

    Billed     = SUM(r_units.monthly_rent) for occupied units only (registry).
    Collected  = PRIMARY: r_companies.monthly_rent_data (from Rent Receivable upload).
                 FALLBACK: r_financial_uploads.pl_data monthlyValues on "Rent -" lines.
    Vacancy    = SUM(r_units.monthly_rent) for vacant/notice/other units.
    AR         = Billed - Collected per company per month.
    Unmatched  = P&L "Rent -" lines whose unit label doesn't match any registry unit.
    """
    import re
    from models.rentals.models import RentalFinancialUpload

    tid = current_user.tenant_id
    companies = db.query(RentalCompany).filter(RentalCompany.tenant_id == tid).all()
    # company_id filter is handled client-side so dropdowns stay populated

    all_units = db.query(RentalUnit).filter(RentalUnit.tenant_id == tid).all()
    units_by_co: dict[str, list] = defaultdict(list)
    for u in all_units:
        units_by_co[str(u.company_id)].append(u)

    all_uploads = db.query(RentalFinancialUpload).filter(RentalFinancialUpload.tenant_id == tid).all()
    uploads_by_co: dict[str, object] = {}
    for up in all_uploads:
        cid = str(up.company_id)
        if cid not in uploads_by_co or up.uploaded_at > uploads_by_co[cid].uploaded_at:
            uploads_by_co[cid] = up

    # Only match individual rent lines (e.g. "Rent - Unit 3A").
    # Deliberately excludes "Total for Rental Income" / "Total for Services" —
    # those are section-total aggregates and would double-count when individual
    # unit lines are also present in the same P&L.
    RENT_RE = re.compile(r'^rent\s*[-–]', re.IGNORECASE)
    # Validates a normalised month key — must be "Mon-YYYY" (3-letter abbrev + 4-digit year).
    # Rejects "Total", "YTD", "Budget", "Prior Year" and other non-month P&L columns that
    # would otherwise be summed into src_b and produce a phantom spike month.
    MONTH_KEY_RE = re.compile(r'^[A-Za-z]{3}-\d{4}$')

    def norm_month(m: str) -> str:
        # "Jan 2026" → "Jan-2026"; "Jan-2026" → "Jan-2026"
        return m.replace(' ', '-') if ' ' in m else m

    def valid_month(mk: str) -> bool:
        return bool(MONTH_KEY_RE.match(mk))

    def month_sort_key(m: str) -> tuple:
        MNAMES = ['Jan','Feb','Mar','Apr','May','Jun','Jul','Aug','Sep','Oct','Nov','Dec']
        parts = m.replace('-', ' ').split()
        if len(parts) == 2:
            try:
                return (int(parts[1]), MNAMES.index(parts[0]))
            except (ValueError, IndexError):
                pass
        return (9999, 0)

    company_summaries = []
    all_months_set: set[str] = set()

    for co in companies:
        cid = str(co.id)
        units = units_by_co.get(cid, [])
        occupied  = [u for u in units if u.status.value == 'occupied']
        non_occ   = [u for u in units if u.status.value != 'occupied']

        billed_mo      = sum(float(u.monthly_rent) for u in occupied)
        vacancy_loss_mo = sum(float(u.monthly_rent) for u in non_occ)

        # Source A: Rent Receivable sync — only include months with actual data (> 0)
        src_a: dict[str, float] = {}
        if co.monthly_rent_data:
            for k, v in co.monthly_rent_data.items():
                val = float(v or 0)
                mk = norm_month(k)
                if val > 0 and valid_month(mk):
                    src_a[mk] = val

        # Source B: P&L financials fallback
        src_b: dict[str, float] = {}
        pl_unmatched: list[str] = []
        registry_norm = {u.unit_number.strip().lower() for u in units}

        upload = uploads_by_co.get(cid)
        if upload and upload.pl_data:
            for item in upload.pl_data:
                label = str(item.get('label', ''))
                if not RENT_RE.match(label):
                    continue
                if item.get('isSectionHeader') or item.get('isTotal') or item.get('children'):
                    continue
                mv: dict = item.get('monthlyValues') or {}
                for raw_k, v in mv.items():
                    mk = norm_month(raw_k)
                    if not valid_month(mk):   # drop "Total", "YTD", "Budget" etc.
                        continue
                    src_b[mk] = src_b.get(mk, 0.0) + float(v or 0)

                # Flag unmatched unit labels
                unit_m = re.search(r'(Unit\s+\S.*?)$', label, re.IGNORECASE)
                if unit_m:
                    extracted = unit_m.group(1).strip().lower()
                    matched = any(
                        extracted == reg or extracted in reg or reg in extracted
                        for reg in registry_norm
                    )
                    if not matched:
                        pl_unmatched.append(label)

        # Build per-month detail — union of both sources
        all_months = sorted(set(list(src_a.keys()) + list(src_b.keys())), key=month_sort_key)
        all_months_set.update(all_months)

        monthly_detail = []
        for m in all_months:
            has_a, has_b = m in src_a, m in src_b

            if has_a:
                collected   = src_a[m]
                data_source = 'rent_receivable'
            elif has_b:
                collected   = src_b[m]
                data_source = 'pl_fallback'
            else:
                continue

            outstanding = max(0.0, billed_mo - collected)
            rate = (collected / billed_mo * 100) if billed_mo > 0 else 0.0

            recon_flag = None
            if has_a and has_b and billed_mo > 0:
                diff_pct = abs(src_a[m] - src_b[m]) / billed_mo * 100
                if diff_pct > 2:
                    recon_flag = {
                        'rent_receivable': round(src_a[m], 2),
                        'pl': round(src_b[m], 2),
                        'diff_pct': round(diff_pct, 1),
                    }

            monthly_detail.append({
                'month': m,
                'billed': round(billed_mo, 2),
                'collected': round(collected, 2),
                'outstanding': round(outstanding, 2),
                'collection_rate': round(rate, 1),
                'data_source': data_source,
                'recon_flag': recon_flag,
            })

        # latest = last month that has actual collected data (> 0)
        # If month filter is active, use that specific month's row instead
        if month:
            latest = next((d for d in reversed(monthly_detail) if d['month'] == month), None)
        else:
            with_data = [d for d in monthly_detail if d['collected'] > 0]
            latest = with_data[-1] if with_data else (monthly_detail[-1] if monthly_detail else None)

        company_summaries.append({
            'company_id': cid,
            'company_name': co.company_name,
            'total_units': len(units),
            'occupied_units': len(occupied),
            'vacant_units': len(non_occ),
            'billed_per_month': round(billed_mo, 2),
            'vacancy_loss_per_month': round(vacancy_loss_mo, 2),
            'last_sync_month': co.last_sync_month,
            'has_rent_receivable': bool(src_a),
            'has_pl_data': bool(src_b),
            'monthly': monthly_detail,
            'latest_month': latest['month'] if latest else None,
            'latest_collected': latest['collected'] if latest else 0.0,
            'latest_outstanding': latest['outstanding'] if latest else 0.0,
            'latest_rate': latest['collection_rate'] if latest else 0.0,
            'pl_lines_unmatched': pl_unmatched,
        })

    # Portfolio totals — use month-filtered latest per company
    total_billed      = sum(c['billed_per_month'] for c in company_summaries)
    total_collected   = sum(c['latest_collected'] for c in company_summaries)
    total_outstanding = max(0.0, total_billed - total_collected)
    port_rate         = (total_collected / total_billed * 100) if total_billed > 0 else 0.0
    total_vac_loss    = sum(c['vacancy_loss_per_month'] for c in company_summaries)
    total_occupied    = sum(c['occupied_units'] for c in company_summaries)
    total_units_all   = sum(c['total_units'] for c in company_summaries)

    # Monthly trend — only months where at least one company has collected > 0
    m_billed: dict[str, float] = {}
    m_collected: dict[str, float] = {}
    for cs in company_summaries:
        for md in cs['monthly']:
            mk = md['month']
            m_billed[mk]    = m_billed.get(mk, 0.0)    + md['billed']
            m_collected[mk] = m_collected.get(mk, 0.0) + md['collected']

    # Only include months that have at least some collected data across portfolio
    months_with_data = {mk for mk, v in m_collected.items() if v > 0}
    monthly_trend = [
        {'month': m, 'billed': round(m_billed[m], 2), 'collected': round(m_collected[m], 2)}
        for m in sorted(months_with_data, key=month_sort_key)
    ]

    # Available months list (for frontend filter dropdown)
    all_available_months = sorted(all_months_set, key=month_sort_key)

    all_unmatched = [
        {'company': c['company_name'], 'label': lbl}
        for c in company_summaries
        for lbl in c['pl_lines_unmatched']
    ]

    return {
        'companies': company_summaries,
        'portfolio': {
            'total_billed':      round(total_billed, 2),
            'total_collected':   round(total_collected, 2),
            'total_outstanding': round(total_outstanding, 2),
            'collection_rate':   round(port_rate, 1),
            'vacancy_loss':      round(total_vac_loss, 2),
            'occupied_units':    total_occupied,
            'total_units':       total_units_all,
        },
        'monthly_trend':      monthly_trend,
        'available_months':   all_available_months,
        'unmatched_lines':    all_unmatched,
        'generated_at':       datetime.utcnow().isoformat(),
    }

