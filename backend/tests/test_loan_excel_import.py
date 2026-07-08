"""Tests for loan_excel_import — synthetic data only."""
import io

import openpyxl

from services.loan_excel_import import parse_loan_workbook


def _workbook_bytes(sheets: dict[str, list[list]]) -> bytes:
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    for title, rows in sheets.items():
        ws = wb.create_sheet(title)
        for r_idx, row in enumerate(rows, start=1):
            for c_idx, val in enumerate(row, start=1):
                ws.cell(row=r_idx, column=c_idx, value=val)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_header_based_import():
    rows = [
        ["Sl No.", "Company Name", "Property Name", "Loan Bank Name", "Loan Date",
         "Loan Amount", "Loan Interest Rate", "Loan EMI", "Lender Name",
         "Loan Maturity Date", "Loan Balance", "Loan EMI Day"],
        [1, "Entity Alpha LLC", "Building One", "First Bank", "01-15-2022",
         500000, "5.25%", 3200, "Jane Doe", "01-15-2032", 480000, 15],
    ]
    parsed = parse_loan_workbook(_workbook_bytes({"Loans": rows}))
    assert len(parsed.rows) == 1
    assert parsed.rows[0].company == "Entity Alpha LLC"
    assert parsed.rows[0].property_name == "Building One"
    assert parsed.rows[0].loan_amount == 500000
    assert parsed.rows[0].loan_interest_rate == 0.0525
    assert parsed.rows[0].loan_emi == 3200


def test_entity_name_and_rental_filter():
    rows = [
        ["Entity", "Entity Name", "Property Name", "Loan Bank Name", "Loan Amount",
         "Loan Interest Rate", "Loan EMI", "Loan Maurity Date", "Loan Balance"],
        ["Rental", "Alpha LLC", "Suite 100", "First Bank", 400000, 5.5, 2800, "01-01-2030", 390000],
        ["Construction", "Beta LLC", "Site A", "Second Bank", 900000, 6.0, 5000, "01-01-2031", 850000],
    ]
    parsed = parse_loan_workbook(_workbook_bytes({"Bank Loan Information": rows}))
    assert len(parsed.rows) == 1
    assert parsed.rows[0].company == "Alpha LLC"
    assert parsed.skipped_non_rental == 1
    assert parsed.has_entity_line_column is True


def test_monthly_balance_columns():
    rows = [
        ["Company Name", "Property Name", "Loan Bank Name", "Loan Amount",
         "Loan Interest Rate", "Loan EMI", "Mar-2026", "Apr-2026"],
        ["Gamma LLC", "Building Two", "Metro Bank", 750000, 6.0, 4500, 700000, 695000],
    ]
    parsed = parse_loan_workbook(_workbook_bytes({"Bank Loan Information": rows}))
    assert len(parsed.rows) == 1
    assert parsed.rows[0].balance_by_month["2026-03"] == 700000
    assert parsed.rows[0].balance_by_month["2026-04"] == 695000
    assert parsed.balance_periods == ["2026-03", "2026-04"]


def test_wwbg_style_typo_headers():
    rows = [
        ["Company Name", "Property Name", "Loan Bank Name", "Loan Acc No", "Loan Amount",
         "Loan interest Rate", "Loan EMI ", "Loan Date", "Loan Maurity Date"],
        ["Delta LLC", "Park Plaza", "Wells Fargo", "12345", 1200000, 4.75, 8500, "03-01-2020", "03-01-2030"],
    ]
    parsed = parse_loan_workbook(_workbook_bytes({"Bank Loan Information": rows}))
    assert len(parsed.rows) == 1
    assert parsed.rows[0].account_no == "12345"
    assert parsed.rows[0].loan_interest_rate == 0.0475
    assert parsed.rows[0].loan_emi == 8500


def test_sheet_title_used_when_company_column_empty():
    rows = [
        ["Sl No.", "Company Name", "Property Name", "Loan Bank Name", "Loan Date",
         "Loan Amount", "Loan Interest Rate", "Loan EMI", "Lender Name",
         "Loan Maturity Date", "Loan Balance", "Loan EMI Day"],
        [1, "", "Building Two", "Second Bank", "06-01-2021",
         750000, 6.0, 4500, "Alex Kim", "06-01-2031", 700000, 1],
    ]
    parsed = parse_loan_workbook(_workbook_bytes({"Gamma Holdings LLC": rows}))
    assert len(parsed.rows) == 1
    assert parsed.rows[0].company == "Gamma Holdings LLC"
    assert parsed.rows[0].property_name == "Building Two"
