"""Tests for ownership_excel_import — synthetic data only."""
import io

import openpyxl

from services.ownership_excel_import import (
    build_import_template_bytes,
    is_rental_entity_line,
    parse_ownership_workbook,
)


def _workbook_bytes(sheets: dict[str, list[list]]) -> bytes:
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    for title, rows in sheets.items():
        ws = wb.create_sheet(title)
        for r_idx, row in enumerate(rows, start=1):
            for c_idx, val in enumerate(row, start=1):
                ws.cell(row=r_idx, column=c_idx, value=val)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_parse_ownership_template_columns():
    rows = [
        [
            "Entity Name", "Owned By", "Property Address", "Property Name",
            "Ownership %", "Entity Structure", "Cost Basis", "Book Value", "Existing Debt",
        ],
        [
            "Alpha LLC", "Partner A", "123 Main St", "Building One",
            0.25, "LLC", 500000, 480000, 300000,
        ],
        [
            "Alpha LLC", "Partner B", "123 Main St", "Building One",
            "75%", "LP", 1500000, 1440000, 900000,
        ],
    ]
    result = parse_ownership_workbook(_workbook_bytes({"Ownership": rows}))
    assert len(result.rows) == 2
    assert result.rows[0].entity_name == "Alpha LLC"
    assert result.rows[0].partner_name == "Partner A"
    assert result.rows[0].property_name == "Building One"
    assert result.rows[0].ownership_pct == 0.25
    assert result.rows[0].cost_basis == 500000
    assert result.rows[1].ownership_pct == 0.75


def test_filters_non_rental_entity_column():
    rows = [
        [
            "Entity", "Entity Name", "Owned By", "Property Name",
            "Ownership %", "Cost Basis", "Book Value",
        ],
        ["Rental", "Alpha LLC", "Partner A", "Building One", 0.5, 100000, 90000],
        ["Construction", "Build Co", "Partner B", "Site A", 0.5, 200000, 180000],
        ["Land", "Land Holdco", "Partner D", "Parcel 9", 1.0, 500000, 480000],
        ["Consulting", "Ack Co", "Partner E", "Office", 1.0, 10000, 10000],
        ["Rental", "Beta LLC", "Partner C", "Suite 2", 0.25, 50000, None],
    ]
    result = parse_ownership_workbook(_workbook_bytes({"Ownership": rows}))
    assert result.has_entity_line_column is True
    assert result.skipped_non_rental == 2
    assert len(result.rows) == 3
    assert {r.entity_name for r in result.rows} == {"Alpha LLC", "Land Holdco", "Beta LLC"}
    assert {r.entity_line for r in result.rows} == {"Rental", "Land"}
    # Blank Book Value on Beta carries prior kept (Land) value — skipped Consulting BV ignored.
    beta = next(r for r in result.rows if r.entity_name == "Beta LLC")
    assert beta.book_value == 480000
    alpha = next(r for r in result.rows if r.entity_name == "Alpha LLC")
    assert alpha.book_value == 90000
    land = next(r for r in result.rows if r.entity_name == "Land Holdco")
    assert land.entity_line == "Land"
    assert land.book_value == 480000


def test_is_rental_entity_line():
    assert is_rental_entity_line("Rental") is True
    assert is_rental_entity_line("RENTAL") is True
    assert is_rental_entity_line("Construction") is False
    assert is_rental_entity_line("Development") is False
    assert is_rental_entity_line(None) is True


def test_is_ownership_entity_line_includes_land():
    from services.ownership_excel_import import is_ownership_entity_line, normalize_entity_line
    assert is_ownership_entity_line("Rental") is True
    assert is_ownership_entity_line("Land") is True
    assert is_ownership_entity_line("Consulting") is False
    assert is_ownership_entity_line("Partner") is False
    assert normalize_entity_line("land") == "Land"
    assert normalize_entity_line("RENTAL") == "Rental"


def test_template_builder_has_expected_headers():
    content = build_import_template_bytes()
    wb = openpyxl.load_workbook(io.BytesIO(content))
    ws = wb.active
    headers = [ws.cell(row=1, column=c).value for c in range(1, 11)]
    assert headers[0] == "Entity"
    assert headers[1] == "Entity Name"
    assert headers[2] == "Owned By"
    assert headers[4] == "Property Name"
    assert headers[9] == "Existing Debt"
    assert ws.cell(row=2, column=1).value == "Rental"


def test_skips_rows_without_partner_or_entity():
    rows = [
        ["Entity Name", "Owned By", "Property Name", "Ownership %", "Cost Basis"],
        ["", "Partner A", "Building One", 0.5, 100000],
        ["Alpha LLC", "", "Building One", 0.5, 100000],
        ["TOTAL", "", "", "", ""],
    ]
    result = parse_ownership_workbook(_workbook_bytes({"Ownership": rows}))
    assert len(result.rows) == 0


def test_legacy_entity_column_as_company_name():
    """Sheets with only 'Entity' (no Entity Name) still map Entity → company."""
    rows = [
        ["Entity", "Owned By", "Property Name", "Ownership %", "Cost Basis"],
        ["Alpha LLC", "Partner A", "Building One", 0.5, 100000],
    ]
    result = parse_ownership_workbook(_workbook_bytes({"Ownership": rows}))
    assert len(result.rows) == 1
    assert result.rows[0].entity_name == "Alpha LLC"
